"""Генератор итогового отчета с цветным визуальным представлением успешности проверок"""

import os
import sys
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import matplotlib.colors as mcolors
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import io

# Добавляем путь для импорта
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

try:
    from utils.symbols import EXCEL_SYMBOLS, TERMINAL_SYMBOLS
except ImportError:
    EXCEL_SYMBOLS = {
        'SUCCESS': '✅', 'ERROR': '❌', 'WARNING': '⚠️',
        'INFO': 'ℹ️', 'ROCKET': '🚀', 'GEAR': '⚙️',
        'MAGNIFYING_GLASS': '🔍', 'CHECKMARK': '✓',
        'CHART_UP': '📈', 'CHART_DOWN': '📉',
        'FILE_FOLDER': '📁', 'PAGE': '📄', 'TABLE': '📊',
        'BOOKS': '📚', 'SAVE': '💾', 'BAR_CHART': '📊',
        'CELEBRATION': '🎉', 'MEMO': '📝', 'TARGET': '🎯',
        'CLOCK': '⏱️', 'SIREN': '🚨', 'TROPHY': '🏆',
        'STAR': '⭐', 'FIRE': '🔥', 'SNOWFLAKE': '❄️',
        'CHECKERED_FLAG': '🏁'
    }
    
    TERMINAL_SYMBOLS = {
        'SUCCESS': '[OK]', 'ERROR': '[ERROR]', 'WARNING': '[WARN]',
        'INFO': '[INFO]', 'ROCKET': '[START]', 'GEAR': '[PROC]',
        'MAGNIFYING_GLASS': '[CHECK]', 'CHECKMARK': '[DONE]',
        'CELEBRATION': '[DONE]', 'MEMO': '[NOTE]', 'TARGET': '[TARGET]',
        'CLOCK': '[TIME]', 'SIREN': '[ALERT]', 'TROPHY': '[BEST]',
        'STAR': '[STAR]', 'FIRE': '[HOT]', 'SNOWFLAKE': '[COLD]',
        'CHECKERED_FLAG': '[FINISH]'
    }

class FinalReportGenerator:
    """Генератор итогового отчета с цветным визуальным представлением"""
    
    def __init__(self, output_dir):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
    def generate_executive_summary(self, results_df):
        """Создает исполняемый итоговый отчет с цветным оформлением"""
        if results_df.empty:
            print(f"{TERMINAL_SYMBOLS['ERROR']} Нет данных для отчета")
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(self.output_dir, f"DQ_FINAL_REPORT_{timestamp}.xlsx")
        
        print(f"\n{TERMINAL_SYMBOLS['ROCKET']} СОЗДАНИЕ ИТОГОВОГО ОТЧЕТА")
        print("=" * 60)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Итоговый отчет"
        
        # Стили для цветового кодирования
        self._define_styles()
        
        # Генерируем отчет
        self._create_executive_summary_sheet(ws, results_df)
        self._create_dashboard_sheet(wb, results_df)
        self._create_rules_status_sheet(wb, results_df)
        self._create_category_summary_sheet(wb, results_df)
        
        # Сохраняем
        wb.save(report_path)
        
        print(f"{TERMINAL_SYMBOLS['CELEBRATION']} ИТОГОВЫЙ ОТЧЕТ СОХРАНЕН: {report_path}")
        
        # Также создаем HTML версию для удобного просмотра
        html_path = self._create_html_dashboard(results_df)
        
        return report_path
    
    def _define_styles(self):
        """Определяет стили для цветового кодирования"""
        self.styles = {
            # Цвета для статусов
            'STATUS_PASSED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'STATUS_FAILED': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'STATUS_WARNING': PatternFill(start_color="FFEB9C", end_color="FFEB9CE", fill_type="solid"),
            'STATUS_CRITICAL': PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
            
            # Цвета для качества (градиент от красного к зеленому)
            'QUALITY_0_50': PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
            'QUALITY_50_70': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'QUALITY_70_85': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'QUALITY_85_95': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'QUALITY_95_100': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            
            # Заголовки
            'HEADER_MAIN': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'HEADER_SECONDARY': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
            'HEADER_TERTIARY': PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
            
            # Границы
            'THIN_BORDER': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            
            # Шрифты
            'FONT_HEADER': Font(bold=True, size=14, color="FFFFFF"),
            'FONT_TITLE': Font(bold=True, size=12, color="1F4E78"),
            'FONT_SUBTITLE': Font(bold=True, size=11, color="2F5496"),
            'FONT_NORMAL': Font(size=10),
            'FONT_BOLD': Font(bold=True, size=10),
            'FONT_SMALL': Font(size=9),
        }
        
    def _create_executive_summary_sheet(self, ws, results_df):
        """Создает главный лист с итоговым отчетом"""
        print(f"   {TERMINAL_SYMBOLS['GEAR']} Создаем главный отчет...")
        
        # ===== ЗАГОЛОВОК =====
        ws.merge_cells('A1:K1')
        ws['A1'] = f"{EXCEL_SYMBOLS['TROPHY']} ИТОГОВЫЙ ОТЧЕТ ПО КАЧЕСТВУ ДАННЫХ"
        ws['A1'].font = Font(bold=True, size=18, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # ===== КЛЮЧЕВЫЕ МЕТРИКИ =====
        ws.merge_cells('A2:K2')
        ws['A2'] = f"Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Рассчитываем ключевые метрики
        total_rules = len(results_df)
        passed_rules = len(results_df[results_df['status'] == 'PASSED'])
        failed_rules = len(results_df[results_df['status'] == 'FAILED'])
        total_errors = results_df['failed'].sum()
        avg_quality = results_df['success_rate_%'].mean()
        total_records = results_df['total_records'].sum()
        
        # Статус системы (цветовой индикатор)
        if avg_quality >= 95:
            system_status = "ОТЛИЧНО"
            status_color = "92D050"
        elif avg_quality >= 85:
            system_status = "ХОРОШО"
            status_color = "C6EFCE"
        elif avg_quality >= 70:
            system_status = "УДОВЛЕТВОРИТЕЛЬНО"
            status_color = "FFEB9C"
        else:
            system_status = "ТРЕБУЕТ ВНИМАНИЯ"
            status_color = "FF9999"
        
        # ===== ПАНЕЛЬ КЛЮЧЕВЫХ ПОКАЗАТЕЛЕЙ =====
        # Заголовок панели
        ws['A4'] = f"{EXCEL_SYMBOLS['BAR_CHART']} КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ"
        ws['A4'].font = self.styles['FONT_TITLE']
        ws['A4'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.merge_cells('A4:C4')
        
        # Создаем таблицу с показателями
        indicators = [
            ["Общее качество данных", f"{avg_quality:.1f}%", system_status, status_color],
            ["Всего проверено правил", str(total_rules), f"{passed_rules} успешно", ""],
            ["Найдено ошибок", f"{total_errors:,}", f"{failed_rules} правил нарушено", ""],
            ["Обработано записей", f"{total_records:,}", "", ""],
            ["Среднее время проверки", f"{results_df['execution_time_sec'].mean():.2f}с", "", ""],
        ]
        
        # Заголовки таблицы
        headers = ["Показатель", "Значение", "Статус", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.styles['FONT_BOLD']
            cell.fill = self.styles['HEADER_TERTIARY']
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center')
        
        # Данные показателей
        for idx, (label, value, status, color) in enumerate(indicators, 1):
            row = 5 + idx
            
            # Показатель
            ws.cell(row=row, column=1, value=label).border = self.styles['THIN_BORDER']
            ws.cell(row=row, column=2, value=value).border = self.styles['THIN_BORDER']
            
            # Статус
            status_cell = ws.cell(row=row, column=3, value=status)
            status_cell.border = self.styles['THIN_BORDER']
            status_cell.alignment = Alignment(horizontal='center')
            
            # Цветовое кодирование
            if color:
                status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                status_cell.font = Font(bold=True, color="FFFFFF" if color == "FF9999" else "000000")
            elif "успешно" in status.lower():
                status_cell.fill = self.styles['STATUS_PASSED']
            elif "нарушено" in status.lower():
                status_cell.fill = self.styles['STATUS_FAILED']
        
        # ===== ВИЗУАЛЬНАЯ СВОДКА ПО ПРАВИЛАМ =====
        start_row = 12
        
        ws.cell(row=start_row, column=1, value=f"{EXCEL_SYMBOLS['MAGNIFYING_GLASS']} ВИЗУАЛЬНАЯ СВОДКА ПРОВЕРОК")
        ws.merge_cells(f'A{start_row}:L{start_row}')
        ws.cell(row=start_row, column=1).font = self.styles['FONT_TITLE']
        ws.cell(row=start_row, column=1).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Заголовки для сводки
        summary_headers = ["№", "Код правила", "Категория", "Таблица", "Колонка", 
                         "Оценено", "Успешно", "Ошибок", "% успеха", 
                         "Статус", "Время (с)"]
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.font = self.styles['FONT_BOLD']
            cell.fill = self.styles['HEADER_SECONDARY']
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные правил с цветовым кодированием
        for idx, (_, row) in enumerate(results_df.sort_values('success_rate_%').iterrows(), 1):
            data_row = start_row + 1 + idx
            
            # Подготовка данных
            row_data = [
                idx,
                row['rule_code'],
                row['quality_category'],
                row['table_name'],
                row.get('matched_column', row.get('column_checked', '')),
                row['total_records'],
                row['passed'],
                row['failed'],
                f"{row['success_rate_%']:.1f}%",
                row['status'],
                f"{row['execution_time_sec']:.2f}"
            ]
            
            # Заполняем строку
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = self.styles['THIN_BORDER']
                
                # Центрируем числовые значения
                if col in [6, 7, 8, 9, 11]:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Цветовое кодирование по проценту успеха
                if col == 10:  # Столбец % успеха
                    percent = float(str(value).replace('%', ''))
                    if percent >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif percent >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif percent >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    elif percent >= 50:
                        cell.fill = self.styles['QUALITY_50_70']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
                
                # Цветовое кодирование по статусу
                elif col == 11:  # Столбец статуса
                    if value == "PASSED":
                        cell.fill = self.styles['STATUS_PASSED']
                        cell.font = Font(bold=True, color="006100")
                    else:
                        cell.fill = self.styles['STATUS_FAILED']
                        cell.font = Font(bold=True, color="9C0006")
                
                # Выделение проблемных правил
                elif col == 2 and row['rule_code'] in ["RCCONF_12.2", "RCCONF_12.3"]:
                    cell.fill = self.styles['STATUS_CRITICAL']
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # ===== СВОДКА ПО КАТЕГОРИЯМ =====
        category_summary_row = start_row + len(results_df) + 3
        
        ws.cell(row=category_summary_row, column=1, 
                value=f"{EXCEL_SYMBOLS['CHECKERED_FLAG']} РЕЗУЛЬТАТЫ ПО КАТЕГОРИЯМ КАЧЕСТВА")
        ws.merge_cells(f'A{category_summary_row}:L{category_summary_row}')
        ws.cell(row=category_summary_row, column=1).font = self.styles['FONT_TITLE']
        ws.cell(row=category_summary_row, column=1).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Группируем по категориям
        categories = results_df.groupby('quality_category').agg({
            'success_rate_%': 'mean',
            'total_records': 'sum',
            'failed': 'sum',
            'rule_code': 'count'
        }).reset_index()
        
        categories.columns = ['Категория', 'Среднее качество', 'Оценено', 'Ошибок', 'Правил']
        
        # Заголовки категорий
        cat_headers = list(categories.columns)
        for col, header in enumerate(cat_headers, 1):
            cell = ws.cell(row=category_summary_row + 1, column=col, value=header)
            cell.font = self.styles['FONT_BOLD']
            cell.fill = self.styles['HEADER_TERTIARY']
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center')
        
        # Данные по категориям
        for idx, (_, cat_row) in enumerate(categories.iterrows(), 1):
            data_row = category_summary_row + 1 + idx
            
            row_data = [
                cat_row['Категория'],
                f"{cat_row['Среднее качество']:.1f}%",
                cat_row['Оценено'],
                cat_row['Ошибок'],
                cat_row['Правил']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = self.styles['THIN_BORDER']
                
                # Центрирование
                if col > 1:
                    cell.alignment = Alignment(horizontal='center')
                
                # Цветовое кодирование качества по категориям
                if col == 2:
                    quality = cat_row['Среднее качество']
                    if quality >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif quality >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif quality >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    elif quality >= 50:
                        cell.fill = self.styles['QUALITY_50_70']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
        
        # Настраиваем ширину колонок
        for col in range(1, len(summary_headers) + 1):
            max_length = 0
            for row in range(start_row + 1, start_row + len(results_df) + 2):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)
        
        # Автофильтр для удобства
        ws.auto_filter.ref = f"A{start_row + 1}:K{start_row + len(results_df) + 1}"
        
        print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Главный отчет создан")
    
    def _create_dashboard_sheet(self, wb, results_df):
        """Создает лист с дашбордом визуализации"""
        print(f"   {TERMINAL_SYMBOLS['GEAR']} Создаем дашборд...")
        
        ws = wb.create_sheet("Дашборд")
        
        # Заголовок
        ws.merge_cells('A1:L1')
        ws['A1'] = f"{EXCEL_SYMBOLS['BAR_CHART']} ДАШБОРД КАЧЕСТВА ДАННЫХ"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Генерируем визуализации
        self._add_visualizations(ws, results_df)
        
        print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Дашборд создан")
    
    def _create_rules_status_sheet(self, wb, results_df):
        """Создает лист со статусом каждого правила"""
        print(f"   {TERMINAL_SYMBOLS['GEAR']} Создаем детальный статус правил...")
        
        ws = wb.create_sheet("Статус правил")
        
        # Заголовок
        ws.merge_cells('A1:M1')
        ws['A1'] = f"{EXCEL_SYMBOLS['MAGNIFYING_GLASS']} ДЕТАЛЬНЫЙ СТАТУС ПРАВИЛ"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Легенда цветов
        ws['A3'] = f"{EXCEL_SYMBOLS['PALETTE']} ЛЕГЕНДА ЦВЕТОВ:"
        ws['A3'].font = Font(bold=True, size=11)
        
        legend_items = [
            ("✅ УСПЕШНО", "C6EFCE", "Правило успешно выполнено"),
            ("❌ НЕУСПЕШНО", "FFC7CE", "Правило нарушено"),
            ("⚠️ ПРОБЛЕМНОЕ", "FF9999", "Правило содержит ошибку в логике"),
            ("📊 КАЧЕСТВО 95-100%", "92D050", "Отличное качество"),
            ("📊 КАЧЕСТВО 85-95%", "C6EFCE", "Хорошее качество"),
            ("📊 КАЧЕСТВО 70-85%", "FFEB9C", "Удовлетворительное качество"),
            ("📊 КАЧЕСТВО <70%", "FF9999", "Низкое качество")
        ]
        
        for idx, (label, color, description) in enumerate(legend_items):
            row = 3 + idx
            ws.cell(row=row, column=2, value=label).font = Font(bold=True)
            ws.cell(row=row, column=3, value=description)
            
            # Цветной квадратик
            color_cell = ws.cell(row=row, column=1)
            color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            color_cell.value = "   "
        
        # Создаем детальную таблицу
        start_row = len(legend_items) + 5
        
        headers = ["Код правила", "Описание", "Категория", "Таблица", 
                  "Статус", "Качество", "Ошибок", "Оценено",
                  "Рекомендация", "Приоритет", "Ссылка на ошибки"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Заполняем данными
        for idx, (_, row) in enumerate(results_df.iterrows(), 1):
            data_row = start_row + idx
            
            # Генерируем рекомендацию и приоритет
            recommendation, priority = self._generate_rule_recommendation(row)
            
            row_data = [
                row['rule_code'],
                row['rule_description'],
                row['quality_category'],
                row['table_name'],
                row['status'],
                f"{row['success_rate_%']:.1f}%",
                row['failed'],
                row['total_records'],
                recommendation,
                priority,
                "Есть" if row['error_file'] else "Нет"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = self.styles['THIN_BORDER']
                
                # Цветовое кодирование
                if col == 5:  # Статус
                    if value == "PASSED":
                        cell.fill = self.styles['STATUS_PASSED']
                        cell.font = Font(bold=True, color="006100")
                    else:
                        cell.fill = self.styles['STATUS_FAILED']
                        cell.font = Font(bold=True, color="9C0006")
                
                elif col == 6:  # Качество
                    percent = float(str(value).replace('%', ''))
                    if percent >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif percent >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif percent >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
                
                elif col == 10:  # Приоритет
                    if value == "Высокий":
                        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif value == "Средний":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True)
                
                # Особое выделение проблемных правил
                if row['rule_code'] in ["RCCONF_12.2", "RCCONF_12.3"]:
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Настраиваем ширину колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(start_row, start_row + len(results_df) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)
        
        print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Статус правил создан")
    
    def _create_category_summary_sheet(self, wb, results_df):
        """Создает лист со сводкой по категориям качества"""
        print(f"   {TERMINAL_SYMBOLS['GEAR']} Создаем сводку по категориям...")
        
        ws = wb.create_sheet("Сводка по категориям")
        
        # Группируем по категориям
        category_stats = results_df.groupby('quality_category').agg({
            'rule_code': 'count',
            'success_rate_%': 'mean',
            'total_records': 'sum',
            'passed': 'sum',
            'failed': 'sum'
        }).round(2)
        
        # Создаем визуальную таблицу
        ws['A1'] = f"{EXCEL_SYMBOLS['CHECKERED_FLAG']} СВОДКА ПО КАТЕГОРИЯМ КАЧЕСТВА"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:F1')
        
        headers = ["Категория качества", "Кол-во правил", "Среднее качество", 
                  "Оценено", "Успешно", "Ошибок"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Заполняем данные
        for idx, (category, stats) in enumerate(category_stats.iterrows(), 1):
            row = 3 + idx
            
            row_data = [
                category,
                int(stats['rule_code']),
                f"{stats['success_rate_%']:.1f}%",
                int(stats['total_records']),
                int(stats['passed']),
                int(stats['failed'])
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Цветовое кодирование качества
                if col == 3:
                    quality = stats['success_rate_%']
                    if quality >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif quality >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif quality >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
        
        print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Сводка по категориям создана")
    
    def _generate_rule_recommendation(self, rule_row):
        """Генерирует рекомендацию для правила"""
        success_rate = rule_row['success_rate_%']
        failed_count = rule_row['failed']
        rule_code = rule_row['rule_code']
        
        # Проблемные правила
        if rule_code in ["RCCONF_12.2", "RCCONF_12.3"]:
            return "Исправить логику правила в конфигурации", "КРИТИЧЕСКИЙ"
        
        if success_rate >= 95:
            return "Поддерживать текущий уровень", "Низкий"
        elif success_rate >= 85:
            if failed_count > 100:
                return "Провести выборочную очистку данных", "Средний"
            else:
                return "Мониторить качество", "Низкий"
        elif success_rate >= 70:
            return f"Требуется очистка {failed_count:,} записей", "Высокий"
        else:
            return f"СРОЧНО исправить {failed_count:,} ошибок", "КРИТИЧЕСКИЙ"
    
    def _add_visualizations(self, ws, results_df):
        """Добавляет визуализации в дашборд"""
        # Здесь можно добавить код для создания графиков matplotlib
        # и их вставки в Excel через openpyxl.drawing.image.Image
        
        # Для простоты создаем текстовые визуализации
        ws['A3'] = "📊 ВИЗУАЛИЗАЦИЯ РЕЗУЛЬТАТОВ:"
        ws['A3'].font = Font(bold=True, size=12)
        
        # Простая текстовая визуализация
        ws['A5'] = "Распределение качества правил:"
        
        # Создаем текстовые "графики"
        for idx, (_, row) in enumerate(results_df.sort_values('success_rate_%').iterrows(), 1):
            quality_bar = "█" * int(row['success_rate_%'] / 5)  # Бар из 20 символов
            ws.cell(row=5 + idx, column=1, value=row['rule_code'])
            ws.cell(row=5 + idx, column=2, value=quality_bar)
            ws.cell(row=5 + idx, column=3, value=f"{row['success_rate_%']:.1f}%")
    
    def _create_html_dashboard(self, results_df):
        """Создает HTML версию дашборда для удобного просмотра"""
        html_path = os.path.join(self.output_dir, "dq_dashboard.html")
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Дашборд качества данных</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
                .dashboard {{ max-width: 1200px; margin: 0 auto; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                         color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
                .kpi-card {{ background: white; padding: 20px; border-radius: 8px; 
                           box-shadow: 0 2px 10px rgba(0,0,0,0.1); margin-bottom: 20px; }}
                .status-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); 
                              gap: 20px; margin: 30px 0; }}
                .status-card {{ padding: 15px; border-radius: 8px; color: white; }}
                .status-passed {{ background: #28a745; }}
                .status-failed {{ background: #dc3545; }}
                .status-warning {{ background: #ffc107; color: #333; }}
                .quality-bar {{ height: 20px; background: #e9ecef; border-radius: 10px; 
                              margin: 5px 0; overflow: hidden; }}
                .quality-fill {{ height: 100%; transition: width 0.5s; }}
                .quality-95 {{ background: #28a745; }}
                .quality-85 {{ background: #20c997; }}
                .quality-70 {{ background: #ffc107; }}
                .quality-low {{ background: #dc3545; }}
                table {{ width: 100%; background: white; border-collapse: collapse; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background: #f8f9fa; font-weight: bold; }}
                tr:hover {{ background: #f5f5f5; }}
                .legend {{ display: flex; gap: 15px; margin: 20px 0; flex-wrap: wrap; }}
                .legend-item {{ display: flex; align-items: center; gap: 5px; }}
                .legend-color {{ width: 20px; height: 20px; border-radius: 3px; }}
            </style>
        </head>
        <body>
            <div class="dashboard">
                <div class="header">
                    <h1>🎯 Дашборд качества данных</h1>
                    <p>Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                
                <div class="status-grid">
                    <div class="kpi-card">
                        <h3>📈 Общее качество</h3>
                        <h2>{results_df['success_rate_%'].mean():.1f}%</h2>
                    </div>
                    <div class="kpi-card">
                        <h3>📋 Проверено правил</h3>
                        <h2>{len(results_df)}</h2>
                    </div>
                    <div class="kpi-card">
                        <h3>⚠️ Найдено ошибок</h3>
                        <h2>{results_df['failed'].sum():,}</h2>
                    </div>
                </div>
                
                <div class="legend">
                    <div class="legend-item"><div class="legend-color" style="background:#28a745"></div> Успешно (>95%)</div>
                    <div class="legend-item"><div class="legend-color" style="background:#20c997"></div> Хорошо (85-95%)</div>
                    <div class="legend-item"><div class="legend-color" style="background:#ffc107"></div> Удовлетворительно (70-85%)</div>
                    <div class="legend-item"><div class="legend-color" style="background:#dc3545"></div> Требует внимания (<70%)</div>
                </div>
                
                <h2>Результаты проверок</h2>
                <table>
                    <tr>
                        <th>Код правила</th>
                        <th>Описание</th>
                        <th>Качество</th>
                        <th>Статус</th>
                        <th>Ошибок</th>
                    </tr>
        """
        
        for _, row in results_df.sort_values('success_rate_%').iterrows():
            quality_class = "quality-low"
            if row['success_rate_%'] >= 95:
                quality_class = "quality-95"
            elif row['success_rate_%'] >= 85:
                quality_class = "quality-85"
            elif row['success_rate_%'] >= 70:
                quality_class = "quality-70"
            
            status_class = "status-passed" if row['status'] == "PASSED" else "status-failed"
            
            html_content += f"""
                    <tr>
                        <td>{row['rule_code']}</td>
                        <td>{row['rule_description'][:50]}...</td>
                        <td>
                            <div class="quality-bar">
                                <div class="quality-fill {quality_class}" style="width:{row['success_rate_%']}%"></div>
                            </div>
                            {row['success_rate_%']:.1f}%
                        </td>
                        <td><span class="{status_class}" style="padding:5px 10px;border-radius:4px">{row['status']}</span></td>
                        <td>{row['failed']:,}</td>
                    </tr>
            """
        
        html_content += """
                </table>
            </div>
        </body>
        </html>
        """
        
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"   {TERMINAL_SYMBOLS['SUCCESS']} HTML дашборд создан: {html_path}")
        
        return html_path