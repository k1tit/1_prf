# # utils/reporting.py
# """Генератор Excel отчетов с объединением ошибок"""

# import os
# import sys
# from datetime import datetime
# import pandas as pd
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

# # Добавляем путь для импорта
# current_dir = os.path.dirname(os.path.abspath(__file__))
# parent_dir = os.path.dirname(current_dir)
# if parent_dir not in sys.path:
#     sys.path.insert(0, parent_dir)

# try:
#     from utils.symbols import EXCEL_SYMBOLS, TERMINAL_SYMBOLS
# except ImportError:
#     # Локальные константы если не удалось импортировать
#     EXCEL_SYMBOLS = {
#         'SUCCESS': '✓', 'ERROR': '✗', 'WARNING': '⚠',
#         'INFO': 'ℹ', 'ROCKET': '🚀', 'GEAR': '⚙',
#         'MAGNIFYING_GLASS': '🔍', 'CHECKMARK': '✓',
#         'CHART_UP': '📈', 'CHART_DOWN': '📉',
#         'FILE_FOLDER': '📁', 'PAGE': '📄', 'TABLE': '📊',
#         'BOOKS': '📚', 'SAVE': '💾', 'BAR_CHART': '📊',
#     }
    
#     TERMINAL_SYMBOLS = {
#         'SUCCESS': '[OK]', 'ERROR': '[ERROR]', 'WARNING': '[WARN]',
#         'INFO': '[INFO]', 'SKIP': '[SKIP]', 'ROCKET': '[START]',
#         'GEAR': '[PROC]', 'MAGNIFYING_GLASS': '[CHECK]',
#         'CHECKMARK': '[DONE]', 'CHART_UP': '[STAT+]', 'CHART_DOWN': '[STAT-]',
#         'CLIPBOARD': '[CLIP]', 'FILE_FOLDER': '[DIR]', 'PAGE': '[FILE]',
#         'TABLE': '[TABLE]', 'COLUMN': '[COL]', 'BOOKS': '[DATA]',
#         'SAVE': '[SAVE]', 'TARGET': '[TARGET]', 'PALETTE': '[STYLE]',
#         'CELEBRATION': '[DONE]', 'BAR_CHART': '[CHART]', 'MEMO': '[NOTE]',
#     }

# class ExcelReportGenerator:
#     """Генератор Excel отчетов с объединением ошибок"""
    
#     def __init__(self, output_dir, error_manager=None):
#         self.output_dir = output_dir
#         self.error_manager = error_manager
    
#     def create_comprehensive_report(self, results_df, include_errors=True):
#         """Создает комплексный Excel отчет со сводкой и ошибками"""
#         if results_df.empty:
#             return None
            
#         timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
#         excel_path = os.path.join(self.output_dir, f"DQ_Full_Report_{timestamp}.xlsx")
        
#         print(f"{TERMINAL_SYMBOLS['PALETTE']} Создаем комплексный Excel отчет...")
        
#         wb = Workbook()
#         wb.remove(wb.active)  # Удаляем дефолтный лист
        
#         # 1. Создаем лист со сводкой проверок
#         self._create_summary_sheet(wb, results_df)
        
#         # 2. Лист с детальной статистикой
#         self._create_detailed_stats_sheet(wb, results_df)
        
#         # 3. Лист с рекомендациями
#         self._create_recommendations_sheet(wb, results_df)
        
#         # 4. Если есть менеджер ошибок, добавляем сводку ошибок
#         if self.error_manager and include_errors:
#             # Сохраняем все ошибки в отдельный файл
#             if hasattr(self.error_manager, 'save_all_errors_to_excel'):
#                 errors_file = self.error_manager.save_all_errors_to_excel()
                
#                 if errors_file and os.path.exists(errors_file):
#                     # Копируем листы с ошибками в основной отчет
#                     self._merge_error_sheets(wb, errors_file)
#                     print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Ошибки добавлены в отчет")
#                 else:
#                     print(f"   {TERMINAL_SYMBOLS['WARNING']} Файл с ошибками не найден")
        
#         # Сохраняем файл
#         wb.save(excel_path)
        
#         print(f"{TERMINAL_SYMBOLS['CELEBRATION']} ПОЛНЫЙ ОТЧЕТ СОХРАНЕН: {excel_path}")
#         print(f"   {EXCEL_SYMBOLS['PAGE']} Листов: {len(wb.sheetnames)}")
        
#         return excel_path
    
#     def _create_summary_sheet(self, wb, results_df):
#         """Создает красивый лист со сводкой"""
#         ws = wb.create_sheet("Сводка проверок")
        
#         # СТИЛИ
#         header_font = Font(bold=True, size=12, color="FFFFFF")
#         title_font = Font(bold=True, size=14, color="1F4E78")
#         normal_font = Font(size=9)
        
#         header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#         light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
#         success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#         error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#         warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
#         thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                            top=Side(style='thin'), bottom=Side(style='thin'))
        
#         center_align = Alignment(horizontal='center', vertical='center')
#         left_align = Alignment(horizontal='left', vertical='center')
        
#         # ЗАГОЛОВОК
#         ws.merge_cells('A1:P1')
#         ws['A1'] = f"{EXCEL_SYMBOLS['BAR_CHART']} СВОДКА ПРОВЕРКИ КАЧЕСТВА ДАННЫХ"
#         ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
#         ws['A1'].alignment = center_align
#         ws['A1'].fill = light_blue_fill
        
#         # Статистика
#         total_errors = results_df["failed"].sum()
#         avg_quality = results_df["success_rate_%"].mean()
#         total_records = results_df["total_records"].sum()
        
#         ws.merge_cells('A2:P2')
#         ws['A2'] = f"Всего ошибок: {total_errors:,} | Среднее качество: {avg_quality:.1f}% | Проверено правил: {len(results_df)} | Обработано записей: {total_records:,}"
#         ws['A2'].font = Font(bold=True, size=11, color="2F5496")
#         ws['A2'].alignment = center_align
        
#         # Информация о файлах с ошибками
#         if self.error_manager and hasattr(self.error_manager, 'get_errors_directory'):
#             errors_dir = self.error_manager.get_errors_directory()
#             if errors_dir:
#                 ws.merge_cells('A3:P3')
#                 ws['A3'] = f"Детальные ошибки сохранены в папке: {errors_dir}"
#                 ws['A3'].font = Font(italic=True, size=10, color="FF0000")
#                 ws['A3'].alignment = center_align
#                 ws['A3'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
#         # Заголовки таблицы
#         headers = list(results_df.columns)
#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=5, column=col, value=header)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = center_align
#             cell.border = thin_border
        
#          # Данные с окрашиванием
#         PROBLEMATIC_RULES = ["RCCONF_12.2", "RCCONF_12.3"]
        
#         for row_idx, (_, row_data) in enumerate(results_df.iterrows(), 6):
#             for col_idx, value in enumerate(row_data, 1):
#                 cell = ws.cell(row=row_idx, column=col_idx, value=value)
#                 cell.border = thin_border
#                 cell.alignment = left_align
#                 cell.font = normal_font
                
#                 # Проверяем проблемные правила
#                 rule_code = row_data['rule_code'] if hasattr(row_data, '__getitem__') else row_data[headers.index('rule_code')]
#                 is_problematic = rule_code in PROBLEMATIC_RULES
                
#                 # Окрашивание строк по статусу
#                 status_col_idx = headers.index("status") + 1 if "status" in headers else -1
#                 failed_col_idx = headers.index("failed") + 1 if "failed" in headers else -1
#                 success_rate_idx = headers.index("success_rate_%") + 1 if "success_rate_%" in headers else -1
                
#                 if is_problematic:
#                     # Проблемные правила - красный фон
#                     cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
#                     cell.font = Font(bold=True, color="FFFFFF")
#                 elif status_col_idx != -1 and col_idx == status_col_idx:
#                     if value == "PASSED":
#                         cell.fill = success_fill
#                     elif value == "FAILED":
#                         cell.fill = error_fill
#                 elif failed_col_idx != -1 and col_idx == failed_col_idx and value > 0:
#                     cell.fill = error_fill
#                 elif success_rate_idx != -1 and col_idx == success_rate_idx:
#                     if value >= 95:
#                         cell.fill = success_fill
#                     elif value >= 80:
#                         cell.fill = warning_fill
#                     else:
#                         cell.fill = error_fill
        
#         # Настраиваем ширину колонок
#         for col in range(1, len(headers) + 1):
#             column_letter = get_column_letter(col)
#             max_length = 0
#             for row in range(5, 5 + len(results_df) + 1):
#                 cell = ws.cell(row=row, column=col)
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
#     def _create_detailed_stats_sheet(self, wb, results_df):
#         """Создает лист с детальной статистикой"""
#         ws = wb.create_sheet("Детальная статистика")
        
#         # Стили
#         header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
#         header_font = Font(bold=True, color="FFFFFF", size=11)
#         title_font = Font(bold=True, size=14, color="1F4E78")
#         problematic_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Красный для проблемных
        
#         # Заголовок
#         ws.merge_cells('A1:L1')
#         ws['A1'] = f"{EXCEL_SYMBOLS['BAR_CHART']} ДЕТАЛЬНАЯ СТАТИСТИКА ПРОВЕРОК"
#         ws['A1'].font = title_font
#         ws['A1'].alignment = Alignment(horizontal="center")

#         PROBLEMATIC_RULES = ["RCCONF_12.2", "RCCONF_12.3"]
        
#         # Сортируем по проценту успеха
#         sorted_df = results_df.sort_values('success_rate_%', ascending=True)
        
#         # Заголовки
#         headers = ["№", "Код правила", "Описание", "Таблица", "Категория", 
#                 "Всего записей", "Успешно", "Ошибок", "% успеха", 
#                 "Статус", "Время (сек)", "Файл ошибок"]
        
#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col, value=header)
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             # Проверяем проблемные правила
#             rule_code = row['rule_code']
#             is_problematic = rule_code in PROBLEMATIC_RULES

#             if is_problematic:
#                 # Проблемные правила - красный фон
#                 cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
#                 cell.font = Font(bold=True, color="FFFFFF")
        
#         # Список проблемных правил, которые всегда неверны
#         PROBLEMATIC_RULES = ["RCCONF_12.2", "RCCONF_12.3"]
        
#         # Данные
#         for idx, (_, row) in enumerate(sorted_df.iterrows(), 1):
#             row_data = [
#                 idx,
#                 row['rule_code'],
#                 row['rule_description'][:50] + "..." if len(row['rule_description']) > 50 else row['rule_description'],
#                 row['table_name'],
#                 row['quality_category'],
#                 row['total_records'],
#                 row['passed'],
#                 row['failed'],
#                 f"{row['success_rate_%']:.1f}%",
#                 row['status'],
#                 f"{row['execution_time_sec']:.2f}",
#                 "Есть" if row['error_file'] else "Нет"
#             ]
            
#             for col, value in enumerate(row_data, 1):
#                 cell = ws.cell(row=3 + idx, column=col, value=value)
                
#                 # Проверяем проблемные правила
#                 rule_code = row['rule_code']
#                 is_problematic = rule_code in PROBLEMATIC_RULES
                
#                 # Цветовое кодирование
#                 if is_problematic:
#                     # Проблемные правила - красный фон
#                     cell.fill = problematic_fill
#                     cell.font = Font(bold=True, color="FFFFFF")
#                 elif col == 9:  # % успеха
#                     percent_value = float(str(value).replace('%', ''))
#                     if percent_value >= 95:
#                         cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#                     elif percent_value >= 80:
#                         cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
#                     else:
#                         cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
#                 elif col == 10:  # Статус
#                     if value == "PASSED":
#                         cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#                     else:
#                       cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
#         # Автоширина
#         for col in range(1, len(headers) + 1):
#             column_letter = get_column_letter(col)
#             max_length = 0
#             for row in range(3, 3 + len(sorted_df) + 1):
#                 cell = ws.cell(row=row, column=col)
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
#         # Замораживаем заголовки
#         ws.freeze_panes = 'A4'
        
#         # Добавляем примечание о проблемных правилах
#         note_row = len(sorted_df) + 5
#         ws.merge_cells(f'A{note_row}:L{note_row}')
#         ws[f'A{note_row}'] = "ПРИМЕЧАНИЕ: Правила, выделенные красным, содержат ошибки в логике и требуют исправления"
#         ws[f'A{note_row}'].font = Font(bold=True, italic=True, color="FF0000")
#         ws[f'A{note_row}'].fill = problematic_fill
#         ws[f'A{note_row}'].alignment = Alignment(horizontal="center")
        
#         # Автоширина
#         for col in range(1, len(headers) + 1):
#             column_letter = get_column_letter(col)
#             max_length = 0
#             for row in range(3, 3 + len(sorted_df) + 1):
#                 cell = ws.cell(row=row, column=col)
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
#         # Замораживаем заголовки
#         ws.freeze_panes = 'A4'
    
#     def _create_recommendations_sheet(self, wb, results_df):
#         """Создает лист с рекомендациями по исправлению"""
#         ws = wb.create_sheet("Рекомендации")
        
#         # Стили
#         title_font = Font(bold=True, size=14, color="1F4E78")
#         header_font = Font(bold=True, size=11)
        
#         # Заголовок
#         ws.merge_cells('A1:D1')
#         ws['A1'] = f"{EXCEL_SYMBOLS['MEMO']} РЕКОМЕНДАЦИИ ПО УЛУЧШЕНИЮ КАЧЕСТВА ДАННЫХ"
#         ws['A1'].font = title_font
#         ws['A1'].alignment = Alignment(horizontal="center")
#         ws['A1'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
#         # Генерируем рекомендации
#         recommendations = self._generate_recommendations(results_df)
        
#         # Заголовки таблица
#         headers = ["Приоритет", "Рекомендация", "Затронутые правила", "Ожидаемый эффект"]
        
#         for col, header in enumerate(headers, 1):
#             cell = ws.cell(row=3, column=col, value=header)
#             cell.font = header_font
#             cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
#             cell.alignment = Alignment(horizontal="center")
        
#         # Рекомендации
#         for idx, rec in enumerate(recommendations, 1):
#             row_data = [
#                 rec['priority'],
#                 rec['recommendation'],
#                 rec['affected_rules'],
#                 rec['expected_effect']
#             ]
            
#             for col, value in enumerate(row_data, 1):
#                 cell = ws.cell(row=3 + idx, column=col, value=value)
                
#                 # Выделяем приоритеты
#                 if col == 1:
#                     if value == "Высокий":
#                         cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#                         cell.font = Font(bold=True, color="9C0006")
#                     elif value == "Средний":
#                         cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
#                         cell.font = Font(bold=True, color="9C6500")
#                     else:
#                         cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
#         # Автоширина
#         for col in range(1, len(headers) + 1):
#             column_letter = get_column_letter(col)
#             max_length = 0
#             for row in range(3, 3 + len(recommendations) + 1):
#                 cell = ws.cell(row=row, column=col)
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
#     def _generate_recommendations(self, results_df):
#         """Генерирует рекомендации на основе результатов проверок"""
#         recommendations = []
        
#         # Находим правила с низким процентом успеха
#         failed_rules = results_df[results_df['success_rate_%'] < 80]
        
#         if len(failed_rules) > 0:
#             # Группируем по категориям
#             for category in failed_rules['quality_category'].unique():
#                 category_rules = failed_rules[failed_rules['quality_category'] == category]
                
#                 if category == "Completeness":
#                     recommendations.append({
#                         'priority': "Высокий",
#                         'recommendation': f"Заполнить обязательные поля в таблицах: {', '.join(category_rules['table_name'].unique())}",
#                         'affected_rules': f"{len(category_rules)} правил",
#                         'expected_effect': "Увеличение полноты данных на 20-40%"
#                     })
#                 elif category == "Conformity":
#                     recommendations.append({
#                         'priority': "Средний",
#                         'recommendation': f"Привести данные к стандартному формату в таблицах: {', '.join(category_rules['table_name'].unique()[:3])}",
#                         'affected_rules': f"{len(category_rules)} правил",
#                         'expected_effect': "Устранение несоответствий справочникам"
#                     })
        
#         # Добавляем общие рекомендации
#         avg_quality = results_df['success_rate_%'].mean()
        
#         if avg_quality < 90:
#             recommendations.append({
#                 'priority': "Высокий",
#                 'recommendation': "Провести очистку данных: удалить дубликаты, исправить опечатки, стандартизировать форматы",
#                 'affected_rules': "Все правила",
#                 'expected_effect': f"Повышение общего качества данных с {avg_quality:.1f}% до 95%+"
#             })
        
#         if len(results_df[results_df['failed'] > 1000]) > 0:
#             recommendations.append({
#                 'priority': "Высокий",
#                 'recommendation': "Внедрить валидацию данных на этапе ввода (маски, справочники, ограничения)",
#                 'affected_rules': "Правила с >1000 ошибок",
#                 'expected_effect': "Снижение ошибок ввода на 70-90%"
#             })
        
#         # Если мало рекомендаций, добавляем общие
#         if len(recommendations) < 3:
#             recommendations.extend([
#                 {
#                     'priority': "Средний",
#                     'recommendation': "Регулярно обновлять справочники и эталонные данные",
#                     'affected_rules': "Правила соответствия",
#                     'expected_effect': "Актуальность проверок соответствия"
#                 },
#                 {
#                     'priority': "Низкий",
#                     'recommendation': "Внедрить автоматические отчеты по качеству данных",
#                     'affected_rules': "Все правила",
#                     'expected_effect': "Своевременное выявление проблем"
#                 }
#             ])
        
#         return recommendations
    
#     def _merge_error_sheets(self, main_wb, errors_file_path):
#         """Копирует листы с ошибками в основной отчет"""
#         try:
#             errors_wb = load_workbook(errors_file_path)
            
#             for sheet_name in errors_wb.sheetnames:
#                 # Пропускаем служебные листы
#                 if sheet_name in ['Сводка ошибок']:
#                     continue
                    
#                 # Копируем лист
#                 source_ws = errors_wb[sheet_name]
#                 new_ws = main_wb.create_sheet(sheet_name)
                
#                 # Копируем данные
#                 for row in source_ws.iter_rows(values_only=True):
#                     new_ws.append(row)
                
#                 # Копируем стили (упрощенно)
#                 for row in source_ws.iter_rows():
#                     for cell in row:
#                         new_cell = new_ws[cell.coordinate]
#                         if cell.has_style:
#                             new_cell.font = cell.font
#                             new_cell.fill = cell.fill
#                             new_cell.border = cell.border
#                             new_cell.alignment = cell.alignment
                
#                 print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Добавлен лист: {sheet_name}")
                
#         except Exception as e:
#             print(f"   {TERMINAL_SYMBOLS['ERROR']} Ошибка при объединении ошибок: {e}")

"""Генератор Excel отчетов с объединением ошибок - УПРОЩЕННАЯ ВЕРСИЯ"""

import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Добавляем путь для импорта
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

try:
    from utils.symbols import EXCEL_SYMBOLS, TERMINAL_SYMBOLS
except ImportError:
    # Локальные константы если не удалось импортировать
    EXCEL_SYMBOLS = {
        'SUCCESS': '[OK]', 'ERROR': '[X]', 'WARNING': '[!]',
        'INFO': '[i]', 'ROCKET': '[>]', 'GEAR': '[*]',
        'MAGNIFYING_GLASS': '[?]', 'CHECKMARK': '[OK]',
        'CHART_UP': '[+]', 'CHART_DOWN': '[-]',
        'FILE_FOLDER': '[DIR]', 'PAGE': '[FILE]', 'TABLE': '[TBL]',
        'BOOKS': '[DATA]', 'SAVE': '[SAVE]', 'BAR_CHART': '[CHART]',
    }
    
    TERMINAL_SYMBOLS = {
        'SUCCESS': '[OK]', 'ERROR': '[ERROR]', 'WARNING': '[WARN]',
        'INFO': '[INFO]', 'SKIP': '[SKIP]', 'ROCKET': '[START]',
        'GEAR': '[PROC]', 'MAGNIFYING_GLASS': '[CHECK]',
        'CHECKMARK': '[DONE]', 'CHART_UP': '[STAT+]', 'CHART_DOWN': '[STAT-]',
        'CLIPBOARD': '[CLIP]', 'FILE_FOLDER': '[DIR]', 'PAGE': '[FILE]',
        'TABLE': '[TABLE]', 'COLUMN': '[COL]', 'BOOKS': '[DATA]',
        'SAVE': '[SAVE]', 'TARGET': '[TARGET]', 'PALETTE': '[STYLE]',
        'CELEBRATION': '[DONE]', 'BAR_CHART': '[CHART]', 'MEMO': '[NOTE]',
    }

class ExcelReportGenerator:
    """Генератор Excel отчетов с объединением ошибок"""
    
    def __init__(self, output_dir, error_manager=None):
        self.output_dir = output_dir
        self.error_manager = error_manager
    
    def create_comprehensive_report(self, results_df, include_errors=True):
        """Создает комплексный Excel отчет со сводкой и ошибками - УПРОЩЕННЫЙ"""
        if results_df.empty:
            return None
            
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        excel_path = os.path.join(self.output_dir, f"DQ_Full_Report_{timestamp}.xlsx")
        
        print(f"{TERMINAL_SYMBOLS['PALETTE']} Создаем комплексный Excel отчет...")
        
        wb = Workbook()
        wb.remove(wb.active)  # Удаляем дефолтный лист
        
        # 1. Создаем лист с цветной сводкой
        self._create_color_summary_sheet(wb, results_df)
        
        # 2. Лист с детальной статистикой
        self._create_simple_stats_sheet(wb, results_df)
        
        # 3. Если есть менеджер ошибок, добавляем сводку ошибок
        if self.error_manager and include_errors:
            # Сохраняем все ошибки в отдельный файл
            if hasattr(self.error_manager, 'save_all_errors_to_excel'):
                errors_file = self.error_manager.save_all_errors_to_excel()
                
                if errors_file and os.path.exists(errors_file):
                    # Копируем листы с ошибками в основной отчет
                    self._merge_error_sheets(wb, errors_file)
                    print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Ошибки добавлены в отчет")
                else:
                    print(f"   {TERMINAL_SYMBOLS['WARNING']} Файл с ошибками не найден")
        
        # Сохраняем файл
        wb.save(excel_path)
        
        print(f"{TERMINAL_SYMBOLS['CELEBRATION']} ПОЛНЫЙ ОТЧЕТ СОХРАНЕН: {excel_path}")
        print(f"   {EXCEL_SYMBOLS['PAGE']} Листов: {len(wb.sheetnames)}")
        
        return excel_path
    
    def _create_color_summary_sheet(self, wb, results_df):
        """Создает лист с цветной сводкой - с выделением проблемных правил"""
        ws = wb.create_sheet("Цветная сводка")
        
        # СТИЛИ
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # ЦВЕТА ДЛЯ СТРОК
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # ЗЕЛЕНЫЙ - нет ошибок
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")        # КРАСНЫЙ - есть ошибки
        bright_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # ЯРКО-КРАСНЫЙ - ошибка в правиле
        dark_red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")   # ТЕМНО-КРАСНЫЙ - критическая ошибка
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ЗАГОЛОВОК
        ws.merge_cells('A1:Q1')
        ws['A1'] = f"{EXCEL_SYMBOLS['BAR_CHART']} ЦВЕТНАЯ СВОДКА ПРОВЕРОК"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Статистика
        total_errors = results_df["failed"].sum()
        avg_quality = results_df["success_rate_%"].mean()
        passed_rules = len(results_df[results_df['failed'] == 0])
        failed_rules = len(results_df[results_df['failed'] > 0])
        
        # Считаем проблемные правила
        problematic_rules = ["RCCONF_12.2", "RCCONF_12.3"]
        problematic_count = len(results_df[results_df['rule_code'].isin(problematic_rules)])
        
        ws.merge_cells('A2:Q2')
        ws['A2'] = f"ЗЕЛЕНЫЕ = нет ошибок | КРАСНЫЕ = есть ошибки | ЯРКО-КРАСНЫЕ = ошибка в логике правила"
        ws['A2'].font = Font(bold=True, size=10, color="2F5496")
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:Q3')
        ws['A3'] = f"[OK] Успешно: {passed_rules} | [!] С ошибками: {failed_rules} | Проблемные: {problematic_count} | Всего ошибок: {total_errors:,}"
        ws['A3'].font = Font(bold=True, size=10, color="000000")
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Заголовки таблицы
        headers = [
            "Код правила", "Описание", "Категория", "Таблица", 
            "Колонка", "Всего записей", "Успешно", "Ошибок", 
            "% успеха", "Статус", "Время (сек)", "Файл ошибок"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # ПРАВИЛА С ЯРКО-КРАСНЫМ ВЫДЕЛЕНИЕМ
        PROBLEMATIC_RULES = ["RCCONF_12.2", "RCCONF_12.3"]
        
        for row_idx, (_, row) in enumerate(results_df.iterrows(), 1):
            data_row = 5 + row_idx
            
            # Определяем цвет строки
            rule_code = str(row['rule_code'])
            
            # Правила 12.2 и 12.3 - ЯРКО-КРАСНЫЕ (ошибка в логике правила)
            if rule_code in PROBLEMATIC_RULES:
                row_fill = bright_red_fill  # ЯРКО-КРАСНЫЙ
                font_color = "FFFFFF"  # Белый текст
                status = "[!] ОШИБКА ПРАВИЛА"
            # Если нет ошибок - зеленый
            elif row['failed'] == 0:
                row_fill = green_fill
                font_color = "000000"  # Черный текст
                status = "[OK] УСПЕШНО"
            # Если есть ошибки - красный
            else:
                row_fill = red_fill
                font_color = "000000"  # Черный текст
                status = "[!] ОШИБКИ"
            
            # Подготавливаем данные для строки
            row_data = [
                row['rule_code'],
                row['rule_description'][:60] + "..." if len(row['rule_description']) > 60 else row['rule_description'],
                row['quality_category'],
                row['table_name'],
                row.get('matched_column', row.get('column_checked', '')),
                row['total_records'],
                row['passed'],
                row['failed'],
                f"{row['success_rate_%']:.1f}%",
                status,
                f"{row['execution_time_sec']:.2f}",
                "Есть" if row.get('error_file') else "Нет"
            ]
            
            # Заполняем строку одним цветом
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = Font(size=9, color=font_color, bold=(rule_code in PROBLEMATIC_RULES))
                cell.fill = row_fill  # Вся строка одним цветом
                
                # Центрирование для числовых колонок
                if col_idx in [6, 7, 8, 9, 10, 12]:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Настраиваем ширину колонок
        column_widths = [15, 40, 12, 12, 15, 12, 12, 10, 10, 10, 12, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Добавляем примечание о проблемных правилах
        note_row = len(results_df) + 7
        ws.merge_cells(f'A{note_row}:M{note_row}')
        ws[f'A{note_row}'] = "[!] ВНИМАНИЕ: Правила RCCONF_12.2 и RCCONF_12.3 выделены ЯРКО-КРАСНЫМ - они содержат ошибки в логике и требуют исправления!"
        ws[f'A{note_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{note_row}'].fill = bright_red_fill
        ws[f'A{note_row}'].alignment = Alignment(horizontal='center')
        
        # Замораживаем заголовки
        ws.freeze_panes = 'A6'
    
    def _create_simple_stats_sheet(self, wb, results_df):
        """Создает лист с простой статистикой"""
        ws = wb.create_sheet("Статистика")
        
        # Стили
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Цвета строк
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{EXCEL_SYMBOLS['CHART_UP']} СТАТИСТИКА ПО КАТЕГОРИЯМ"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Группируем по категориям
        category_stats = results_df.groupby('quality_category').agg({
            'rule_code': 'count',
            'success_rate_%': 'mean',
            'total_records': 'sum',
            'passed': 'sum',
            'failed': 'sum'
        }).round(2)
        
        # Заголовки
        headers = ["Категория", "Правил", "Среднее качество", "Всего записей", "Ошибок", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Данные по категориям
        for idx, (category, stats) in enumerate(category_stats.iterrows(), 1):
            row = 3 + idx
            
            # Определяем цвет строки
            avg_quality = stats['success_rate_%']
            if avg_quality >= 95:
                row_fill = green_fill
                status = "ОТЛИЧНО"
            elif avg_quality >= 80:
                row_fill = green_fill
                status = "ХОРОШО"
            elif avg_quality >= 60:
                row_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status = "СРЕДНЕ"
            else:
                row_fill = red_fill
                status = "ПЛОХО"
            
            row_data = [
                category,
                int(stats['rule_code']),
                f"{stats['success_rate_%']:.1f}%",
                int(stats['total_records']),
                int(stats['failed']),
                status
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Настраиваем ширину
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_simple_color_report(self, results_df):
        """Создает ОЧЕНЬ ПРОСТОЙ цветной отчет (только один лист)"""
        if results_df.empty:
            return None
            
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        excel_path = os.path.join(self.output_dir, f"DQ_Color_Simple_{timestamp}.xlsx")
        
        print(f"{TERMINAL_SYMBOLS['PALETTE']} Создаем ПРОСТОЙ цветной отчет...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты проверок"
        
        # ЦВЕТА
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # ЗАГОЛОВОК
        ws['A1'] = "🎯 РЕЗУЛЬТАТЫ ПРОВЕРОК КАЧЕСТВА ДАННЫХ"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:F1')
        
        # Подзаголовок
        total_errors = results_df["failed"].sum()
        passed_rules = len(results_df[results_df['failed'] == 0])
        failed_rules = len(results_df[results_df['failed'] > 0])
        
        ws['A2'] = f"[OK] Успешно: {passed_rules} правил | [!] С ошибками: {failed_rules} правил | Всего ошибок: {total_errors:,}"
        ws.merge_cells('A2:F2')
        ws['A2'].font = Font(bold=True, size=10)
        
        # Заголовки таблицы
        headers = ["Код правила", "Описание", "Таблица", "Ошибок", "% успеха", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # ДАННЫЕ С ЦВЕТНЫМИ СТРОКАМИ
        for idx, (_, row) in enumerate(results_df.iterrows(), 1):
            row_idx = 4 + idx
            
            # Определяем цвет строки
            rule_code = str(row['rule_code'])
            
            # Правила 12.2 и 12.3 всегда красные
            if rule_code in ["RCCONF_12.2", "RCCONF_12.3"]:
                row_fill = red_fill
                status = "[!] ОШИБКА ПРАВИЛА"
            # Если нет ошибок - зеленый
            elif row['failed'] == 0:
                row_fill = green_fill
                status = "[OK] УСПЕШНО"
            # Если есть ошибки - красный
            else:
                row_fill = red_fill
                status = "[!] ОШИБКИ"
            
            # Данные для строки
            row_data = [
                row['rule_code'],
                row['rule_description'][:40],
                row['table_name'],
                row['failed'],
                f"{row['success_rate_%']:.1f}%",
                status
            ]
            
            # Заполняем строку одним цветом
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 15  # Код правила
        ws.column_dimensions['B'].width = 40  # Описание
        ws.column_dimensions['C'].width = 12  # Таблица
        ws.column_dimensions['D'].width = 10  # Ошибок
        ws.column_dimensions['E'].width = 12  # % успеха
        ws.column_dimensions['F'].width = 15  # Статус
        
        # Замораживаем заголовки
        ws.freeze_panes = 'A5'
        
        # Сохраняем
        wb.save(excel_path)
        
        print(f"{TERMINAL_SYMBOLS['SUCCESS']} ПРОСТОЙ цветной отчет сохранен: {excel_path}")
        return excel_path
    
    def _merge_error_sheets(self, main_wb, errors_file_path):
        """Копирует листы с ошибками в основной отчет"""
        try:
            errors_wb = load_workbook(errors_file_path)
            
            for sheet_name in errors_wb.sheetnames:
                # Пропускаем служебные листы
                if sheet_name in ['Сводка ошибок']:
                    continue
                    
                # Копируем лист
                source_ws = errors_wb[sheet_name]
                new_ws = main_wb.create_sheet(sheet_name)
                
                # Копируем данные
                for row in source_ws.iter_rows(values_only=True):
                    new_ws.append(row)
                
                print(f"   {TERMINAL_SYMBOLS['SUCCESS']} Добавлен лист: {sheet_name}")
                
        except Exception as e:
            print(f"   {TERMINAL_SYMBOLS['ERROR']} Ошибка при объединении ошибок: {e}")