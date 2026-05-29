"""Основной класс проверки качества данных с поддержкой специальных обработчиков таблиц - ИСПРАВЛЕННАЯ ВЕРСИЯ"""

import os
import json
import pandas as pd
import logging
import traceback
import re
from datetime import datetime
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

try:
    from utils.empty_rows import fully_empty_rows_mask
    from utils.sqlite_safe import connect_sqlite
    from utils.symbols import Symbols
    from utils.column_matcher import ColumnMatcher
    from utils.file_manager import ErrorFileManager
    from core.memory_manager import MemoryManager
    from validators.completeness import CompletenessValidator
    from validators.conformity import ConformityValidator
    from validators.cross_column import CrossColumnEqualityValidator
    from validators.cross_column_equality import CrossColumnEqualityCheckValidator
    from validators.payment_terms_consistency import PaymentTermsConsistencyValidator
    from validators.recon_account_consistency import ReconAccountConsistencyValidator
    from validators.text_validators import (
        SpecialCharactersValidator, 
        ConsecutiveSpacesValidator,
        UppercaseValidator
    )
    from validators.advanced_special_characters import AdvancedSpecialCharactersValidator
    from validators.logical_validator import LogicalValidator
except ImportError as e:
    print(f"Ошибка импорта: {e}")
    raise

class FastDataQualityChecker:
    # KNB1: RCCOMP_113.1 / 115.1 — оценка только при account_group_code = 9038 (KNA1.KTOKD)
    RULES_KTOKD_ONLY_9038_SCOPE = frozenset({"RCCOMP_113.1", "RCCOMP_115.1"})
    RULES_FORCE_KNA1_KTOKD_JOIN = frozenset(
        {"RCCONF_113.1", "RCCONF_115.11", "RCCONF_24.1", "RCCOMP_113.1", "RCCOMP_115.1"}
    )
    # В выгрузку ошибок всегда подтягиваем KTOKD из KNA1 (Group_1)
    RULES_ERROR_EXPORT_KNA1_KTOKD = frozenset(
        {"RCCOMP_113.1", "RCCOMP_115.1", "RCCONF_113.1", "RCCONF_24.1", "RCCONF_115.11"}
    )

    def __init__(self, db_path: str, rules_file: str, output_dir: str = "quality_reports",
                 parallel_tables: int = 0, use_async_load: bool = False, debug: bool = False,
                 reference_datetime=None):
        self.db_path = db_path
        self.rules_file = rules_file
        self.output_dir = output_dir
        self.parallel_tables = max(0, int(parallel_tables))
        self.use_async_load = bool(use_async_load)
        self.debug = bool(debug)
        # Опорная дата для правил «на дату» (например RCCONF_173.1). None = datetime.now() в правиле.
        self.reference_datetime = reference_datetime
        self._parallel_lock = threading.Lock() if self.parallel_tables else None
        
        self.memory_manager = MemoryManager(db_path)
        self.error_manager = ErrorFileManager(output_dir)
        self.column_matcher = ColumnMatcher()
        self.symbols = Symbols()
        
        self.results = []
        self.rule_errors = {}  
        self.suspicious_rules = []  
        self.processed_rules = 0
        self.skipped_rules = 0
        self.logger = logging.getLogger("FastDQChecker")
        
        self.MAX_ERRORS_TO_SAVE = 100000
        self.EXCEL_MAX_ROWS = 1_048_576  # лимит строк в .xlsx (Excel 2007+)
        self.MASS_ERROR_THRESHOLD = 0.5
        
        # Загружаем маппинг колонок
        self.column_map = self._load_column_map()
        # Маппинг AUSP: column_name_checked -> ATINN (правила применяются к строкам с этим ATINN, проверяется ATWRT)
        self.ausp_atinn_mapping = self._load_ausp_atinn_mapping()
        
        self.colors = {
            'green': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'red': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'orange': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'dark_red': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
            'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'header_font': Font(color='FFFFFF', bold=True),
            'normal_font': Font(name='Calibri', size=11),
            'bold_font': Font(bold=True),
            'error_font': Font(color='FF0000', bold=True),
            'success_font': Font(color='00B050', bold=True),
        }
        
        # Статистика в реальном времени
        self.current_table = None
        self.current_rule = None
        self.start_time = None
        self.table_start_time = None
        
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(os.path.join(output_dir, "errors"), exist_ok=True)
        
        logging.basicConfig(level=logging.INFO, format="%(levelname)s:%(name)s: %(message)s")
        
        self.table_handlers = self._load_table_handlers()
    
    def _load_column_map(self):
        """Загружает маппинг колонок из column_map.json"""
        column_map_path = os.path.join(parent_dir, "json files", "column_map.json")
        try:
            if os.path.exists(column_map_path):
                with open(column_map_path, "r", encoding="utf-8") as f:
                    column_map = json.load(f)
                print(f"[INFO] Загружен маппинг колонок для {len(column_map)} таблиц")
                return column_map
            else:
                print(f"[WARN] Файл column_map.json не найден: {column_map_path}")
                return {}
        except Exception as e:
            print(f"[WARN] Ошибка загрузки column_map.json: {e}")
            return {}
    
    def _load_ausp_atinn_mapping(self):
        """Загружает маппинг AUSP: column_name_checked→ATINN и ATINN→временное имя колонки (conf_ausp_atinn_mapping.json)."""
        conf_path = os.path.join(parent_dir, "json files", "conf_ausp_atinn_mapping.json")
        try:
            if os.path.exists(conf_path):
                with open(conf_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                mapping = data.get("column_name_checked_to_atinn", data) if isinstance(data, dict) else {}
                self.ausp_atinn_to_temporary_name = data.get("atinn_to_temporary_column_name", {}) if isinstance(data, dict) else {}
                if mapping:
                    print(f"[INFO] Загружен маппинг AUSP ATINN для {len(mapping)} правил, временные имена колонок: {len(self.ausp_atinn_to_temporary_name)}")
                return mapping
            self.ausp_atinn_to_temporary_name = {}
            return {}
        except Exception as e:
            print(f"[WARN] Ошибка загрузки conf_ausp_atinn_mapping.json: {e}")
            self.ausp_atinn_to_temporary_name = {}
            return {}
    
    def _normalize_ausp_name(self, name):
        """Нормализация имени для AUSP: без учёта регистра, опционально без префикса Z / Z_ (поля могут быть с Z или без)."""
        if not name:
            return ""
        s = (name or "").strip().upper().replace(" ", "_")
        if s.startswith("Z_"):
            return s[2:]
        if s.startswith("Z") and len(s) > 1 and s[1] != "_":
            return s[1:]
        return s

    def _find_ausp_columns(self, columns, table_name):
        """В таблице AUSP колонки ATINN (код характеристики) и ATWRT (значение). Поиск по имени, подстроке и по позиции (2-я и 3-я колонка)."""
        cols = [c for c in columns if c is not None]
        atinn_col = None
        atwrt_col = None
        for c in cols:
            raw = str(c).strip().upper()
            cu = raw.replace(" ", "").replace("_", "")
            if cu == "ATINN":
                atinn_col = c
            if cu == "ATWRT":
                atwrt_col = c
        if not atinn_col:
            atinn_col = self._find_column_alternative(cols, "ATINN", table_name)
        if not atwrt_col:
            atwrt_col = self._find_column_alternative(cols, "ATWRT", table_name)
        # По подстроке: заголовок может содержать ATINN/ATWRT (например "ATINN_Code" или локализованное имя)
        if not atinn_col:
            for c in cols:
                if "ATINN" in str(c).upper() or re.sub(r"[^A-Za-z0-9]", "", str(c).upper()) == "ATINN":
                    atinn_col = c
                    break
        if not atwrt_col:
            for c in cols:
                if "ATWRT" in str(c).upper() or re.sub(r"[^A-Za-z0-9]", "", str(c).upper()) == "ATWRT":
                    atwrt_col = c
                    break
        # Запасной вариант: колонка = только буквы ATINN/ATWRT (BOM/лишние символы)
        if not atinn_col:
            for c in cols:
                if re.sub(r"[^A-Za-z0-9]", "", str(c).upper()) == "ATINN":
                    atinn_col = c
                    break
        if not atwrt_col:
            for c in cols:
                if re.sub(r"[^A-Za-z0-9]", "", str(c).upper()) == "ATWRT":
                    atwrt_col = c
                    break
        # Если не нашли по имени — для AUSP берём 2-ю и 3-ю колонку (слева): часто это ATINN и ATWRT
        if (table_name or "").strip().upper() == "AUSP" and len(cols) >= 3:
            if not atinn_col:
                atinn_col = cols[1]
            if not atwrt_col:
                atwrt_col = cols[2]
        if (table_name or "").strip().upper() == "AUSP" and (not atinn_col or not atwrt_col):
            self._debug_ausp_columns(columns, table_name)
        return atinn_col, atwrt_col

    def _debug_ausp_columns(self, columns, table_name):
        """Выводит отладочную информацию по колонкам AUSP: что пришло, почему не совпало с ATINN/ATWRT."""
        cols = [c for c in columns if c is not None]
        lines = [
            "",
            "   [AUSP DEBUG] === почему не найдены колонки ATINN/ATWRT ===",
            f"   [AUSP DEBUG] Всего колонок: {len(cols)}",
        ]
        for i, c in enumerate(cols):
            raw = str(c).strip().upper()
            cu = raw.replace(" ", "").replace("_", "")
            clean = re.sub(r"[^A-Za-z0-9]", "", raw)
            repr_c = repr(c)[:80]
            lines.append(f"   [AUSP DEBUG]   {i}: repr={repr_c}")
            lines.append(f"   [AUSP DEBUG]       upper={raw!r}  replace(space,_)={cu!r}  clean(only letters)={clean!r}")
            lines.append(f"   [AUSP DEBUG]       ==ATINN? {cu == 'ATINN'}  ==ATWRT? {cu == 'ATWRT'}  'ATINN' in name? {'ATINN' in raw}  'ATWRT' in name? {'ATWRT' in raw}")
        lines.append("   [AUSP DEBUG] === конец отладки ===")
        print("\n".join(lines))

    # --- AUSP: явные условия. Колонка ATINN — в ней значение; по значению определяем имя колонки для проверки. ---
    AUSP_ATINN_TO_COLUMN = {
        "143": "CCAF",           # если ATINN = 143 → считаем колонку как CCAF
        "604": "RED_OUTLET",     # если ATINN = 604 → RED_OUTLET
        "148": "ZGLOBAL_CUSTOMER",
        "151": "ZTRADE_NAME",
    }

    def _ausp_atinn_to_column_name(self, atinn_value):
        """По значению ATINN возвращаем имя колонки для проверки. Условия заданы явно в коде."""
        if atinn_value is None:
            return None
        return self.AUSP_ATINN_TO_COLUMN.get(str(atinn_value).strip())

    def _normalize_atinn_for_filter(self, value):
        """Нормализация значения ATINN для фильтрации: 143, '143', 143.0, ' 143 ' → '143'. По виду цифр сравниваем единообразно."""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        s = str(value).strip()
        try:
            return str(int(float(s)))
        except (ValueError, TypeError):
            return re.sub(r"\.0+$", "", s) if s else ""

    def _ausp_atinn_mask(self, series, atinn_value):
        """Булева маска: строки, где ATINN совпадает с atinn_value (нормализация: 143, 143.0, '143' → одно сравнение)."""
        target = self._normalize_atinn_for_filter(atinn_value)
        if not target:
            return pd.Series(False, index=series.index)
        normalized = series.apply(lambda x: self._normalize_atinn_for_filter(x))
        return normalized == target

    def _resolve_ausp_atinn_value(self, rule):
        """По правилу определяем значение ATINN (143, 148, 151, 604). Сначала явные условия, потом маппинг из конфига."""
        column_to_check = (rule.get("column_name_checked") or "").strip()
        biz = (rule.get("business_attribute_name") or "").strip()
        atinn_value = None
        # Явные условия: по содержимому column_name_checked или business_attribute_name
        if column_to_check:
            cu = column_to_check.upper()
            if "143" in cu or "ATINN(143)" in cu or "ATINN=143" in cu or "ATINN =143" in cu:
                atinn_value = "143"
            elif "604" in cu or "ATINN=604" in cu or "ATINN = 604" in cu:
                atinn_value = "604"
            elif "148" in cu or "ATINN=148" in cu:
                atinn_value = "148"
            elif "151" in cu or "ATINN=151" in cu:
                atinn_value = "151"
        if atinn_value:
            return atinn_value
        if biz:
            bu = biz.upper()
            if "CCAF" in bu or "CAF" in bu:
                return "143"
            if "RED" in bu and "OUTLET" in bu or bu == "OUTLET":
                return "604"
            if "GLOBAL" in bu and "CUSTOMER" in bu:
                return "148"
            if "TRADE" in bu and "NAME" in bu or "TRADING" in bu and "GROUP" in bu:
                return "151"
        # По маппингу из конфига (если есть)
        candidates = [column_to_check] if column_to_check else []
        if biz and biz not in candidates:
            candidates.append(biz)
        if self.ausp_atinn_mapping and candidates:
            for key, val in self.ausp_atinn_mapping.items():
                key_norm = self._normalize_ausp_name(key)
                key_upper = (key or "").strip().upper()
                for cand in candidates:
                    if not cand:
                        continue
                    cand_norm = self._normalize_ausp_name(cand)
                    cand_upper = cand.strip().upper()
                    if key_upper == cand_upper or key_norm == cand_norm:
                        atinn_value = str(val).strip()
                        break
                if atinn_value:
                    break
        if not atinn_value and column_to_check:
            atinn_match = re.search(r"ATINN\s*[=\(]\s*(\d+)", column_to_check, re.IGNORECASE)
            if atinn_match:
                atinn_value = atinn_match.group(1).strip()
        return atinn_value

    def _build_ausp_split(self, df, table_name):
        """
        AUSP: разбиваем по значению колонки ATINN. Имя колонки для проверки задаём условиями:
        если ATINN = 143 → CCAF, если ATINN = 604 → RED_OUTLET, 148 → ZGLOBAL_CUSTOMER, 151 → ZTRADE_NAME.
        """
        if df is None or df.empty or (table_name or "").strip().upper() != "AUSP":
            return None
        atinn_col, atwrt_col = self._find_ausp_columns(df.columns, table_name)
        if not atinn_col or not atwrt_col:
            return None
        out = {}
        for atinn_val, temp_name in self.AUSP_ATINN_TO_COLUMN.items():
            mask = self._ausp_atinn_mask(df[atinn_col], atinn_val)
            slice_df = df.loc[mask].copy()
            slice_df = slice_df.rename(columns={atwrt_col: temp_name})
            out[atinn_val] = (slice_df, temp_name)
        return out

    def _apply_ausp_filter(self, df, column_to_check, table_name, rule=None):
        """
        В таблице AUSP колонка называется ATINN — в ней значения (143, 148, 151, 604 и т.д.).
        column_name_checked в правиле (ATINN(143)1-9, ATINN=143, ZGLOBAL_CUSTOMER и т.д.) — это идентификатор,
        по которому мы определяем, какое значение ATINN фильтровать; проверяем затем колонку ATWRT.
        Возвращает (filtered_df, atwrt_column, temporary_column_name) или (None, None, None).
        """
        if df is None or df.empty:
            return None, None, None
        t = (table_name or "").strip().upper()
        if t != "AUSP":
            return None, None, None
        # Значение ATINN для фильтра: из маппинга или из строки вида ATINN(143), ATINN=604
        candidates = [(column_to_check or "").strip()]
        if rule:
            biz = (rule.get("business_attribute_name") or "").strip()
            if biz and biz not in candidates:
                candidates.append(biz)
        atinn_value = None
        if self.ausp_atinn_mapping:
            for key, val in self.ausp_atinn_mapping.items():
                key_norm = self._normalize_ausp_name(key)
                key_upper = (key or "").strip().upper()
                for cand in candidates:
                    if not cand:
                        continue
                    cand_norm = self._normalize_ausp_name(cand)
                    cand_upper = cand.strip().upper()
                    if key_upper == cand_upper or key_norm == cand_norm:
                        atinn_value = str(val).strip()
                        break
                if atinn_value:
                    break
        if not atinn_value and column_to_check:
            atinn_match = re.search(r"ATINN\s*[=\(]\s*(\d+)", (column_to_check or ""), re.IGNORECASE)
            if atinn_match:
                atinn_value = atinn_match.group(1).strip()
        if not atinn_value:
            return None, None, None
        atinn_col, atwrt_col = self._find_ausp_columns(df.columns, table_name)
        if not atinn_col or not atwrt_col:
            return None, None, None
        mask = self._ausp_atinn_mask(df[atinn_col], atinn_value)
        filtered = df.loc[mask].copy()
        # Имя колонки по условиям: если ATINN = 143 → CCAF, 604 → RED_OUTLET и т.д.
        temporary_name = self._ausp_atinn_to_column_name(atinn_value)
        if not temporary_name and rule:
            temporary_name = (rule.get("business_attribute_name") or "").strip()
        if not temporary_name:
            temporary_name = atwrt_col
        return filtered, atwrt_col, temporary_name
    
    def _print_progress_bar(self, iteration, total, prefix='', suffix='', length=50, fill='█', print_end="\r"):
        """Выводит красивый прогресс-бар"""
        percent = ("{0:.1f}").format(100 * (iteration / float(total)))
        filled_length = int(length * iteration // total)
        bar = fill * filled_length + '░' * (length - filled_length)
        sys.stdout.write(f'\r{prefix} |{bar}| {percent}% {suffix}')
        sys.stdout.flush()
        if iteration == total: 
            sys.stdout.write('\n')
    
    def _print_rule_stats(self, rule_code, total_rows, error_count, exec_time, is_suspicious=False, mass_error=False):
        """Выводит статистику по правилу. Всего записей = passed + failed; качество от всего записей."""
        total_records = (total_rows - error_count) + error_count  # всего записей = passed + failed
        if total_records > 0:
            success_rate = ((total_rows - error_count) / total_records * 100)
            error_percent = (error_count / total_records * 100)
        else:
            success_rate = 0
            error_percent = 0
        
        # Цветовая индикация
        if error_count == 0:
            color = "\033[92m"  # зеленый
            status = "[OK] УСПЕШНО"
        elif mass_error:
            color = "\033[91m"  # красный
            status = "[!] МАССОВЫЕ"
        elif is_suspicious:
            color = "\033[93m"  # желтый
            status = "[!] ПОДОЗР."
        else:
            color = "\033[91m"  # красный
            status = "[!] ОШИБКИ"
        
        # Форматирование (total_rows = только оценённые по правилу, без skip)
        print(f"\r    {color}{status}\033[0m {rule_code:20} | "
              f"Оценено: {total_rows:8,} | "
              f"Успех: {success_rate:6.1f}% | "
              f"Ошибок: {error_count:8,} ({error_percent:5.1f}%) | "
              f"Время: {exec_time:6.2f}с")
    
    def _print_table_header(self, table_name, rule_count, row_count):
        """Выводит заголовок для таблицы"""
        print(f"\n{'='*100}")
        print(f"ТАБЛИЦА: \033[1m{table_name}\033[0m")
        print(f"  Правил: {rule_count:3d} | Строк: {row_count:,} | Начало: {datetime.now().strftime('%H:%M:%S')}")
        print(f"{'-'*100}")
    
    def _print_table_summary(self, table_name, elapsed_time, success_count, error_count, suspicious_count):
        """Выводит итог по таблице"""
        total_rules = success_count + error_count + suspicious_count
        success_percent = (success_count / total_rules * 100) if total_rules > 0 else 0
        
        print(f"{'-'*100}")
        print(f"ИТОГ ТАБЛИЦЫ \033[1m{table_name}\033[0m:")
        print(f"  Всего правил: {total_rules:3d} | Время: {elapsed_time:.2f}с")
        print(f"  [OK] Успешно:    {success_count:3d} ({success_percent:.1f}%)")
        print(f"  [!] Ошибки:     {error_count:3d}")
        print(f"  [!] Подозрительные: {suspicious_count:3d}")
        print(f"{'='*100}\n")
    
    def _load_table_handlers(self, silent=False):
        handlers = {}
        import importlib

        def _log(msg):
            if not silent:
                print(msg)

        try:
            module = importlib.import_module('table_scripts.but000_handler')
            if hasattr(module, 'BUT000Handler'):
                handlers['BUT000'] = module.BUT000Handler
                _log(f"   [INFO] Загружен обработчик BUT000")
        except ImportError as e:
            _log(f"   [WARN] Не удалось загрузить обработчик BUT000: {e}")

        try:
            module = importlib.import_module('table_scripts.adrc_handler')
            if hasattr(module, 'ADRCHandler'):
                handlers['ADRC'] = module.ADRCHandler
                _log(f"   [INFO] Загружен обработчик ADRC")
        except ImportError as e:
            _log(f"   [WARN] Не удалось загрузить обработчик ADRC: {e}")

        try:
            module = importlib.import_module('table_scripts.kna1_handler')
            if hasattr(module, 'KNA1Handler'):
                handlers['KNA1'] = module.KNA1Handler
                _log(f"   [INFO] Загружен обработчик KNA1")
        except ImportError as e:
            _log(f"   [WARN] Не удалось загрузить обработчик KNA1: {e}")

        try:
            module = importlib.import_module('table_scripts.taxnum_handler')
            if hasattr(module, 'TaxNumHandler'):
                handlers['DFKKBPTAXNUM'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM1'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM2'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM3'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM4'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM5'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM6'] = module.TaxNumHandler
                _log(f"   [INFO] Загружен обработчик TaxNumHandler")
        except ImportError as e:
            _log(f"   [WARN] Не удалось загрузить обработчик TaxNumHandler: {e}")

        _log(f"   [INFO] Всего загружено обработчиков: {len(handlers)}")
        return handlers

    def reload_table_handlers(self):
        """
        Перезагрузить модули table_scripts/*_handler с диска и обновить self.table_handlers.
        Нужно для интерактивного режима: без перезапуска процесса подхватываются правки в KNA1Handler и др.
        """
        import importlib
        import sys
        mod_names = (
            'table_scripts.but000_handler',
            'table_scripts.adrc_handler',
            'table_scripts.kna1_handler',
            'table_scripts.taxnum_handler',
        )
        for name in mod_names:
            if name in sys.modules:
                try:
                    importlib.reload(sys.modules[name])
                except Exception as e:
                    print(f"   [WARN] reload {name}: {e}")
        self.table_handlers = self._load_table_handlers(silent=True)
        print(f"   [INFO] Обработчики таблиц обновлены с диска ({len(self.table_handlers)} шт.)")

    def _apply_rule_time_column_map(self, df, table_name: str):
        """
        Переименование колонок по column_map.json только на время проверки правил.
        БД и data_cache не изменяются.
        """
        try:
            from utils.column_map_resolver import apply_column_headers_for_rules
            return apply_column_headers_for_rules(
                df,
                table_name,
                self.column_map,
                parent_dir,
                log_renames=True,
            )
        except ImportError:
            return df.copy() if df is not None else df

    def _get_table_for_rules(self, table_name: str):
        """
        Таблица из кэша с шапкой по column_map.json (копия для правил).
        Исходные имена колонок в SQLite не меняются.
        """
        if not hasattr(self, "_rule_time_column_cache"):
            self._rule_time_column_cache = {}
        cache_key = str(table_name or "").strip().upper()
        if cache_key in self._rule_time_column_cache:
            return self._rule_time_column_cache[cache_key].copy()
        raw = self.memory_manager.get_table(table_name)
        if raw is None or raw.empty:
            return raw
        mapped = self._apply_rule_time_column_map(raw, table_name)
        self._rule_time_column_cache[cache_key] = mapped
        return mapped.copy()

    def _get_mapped_column_name(self, table_name, column_name):
        """Получает SAP-имя колонки из column_map.json (логическое -> SAP)."""
        try:
            from utils.column_map_resolver import map_logical_to_sap
            mapped = map_logical_to_sap(table_name, column_name, self.column_map, parent_dir)
            if mapped and mapped != column_name:
                print(f"      [MAP] {table_name}: '{column_name}' -> '{mapped}' (из column_map.json)")
            return mapped
        except ImportError:
            pass
        if not self.column_map:
            return column_name
        table_mapping = None
        if table_name in self.column_map:
            table_mapping = self.column_map[table_name]
        else:
            tn = str(table_name or "").strip().upper()
            for k, v in self.column_map.items():
                if str(k).strip().upper() == tn and isinstance(v, dict):
                    table_mapping = v
                    break
        if not table_mapping:
            return column_name
        for logical_name, real_name in table_mapping.items():
            if str(logical_name).startswith("_"):
                continue
            if str(real_name).upper() == str(column_name or "").upper():
                return column_name
        if column_name in table_mapping:
            return table_mapping[column_name]
        for logical_name, real_name in table_mapping.items():
            if str(logical_name).startswith("_"):
                continue
            if str(logical_name).upper() == str(column_name or "").upper():
                return real_name
        return column_name

    def _resolve_column_for_rule(self, df, column_name, table_name):
        """Единая точка: column_map + физические заголовки выгрузки -> колонка в DataFrame."""
        if not column_name or df is None:
            return None
        try:
            from utils.column_map_resolver import resolve_column_in_df, map_logical_to_sap
            sap_name = map_logical_to_sap(table_name, column_name, self.column_map, parent_dir)
            for target in (sap_name, column_name):
                if not target:
                    continue
                found = resolve_column_in_df(df, target, table_name, self.column_map, parent_dir)
                if found:
                    if target != column_name or found != column_name:
                        print(f"      [MAP] Колонка по column_map: '{column_name}' -> '{found}'")
                    return found
        except ImportError:
            pass
        return self._find_column_alternative(df.columns, column_name, table_name)
    
    def _fix_column_name_for_taxnum(self, column_name, table_name):
        if table_name.startswith('DFKKBPTAXNUM') and len(table_name) > 12 and table_name[12:].isdigit():
            if column_name.upper().startswith('TAXNUM') and len(column_name) > 6 and column_name[6:].isdigit():
                return 'TAXNUM'
        if table_name == 'DFKKBPTAXNUM' and column_name.upper().startswith('TAXNUM'):
            return 'TAXNUM'
        return column_name
    
    def load_configuration(self):
        try:
            with open(self.rules_file, "r", encoding="utf-8") as f:
                rules = json.load(f)
            total_tables = len(rules)
            total_rules = sum(len(rules[table]) for table in rules)
            print(f"\n\033[1m[INFO]\033[0m Загружено {total_tables} таблиц, {total_rules} правил")
            return rules
        except Exception as e:
            self.logger.error(f"Ошибка загрузки конфигурации: {e}")
            return {}
    
    def run_quality_checks_fast(self, specific_table: str = None, table_list: list = None, only_rule_codes: set = None):
        """
        Запускает проверки качества данных.
        specific_table — одна таблица; table_list — список таблиц; оба None — все таблицы.
        only_rule_codes — если задано, выполняются только правила с указанными кодами (изолированный запуск).
        """
        def _filter_rules(rules, codes):
            if not codes:
                return rules
            return [r for r in rules if r.get('rule_code') in codes]

        print(f"\n" + "="*100)
        print(f"\033[1mЗАПУСК ПРОВЕРОК КАЧЕСТВА ДАННЫХ\033[0m")
        print(f"="*100)
        
        self.start_time = time.time()
        self._rule_time_column_cache = {}

        rules_config = self.load_configuration()
        if not rules_config:
            self.logger.error("[ERROR] Не удалось загрузить конфигурацию правил")
            self._save_totals_by_table()
            return pd.DataFrame()
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.results = []
        self.rule_errors = {}
        self.suspicious_rules = []
        self.processed_rules = 0
        self.skipped_rules = 0
        
        if table_list:
            print(f"[INFO] Проверяем выбранные таблицы: {len(table_list)} шт.")
        elif specific_table:
            print(f"[INFO] Проверяем только таблицу: {specific_table}")
        else:
            print(f"[INFO] Проверяем все таблицы")
        if only_rule_codes:
            print(f"[INFO] Изолированный режим: только правила {sorted(only_rule_codes)}")
        
        print(f"[INFO] Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило")
        
        print(f"\n[INFO] Загружаем данные из базы...")
        load_start = time.time()
        
        if hasattr(self.memory_manager, 'load_all_data_to_ram'):
            if specific_table:
                tables_to_load = self._expand_ausp_for_load([specific_table])
                if self.use_async_load and hasattr(self.memory_manager, 'load_selected_tables_to_ram_async_sync'):
                    self.memory_manager.load_selected_tables_to_ram_async_sync(tables_to_load)
                elif hasattr(self.memory_manager, 'load_selected_tables_to_ram'):
                    self.memory_manager.load_selected_tables_to_ram(tables_to_load)
                else:
                    self.memory_manager.load_all_data_to_ram()
            elif table_list:
                tables_to_load = self._expand_ausp_for_load(list(table_list))
                if self.use_async_load and hasattr(self.memory_manager, 'load_selected_tables_to_ram_async_sync'):
                    self.memory_manager.load_selected_tables_to_ram_async_sync(tables_to_load)
                elif hasattr(self.memory_manager, 'load_selected_tables_to_ram'):
                    self.memory_manager.load_selected_tables_to_ram(tables_to_load)
                else:
                    self.memory_manager.load_all_data_to_ram()
            else:
                if self.use_async_load and hasattr(self.memory_manager, 'load_selected_tables_to_ram_async_sync'):
                    all_tables = self.memory_manager._get_all_table_names()
                    self.memory_manager.load_selected_tables_to_ram_async_sync(all_tables, add_reference_tables=True)
                else:
                    self.memory_manager.load_all_data_to_ram()
        
        available_tables = []
        if hasattr(self.memory_manager, 'data_cache'):
            available_tables = list(self.memory_manager.data_cache.keys())
        
        load_time = time.time() - load_start
        
        print(f"   [INFO] Загрузка завершена за {load_time:.2f} сек")
        print(f"   [INFO] Доступно таблиц в памяти: {len(available_tables)}")
        
        if available_tables:
            table_sizes = []
            for table in available_tables[:10]:
                df = self.memory_manager.get_table(table)
                if df is not None:
                    table_sizes.append((table, len(df)))
            
            if table_sizes:
                print(f"\n   [INFO] Размеры таблиц (первые {len(table_sizes)}):")
                for table, size in sorted(table_sizes, key=lambda x: x[1], reverse=True)[:5]:
                    print(f"      {table:25} -> {size:10,} строк")
            # Подсчёт строк по таблицам-алиасам DFKKBPTAXNUM — по каждой в отдельности (там где они используются)
            if "DFKKBPTAXNUM" in self.memory_manager.data_cache:
                print(f"\n   [INFO] Подсчёт строк по таблицам-алиасам DFKKBPTAXNUM (по каждой в отдельности):")
                aliases_to_show = getattr(self.memory_manager, 'DFKKBPTAXNUM_TABLES', ())
                for alias in aliases_to_show:
                    if alias in self.memory_manager.data_cache:
                        alias_df = self.memory_manager.get_table(alias)
                        n = len(alias_df) if alias_df is not None else 0
                        print(f"      {alias:25} -> {n:10,} строк")
        
        if self.parallel_tables:
            print(f"   [INFO] Параллельная обработка таблиц: {self.parallel_tables} потоков")
        print(f"\n\033[1m[INFO]\033[0m Обрабатываем правила:")
        
        if specific_table:
            if specific_table in rules_config:
                table_rules = _filter_rules(rules_config[specific_table], only_rule_codes)
                if not table_rules:
                    print(f"\n[WARN] Для таблицы '{specific_table}' нет правил с кодами {only_rule_codes}. Завершение.")
                    self._save_totals_by_table()
                    return pd.DataFrame()
                self._process_table_rules(specific_table, table_rules, 
                                        available_tables, timestamp)
            elif specific_table == "AUSP":
                for t in self.AUSP_TABLE_GROUP:
                    if t in rules_config and t in available_tables:
                        table_rules = _filter_rules(rules_config[t], only_rule_codes)
                        if table_rules:
                            self._process_table_rules(t, table_rules, available_tables, timestamp)
            elif specific_table == "DFKKBPTAXNUM":
                # Одна таблица DFKKBPTAXNUM в БД — после загрузки создаются постоянные DFKKBPTAXNUM1..6, проверяем по ним.
                # Также здесь же обрабатываем виртуальную таблицу DFKKBPTAXNUM_ALL (правила «по всем taxnum»).
                for alias in self.DFKKBPTAXNUM_ALIASES:
                    if alias in rules_config and alias in available_tables:
                        table_rules = _filter_rules(rules_config[alias], only_rule_codes)
                        if table_rules:
                            self._process_table_rules(alias, table_rules, available_tables, timestamp)
            else:
                print(f"\n[ERROR] В конфигурации нет правил для таблицы '{specific_table}'")
                print(f"   Доступные таблицы: {list(rules_config.keys())}")
                self._save_totals_by_table()
                return pd.DataFrame()
        elif table_list:
            tables_to_process = []
            for t in table_list:
                if t == "AUSP":
                    for a in self.AUSP_TABLE_GROUP:
                        if a in rules_config:
                            tr = _filter_rules(rules_config[a], only_rule_codes)
                            if tr:
                                tables_to_process.append((a, tr))
                elif t in rules_config:
                    tr = _filter_rules(rules_config[t], only_rule_codes)
                    if tr:
                        tables_to_process.append((t, tr))
            missing = [t for t in table_list if t != "AUSP" and t not in rules_config]
            if missing:
                print(f"\n[WARN] Таблицы не найдены в конфигурации: {missing}")
            total_tables = len(tables_to_process)
            print(f"   [INFO] Таблиц для проверки: {total_tables}")
            self._run_tables_loop(tables_to_process, available_tables, timestamp, total_tables)
        else:
            # Полная проверка: все таблицы из конфига, в т.ч. AUSP_143, AUSP_604, AUSP_148, AUSP_151 — отдельно
            tables_to_process = []
            for table_name, table_rules in rules_config.items():
                tr = _filter_rules(table_rules, only_rule_codes)
                if tr:
                    tables_to_process.append((table_name, tr))
            total_tables = len(tables_to_process)
            print(f"   [INFO] Всего таблиц для проверки: {total_tables}")
            self._run_tables_loop(tables_to_process, available_tables, timestamp, total_tables)
        
        overall_time = time.time() - self.start_time
        
        # Создаем единый timestamp для всех файлов в этой сессии
        file_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
        self._print_final_statistics()
        self._save_rule_errors(file_timestamp)
        try:
            self._save_totals_by_table(file_timestamp)
        except Exception as e:
            print(f"\n[ERROR] Ошибка записи в папку total: {e}")
            traceback.print_exc()
        
        report_name = "quality_check_report"
        if specific_table:
            report_name = f"quality_check_report_{self._safe_filename_token(specific_table)}"
        self._create_correct_report(report_name, file_timestamp)
        
        # Экспорт отфильтрованных данных ADR2 (папка quality_reports/adr2_filtered)
        # временно отключён по требованию: не создаём дополнительные adr2_filtered файлы.
        
        print(f"\n" + "="*100)
        print(f"\033[1mПРОВЕРКА ЗАВЕРШЕНА\033[0m")
        print(f"   Общее время: {overall_time:.2f} сек")
        print(f"   Скорость: {self.processed_rules/overall_time:.1f} правил/сек" if overall_time > 0 else "")
        print(f"="*100)

        results_df = pd.DataFrame(self.results)
        return results_df

    def _run_tables_loop(self, tables_to_process, available_tables, timestamp, total_tables):
        """Запускает обработку таблиц: последовательно или параллельно (по self.parallel_tables)."""
        def do_one(item):
            i, (table_name, table_rules) = item
            try:
                if self._parallel_lock:
                    with self._parallel_lock:
                        print(f"\n[ПРОГРЕСС] Таблица {i}/{total_tables}: {table_name}")
                else:
                    print(f"\n[ПРОГРЕСС] Таблица {i}/{total_tables}: {table_name}")
                self._process_table_rules(table_name, table_rules, available_tables, timestamp)
            except Exception as e:
                if self._parallel_lock:
                    with self._parallel_lock:
                        print(f"   \033[91m[ERROR]\033[0m Ошибка при обработке таблицы {table_name}: {str(e)}")
                else:
                    print(f"   \033[91m[ERROR]\033[0m Ошибка при обработке таблицы {table_name}: {str(e)}")
                traceback.print_exc()

        if self.parallel_tables and len(tables_to_process) > 1:
            enumerated = list(enumerate(tables_to_process, 1))
            with ThreadPoolExecutor(max_workers=self.parallel_tables) as executor:
                list(executor.map(do_one, enumerated))
        else:
            for i, (table_name, table_rules) in enumerate(tables_to_process, 1):
                do_one((i, (table_name, table_rules)))
    
    def _process_table_rules(self, table_name, table_rules, available_tables, timestamp):
        # AUSP в интерфейсе одна; внутри обрабатываем по производным таблицам AUSP_143, AUSP_604, AUSP_148, AUSP_151
        if table_name == "AUSP" and table_rules:
            by_table = {}
            for r in table_rules:
                t = r.get("table_name_checked") or r.get("table_name") or ""
                if t not in by_table:
                    by_table[t] = []
                by_table[t].append(r)
            for t in self.AUSP_TABLE_GROUP:
                if t in by_table and by_table[t]:
                    self._process_table_rules(t, by_table[t], available_tables, timestamp)
            return
        self.current_table = table_name
        self.table_start_time = time.time()
        if str(table_name or "").strip().upper() == "KNB1":
            setattr(self, "_kna1_ktokd_lookup_df", None)

        # Таблица считается доступной, если есть в списке или для ADRC — по совпадению имени без учёта регистра
        def _table_available(tname, avail):
            if tname in avail:
                return True
            if str(tname or "").strip().upper() == "ADRC":
                return any(str(t).strip().upper() == "ADRC" for t in avail)
            return False
        if not _table_available(table_name, available_tables):
            print(f"   \033[91m[ERROR]\033[0m Таблица '{table_name}' НЕ НАЙДЕНА в БД!")
            print(f"   [DEBUG] Доступные таблицы в БД: {sorted(available_tables)[:10]}{'...' if len(available_tables) > 10 else ''}")
            print(f"   [DEBUG] Всего таблиц в БД: {len(available_tables)}")
            # Проверяем, может быть таблица есть, но с другим именем (например, с экранированием)
            similar_tables = [t for t in available_tables if table_name.replace('/', '_') in t or t.replace('/', '_') == table_name.replace('/', '_')]
            if similar_tables:
                print(f"   [DEBUG] Найдены похожие таблицы: {similar_tables}")
            if self._parallel_lock:
                with self._parallel_lock:
                    for _ in table_rules:
                        self.skipped_rules += 1
                    for rule in table_rules:
                        self._log_skipped_rule(rule, table_name, "Таблица не найдена в БД", timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, "Таблица не найдена в БД", timestamp)
            return
        
        df_raw = self.memory_manager.get_table(table_name)
        if df_raw is None or df_raw.empty:
            # Для производных AUSP_143 и т.д. уточняем причину: срез по ATINN пуст
            skip_reason = "Таблица пуста"
            if table_name in self.AUSP_TABLE_GROUP:
                atinn_val = table_name.replace("AUSP_", "")
                skip_reason = f"Таблица пуста (нет строк с ATINN={atinn_val} в AUSP)"
            print(f"   \033[93m[WARN]\033[0m Таблица {table_name} пуста! Пропускаем...")
            if self._parallel_lock:
                with self._parallel_lock:
                    for rule in table_rules:
                        self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, skip_reason, timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, skip_reason, timestamp)
            return

        # Шапка из выгрузки → SAP-имена для rules.json (копия; в БД имена из Excel/CSV)
        df = self._get_table_for_rules(table_name)
        if df is None or df.empty:
            print(f"   \033[93m[WARN]\033[0m Таблица {table_name} пуста после маппинга колонок. Пропускаем...")
            for rule in table_rules:
                self.skipped_rules += 1
                self._log_skipped_rule(rule, table_name, "Таблица пуста после маппинга колонок", timestamp)
            return
        
        # ADRC: исключаем клиентов с NAME1 = RESERVED (не учитываем в правилах, в ошибках и в общем подсчёте)
        if (table_name or "").strip().upper() == "ADRC":
            name1_col = None
            for c in df.columns:
                if str(c).strip().upper() == "NAME1":
                    name1_col = c
                    break
            if name1_col is None:
                name1_col = self._find_column_alternative(df.columns, "NAME1", table_name)
            if name1_col is None:
                name1_col = self._find_most_similar_column(df.columns, "NAME1")
            if name1_col and name1_col in df.columns:
                before = len(df)
                val_str = df[name1_col].astype(str).str.strip().str.upper()
                df = df[val_str != "RESERVED"].copy()
                dropped = before - len(df)
                if dropped > 0:
                    print(f"   [ADRC] Исключены строки с NAME1=RESERVED: {dropped:,} (осталось {len(df):,})")
                if df.empty:
                    print(f"   \033[93m[WARN]\033[0m После фильтра NAME1!=RESERVED таблица ADRC пуста. Пропускаем...")
                    if self._parallel_lock:
                        with self._parallel_lock:
                            for rule in table_rules:
                                self.skipped_rules += 1
                                self._log_skipped_rule(rule, table_name, "Нет строк после исключения NAME1=RESERVED", timestamp)
                    else:
                        for rule in table_rules:
                            self.skipped_rules += 1
                            self._log_skipped_rule(rule, table_name, "Нет строк после исключения NAME1=RESERVED", timestamp)
                    return
            else:
                print(f"   [WARN] В ADRC не найдена колонка NAME1, фильтр RESERVED не применён (колонки: {list(df.columns)[:15]})")
        
        # Удаление полностью пустых строк (как «хвост» пустых строк в Excel): не считаем их строками
        # таблицы для DQ — проверять нечего, иначе раздуваются total_rows и ложные срабатывания.
        before_empty_filter = len(df)
        empty_mask = fully_empty_rows_mask(df)
        df = df.loc[~empty_mask].copy()
        empty_rows_dropped = before_empty_filter - len(df)
        if empty_rows_dropped > 0:
            print(f"   [{table_name}] Удалено полностью пустых строк: {empty_rows_dropped:,} (осталось {len(df):,})")
        
        if df.empty:
            print(f"   \033[93m[WARN]\033[0m После фильтрации пустых строк таблица {table_name} пуста. Пропускаем...")
            if self._parallel_lock:
                with self._parallel_lock:
                    for rule in table_rules:
                        self.skipped_rules += 1
                        self._log_skipped_rule(rule, table_name, "Таблица пуста после фильтрации", timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, "Таблица пуста после фильтрации", timestamp)
            return
        
        # Для ZBUT0000P3VVI9 и аналогов в колонке «Всего записей» показываем количество кастомеров (из отдельной таблицы _partners)
        display_row_count = len(df)
        if table_name in self.TABLE_UNIQUE_PARTNER:
            count = getattr(self.memory_manager, 'get_unique_partner_count', lambda t: None)(table_name)
            if count is not None:
                display_row_count = count
            else:
                partner_col = self._find_partner_column(df, table_name=table_name)
                if partner_col and partner_col in df.columns:
                    display_row_count = int(df[partner_col].nunique())
        
        # Выводим заголовок таблицы
        self._print_table_header(table_name, len(table_rules), display_row_count)
        
        # AUSP: сразу разбиваем таблицу по ATINN в «временные таблицы» (ATINN=143 → CCAF, 604 → Red Outlet и т.д.)
        ausp_split = None
        if (table_name or "").strip().upper() == "AUSP":
            ausp_split = self._build_ausp_split(df, table_name)
            if ausp_split:
                total_slices = sum(len(s[0]) for s in ausp_split.values())
                print(f"   [AUSP] Таблица разбита по ATINN на {len(ausp_split)} срезов (всего строк в срезах: {total_slices:,})")
            else:
                atinn_c, atwrt_c = self._find_ausp_columns(df.columns, table_name)
                names = list(df.columns) if hasattr(df.columns, '__iter__') else []
                print(f"   [AUSP] Не удалось разбить по имени: колонки ATINN/ATWRT не найдены. Заголовки ({len(names)}): {names[:15]}{'...' if len(names) > 15 else ''}")
                self._debug_ausp_columns(df.columns, table_name)
        
        if table_name in self.table_handlers:
            # Для DFKKBPTAXNUM* правила «same value as other» (в одной строке) выполняем стандартным валидатором (other_columns);
            # остальные правила — через TaxNumHandler (формат 63.1 и т.д.)
            # Почему «Всего записей» может сильно отличаться между RCCONF_50.11 и RCCONF_63.7:
            # 1) Разный размер среза: DFKKBPTAXNUM1 (taxtype=1) vs DFKKBPTAXNUM5 (taxtype=5) — разное число строк в таблице.
            # 2) Оценённые строки = только те, где заполнены ВСЕ tax_0..tax_6 (IF any NULL THEN ''). Если в данных
            #    taxnum2/taxnum5 и др. часто NULL, в «оценённые» попадёт меньше строк — отсюда разница в итогах и в ошибках.
            TAXNUM_SAME_ROW_RULES = {"RCCONF_50.11", "RCCONF_52.11", "RCCONF_54.9", "RCCONF_63.7"}
            is_taxnum_table = str(table_name or "").strip().upper().startswith("DFKKBPTAXNUM")
            rule_codes_in_table = {str(r.get("rule_code") or "").strip() for r in (table_rules or []) if r}
            same_row_here = TAXNUM_SAME_ROW_RULES & rule_codes_in_table
            if is_taxnum_table:
                print(f"   [DEBUG] {table_name}: кодов правил в таблице: {len(rule_codes_in_table)}, same_row: {same_row_here}")
            if is_taxnum_table and same_row_here:
                standard_rules = [r for r in table_rules if str(r.get("rule_code") or "").strip() in TAXNUM_SAME_ROW_RULES]
                handler_rules = [r for r in table_rules if str(r.get("rule_code") or "").strip() not in TAXNUM_SAME_ROW_RULES]
                success_count, error_count, suspicious_count = 0, 0, 0
                if standard_rules:
                    print(f"   [INFO] Правила «same value as other» (в одной строке): стандартный метод ({len(standard_rules)} правил)")
                    s, e, sus = self._process_with_standard_method(table_name, df, standard_rules, timestamp, ausp_split=None)
                    success_count += s
                    error_count += e
                    suspicious_count += sus
                if handler_rules:
                    print(f"   [INFO] Остальные правила DFKKBPTAXNUM: специальный обработчик ({len(handler_rules)} правил)")
                    s, e, sus = self._process_with_table_handler(table_name, df, handler_rules, timestamp)
                    success_count += s
                    error_count += e
                    suspicious_count += sus
            else:
                print(f"   [INFO] Используем специальный обработчик")
                success_count, error_count, suspicious_count = self._process_with_table_handler(table_name, df, table_rules, timestamp)
        else:
            print(f"   [INFO] Используем стандартный метод проверки")
            success_count, error_count, suspicious_count = self._process_with_standard_method(table_name, df, table_rules, timestamp, ausp_split=ausp_split)
        
        # Выводим итог по таблице
        elapsed_time = time.time() - self.table_start_time
        self._print_table_summary(table_name, elapsed_time, success_count, error_count, suspicious_count)
        
        return success_count, error_count, suspicious_count
    
    def _process_with_table_handler(self, table_name, df, table_rules, timestamp):
        handler_class = self.table_handlers[table_name]
        success_count = 0
        error_count = 0
        suspicious_count = 0
        
        try:
            handler = handler_class(table_name, df, self.memory_manager, self)
            
            total_rules = len(table_rules)
            for i, rule in enumerate(table_rules, 1):
                if self._parallel_lock:
                    with self._parallel_lock:
                        self.processed_rules += 1
                else:
                    self.processed_rules += 1
                self.current_rule = rule.get('rule_code', 'UNKNOWN')
                
                # Показываем прогресс
                sys.stdout.write(f"\r    [{i:3d}/{total_rules:3d}] {self.current_rule:20} | ")
                sys.stdout.flush()
                
                rule_start_time = time.time()
                
                # Выполняем проверку
                result = handler.validate_rule(rule)
                
                execution_time = time.time() - rule_start_time
                
                if result and isinstance(result, dict):
                    error_df_res = result.get('error_df', pd.DataFrame())
                    error_count_result = int(result.get('error_count', result.get('failed', 0)))
                    failed = int(result.get('failed', error_count_result))
                    passed = int(result.get('passed', 0))
                    # Только строки в scope правила (passed + failed). Не len(df) и не размер таблицы.
                    if result.get('total_records') is not None:
                        total_rows = int(result['total_records'])
                    elif result.get('total_evaluated') is not None:
                        total_rows = int(result['total_evaluated'])
                    else:
                        total_rows = passed + failed
                    if passed + failed != total_rows and total_rows > 0:
                        passed = max(total_rows - failed, 0)
                    is_suspicious = self._check_if_suspicious(self.current_rule, error_count_result, total_rows)
                    mass_error = error_count_result > self.MAX_ERRORS_TO_SAVE
                    
                    # Обновляем счетчики (локальные)
                    if error_count_result == 0:
                        success_count += 1
                    elif is_suspicious or mass_error:
                        suspicious_count += 1
                    else:
                        error_count += 1
                    
                    # Выводим статистику
                    self._print_rule_stats(self.current_rule, total_rows, error_count_result, 
                                         execution_time, is_suspicious, mass_error)
                    
                    # Сохраняем результат и ошибки (общее состояние — под блокировкой при параллелизме)
                    result['check_date'] = timestamp
                    result['passed'] = passed
                    result['failed'] = failed
                    result['total_records'] = total_rows
                    result['total_evaluated'] = total_rows
                    result['error_count'] = error_count_result
                    # total_records / total_evaluated: только scope правила, без IF ... THEN '' skip
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self.results.append(result)
                            if error_count_result > 0:
                                key = f"{self.current_rule}_{table_name}"
                                error_df = result.get('error_df', pd.DataFrame())
                                if error_df is not None and not error_df.empty:
                                    if len(error_df) > error_count_result * 1.1:
                                        error_df = error_df.head(error_count_result)
                                    self._save_rule_error_with_limit(
                                        self.current_rule, table_name, error_df,
                                        error_count_result, is_suspicious, total_rows
                                    )
                                else:
                                    self.rule_errors[key] = {
                                        'rule_code': self.current_rule,
                                        'table_name': table_name,
                                        'error_df': pd.DataFrame(),
                                        'error_count': error_count_result,
                                        'is_suspicious': is_suspicious,
                                        'total_rows': total_rows
                                    }
                    else:
                        self.results.append(result)
                        if error_count_result > 0:
                            key = f"{self.current_rule}_{table_name}"
                            error_df = result.get('error_df', pd.DataFrame())
                            if error_df is not None and not error_df.empty:
                                if len(error_df) > error_count_result * 1.1:
                                    error_df = error_df.head(error_count_result)
                                self._save_rule_error_with_limit(
                                    self.current_rule, table_name, error_df,
                                    error_count_result, is_suspicious, total_rows
                                )
                            else:
                                self.rule_errors[key] = {
                                    'rule_code': self.current_rule,
                                    'table_name': table_name,
                                    'error_df': pd.DataFrame(),
                                    'error_count': error_count_result,
                                    'is_suspicious': is_suspicious,
                                    'total_rows': total_rows
                                }
                else:
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self.skipped_rules += 1
                            self._log_skipped_rule(rule, table_name, "Обработчик вернул неверный формат", timestamp)
                    else:
                        self.skipped_rules += 1
                        self._log_skipped_rule(rule, table_name, "Обработчик вернул неверный формат", timestamp)
                
        except Exception as e:
            print(f"\n   \033[91m[ERROR]\033[0m Ошибка в обработчике {table_name}: {str(e)}")
            traceback.print_exc()
            if self._parallel_lock:
                with self._parallel_lock:
                    for rule in table_rules:
                        self.skipped_rules += 1
                        self._log_skipped_rule(rule, table_name, f"Ошибка обработчика: {str(e)}", timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, f"Ошибка обработчика: {str(e)}", timestamp)
        
        return success_count, error_count, suspicious_count
    
    def _process_with_standard_method(self, table_name, df, table_rules, timestamp, ausp_split=None):
        success_count = 0
        error_count = 0
        suspicious_count = 0
        
        total_rules = len(table_rules)
        for i, rule in enumerate(table_rules, 1):
            if self._parallel_lock:
                with self._parallel_lock:
                    self.processed_rules += 1
            else:
                self.processed_rules += 1
            self.current_rule = rule.get('rule_code', 'UNKNOWN')
            
            # Показываем прогресс
            sys.stdout.write(f"\r    [{i:3d}/{total_rules:3d}] {self.current_rule:20} | ")
            sys.stdout.flush()
            
            rule_start_time = time.time()
            
            # Обрабатываем правило (для AUSP передаём ausp_split — подставляем уже готовый срез и имя колонки)
            error_count_result, total_rows = self._process_single_rule(rule, table_name, df, timestamp, ausp_split=ausp_split)
            
            execution_time = time.time() - rule_start_time
            
            is_suspicious = self._check_if_suspicious(self.current_rule, error_count_result, total_rows)
            mass_error = error_count_result > self.MAX_ERRORS_TO_SAVE
            
            # Обновляем счетчики
            if error_count_result == 0:
                success_count += 1
            elif is_suspicious or mass_error:
                suspicious_count += 1
            else:
                error_count += 1
            
            # Выводим статистику
            self._print_rule_stats(self.current_rule, total_rows, error_count_result, 
                                 execution_time, is_suspicious, mass_error)
        
        return success_count, error_count, suspicious_count
    
    def _process_single_rule_without_save(self, rule, table_name, df, timestamp):
        """Обрабатывает одно правило БЕЗ сохранения результата в self.results"""
        return self._process_single_rule(rule, table_name, df, timestamp, save_result=False)
    
    def _process_single_rule(self, rule, table_name, df, timestamp, save_result=True, ausp_split=None):
        import re  # привязка в начале функции, чтобы избежать UnboundLocalError при использовании re выше по коду
        rule_code_raw = str(rule.get("rule_code", "UNKNOWN"))
        # Нормализуем код правила, чтобы исключить скрытые символы/мусор из Excel/JSON.
        rule_code = re.sub(r"[^A-Za-z0-9._-]", "", rule_code_raw).upper()
        rule_description = rule.get("rule_description", "Unknown rule")
        rule_desc_lower = str(rule_description).lower()
        quality_category = rule.get("quality_category", "Unknown")
        column_to_check = rule.get("column_name_checked", "")
        value_checked = rule.get("value_checked", "")

        # RCCONF_63.1 — только через TaxNumHandler (длины из conf_tax_number_format)
        tn_tax = str(table_name or "").strip().upper()
        if rule_code == "RCCONF_63.1" and tn_tax.startswith("DFKKBPTAXNUM"):
            handler_cls = self.table_handlers.get(table_name) or self.table_handlers.get(tn_tax)
            if handler_cls is not None:
                try:
                    h = handler_cls(table_name, df, self.memory_manager, self)
                    result = h.validate_rule(rule)
                    if result and isinstance(result, dict):
                        err = int(result.get("failed", result.get("error_count", 0)))
                        tot = int(result.get("total_records", 0))
                        if save_result:
                            result["check_date"] = timestamp
                            result["table_name"] = table_name
                            self.results.append(result)
                            if err > 0 and result.get("error_df") is not None:
                                self._save_rule_error_with_limit(
                                    rule_code, table_name, result["error_df"], err,
                                    self._check_if_suspicious(rule_code, err, tot), tot,
                                )
                        return err, tot
                except Exception as e:
                    print(f"      [WARN] RCCONF_63.1 TaxNumHandler: {e}")
        
        # AUSP: если передано разбиение по ATINN — сразу подставляем готовый срез и имя колонки (CCAF, Red Outlet и т.д.)
        matched_column = None
        actual_column_to_check = column_to_check  # по умолчанию для rule_info
        if (table_name or "").strip().upper() == "AUSP" and ausp_split:
            atinn_value = self._resolve_ausp_atinn_value(rule)
            if atinn_value and atinn_value in ausp_split:
                df_slice, col_name = ausp_split[atinn_value]
                df = df_slice
                matched_column = col_name
                actual_column_to_check = col_name
                if df.empty:
                    self._log_skipped_rule(rule, table_name, f"Нет строк с ATINN={atinn_value} ({column_to_check})", timestamp)
                    return 0, 0
                print(f"      [AUSP] Правило → срез ATINN={atinn_value}, колонка '{col_name}', строк: {len(df):,}")
        
        if not matched_column:
            # Определяем реальную колонку для проверки
            actual_column_to_check = None
            table_name_norm = str(table_name or "").strip().upper()
            if table_name_norm == "KNB5" and column_to_check == "MAHNA":
                actual_column_to_check = column_to_check
                print(f"      [MAP] KNB5: используем column_name_checked '{column_to_check}' напрямую (альтернативы через _find_column_alternative)")
            elif table_name_norm == "KNB1" and column_to_check in ("AKONT", "FDGRV", "ZTERM"):
                actual_column_to_check = column_to_check
                print(f"      [MAP] KNB1: используем column_name_checked '{column_to_check}' (AKONT/FDGRV/ZTERM), альтернативы через _find_column_alternative)")
            elif value_checked and self.column_map:
                table_mapping = None
                tn = str(table_name or "").strip().upper()
                if table_name in self.column_map:
                    table_mapping = self.column_map[table_name]
                else:
                    for k, v in self.column_map.items():
                        if str(k).strip().upper() == tn and isinstance(v, dict):
                            table_mapping = v
                            break
                if table_mapping:
                    logical_keys = sorted(
                        (k for k in table_mapping.keys() if not str(k).startswith("_")),
                        key=lambda x: len(str(x)),
                        reverse=True,
                    )
                    for logical_name in logical_keys:
                        pattern = rf'(?:^|[\.\s\+\-_])({re.escape(logical_name)})(?:[\s\+\-_\.]|$)'
                        if re.search(pattern, value_checked, re.IGNORECASE):
                            actual_column_to_check = table_mapping[logical_name]
                            print(f"      [MAP] Найдено логическое имя в value_checked: '{logical_name}' -> '{actual_column_to_check}'")
                            break
            
            if not actual_column_to_check:
                actual_column_to_check = self._get_mapped_column_name(table_name, column_to_check)
            actual_column_to_check = self._fix_column_name_for_taxnum(actual_column_to_check, table_name)
            
            print(f"      [COL] Ищем колонку: '{column_to_check}' (value_checked: '{value_checked}') -> '{actual_column_to_check}'")
            
            matched_column = None
            # AUSP без предразбиения: фильтр по ATINN и переименование ATWRT (если не передали ausp_split)
            ausp_filtered, ausp_atwrt_col, ausp_temporary_name = self._apply_ausp_filter(df, column_to_check, table_name, rule=rule)
            if ausp_filtered is not None and ausp_atwrt_col is not None and ausp_temporary_name:
                if ausp_filtered.empty:
                    atinn_val = None
                    if self.ausp_atinn_mapping:
                        for k, v in self.ausp_atinn_mapping.items():
                            if (k or "").strip().upper() == (column_to_check or "").strip().upper():
                                atinn_val = v
                                break
                    if atinn_val is None and column_to_check:
                        m = re.search(r"ATINN\s*[=\(]\s*(\d+)", (column_to_check or ""), re.IGNORECASE)
                        if m:
                            atinn_val = m.group(1)
                    self._log_skipped_rule(rule, table_name, f"Нет строк с ATINN={atinn_val} ({column_to_check})", timestamp)
                    return 0, 0
                df = ausp_filtered.copy()
                if ausp_atwrt_col in df.columns and ausp_temporary_name != ausp_atwrt_col:
                    df = df.rename(columns={ausp_atwrt_col: ausp_temporary_name})
                matched_column = ausp_temporary_name
                print(f"      [AUSP] ATINN отфильтрован, колонка ATWRT временно переименована в '{ausp_temporary_name}', строк: {len(df):,}")
            
            # Для AUSP без предразбиения: не ищем колонку по column_name_checked — только ATINN и ATWRT
            if (table_name or "").strip().upper() == "AUSP" and not matched_column:
                atinn_col, atwrt_col = self._find_ausp_columns(df.columns, table_name)
                if not atinn_col or not atwrt_col:
                    self._log_skipped_rule(rule, table_name, "В таблице AUSP не найдены колонки ATINN или ATWRT", timestamp)
                    return 0, 0
                self._log_skipped_rule(rule, table_name, f"Для правила не определено значение ATINN (column_name_checked='{column_to_check}'). Добавьте запись в conf_ausp_atinn_mapping.json", timestamp)
                return 0, 0
            
            if not matched_column:
                matched_column = self._resolve_column_for_rule(df, actual_column_to_check, table_name)

            if not matched_column and actual_column_to_check != column_to_check:
                matched_column = self._resolve_column_for_rule(df, column_to_check, table_name)

            if not matched_column:
                matched_column = self._find_most_similar_column(df.columns, actual_column_to_check)
                if not matched_column:
                    self._log_skipped_rule(rule, table_name, f"Колонка '{actual_column_to_check}' не найдена", timestamp)
                    return 0, 0
        
        rule_info = {
            'table_name': table_name,
            'rule_code': rule_code,
            'rule_description': rule_description,
            'quality_category': quality_category,
            'matched_column': matched_column,
            'original_column': column_to_check,
            'actual_column_searched': actual_column_to_check
        }
        
        validator = self._get_validator_for_rule(rule_description, quality_category, rule_info)
        is_recon_1131 = (
            self._normalize_rule_code(rule_code) == "RCCONF_113.1"
            or ("recon" in rule_desc_lower and "account group" in rule_desc_lower)
        )
        # Защита от ошибочного выбора generic cross-column валидатора для RCCONF_119.2
        if rule_code == "RCCONF_119.2" and not isinstance(validator, PaymentTermsConsistencyValidator):
            print(
                f"      [WARN] RCCONF_119.2: выбран {validator.__class__.__name__}, "
                "принудительно переключаем на PaymentTermsConsistencyValidator"
            )
            validator = PaymentTermsConsistencyValidator(rule_info)
        # Защита от ошибочного выбора generic валидатора для RCCONF_113.1
        if is_recon_1131 and not isinstance(validator, ReconAccountConsistencyValidator):
            print(
                f"      [WARN] RCCONF_113.1: выбран {validator.__class__.__name__}, "
                "принудительно переключаем на ReconAccountConsistencyValidator"
            )
            validator = ReconAccountConsistencyValidator(rule_info)
        if rule_code == "RCCONF_119.2":
            print(f"      [DEBUG] RCCONF_119.2: validator={validator.__class__.__name__}")
        if is_recon_1131:
            print(f"      [DEBUG] RCCONF_113.1: validator={validator.__class__.__name__}")
        
        params = {}
        need_second_column = False
        
        # Проверяем, нужна ли вторая колонка для валидатора.
        # RCCONF_119.2 — отдельный сценарий: second_column не используется.
        if rule_code == "RCCONF_119.2" or is_recon_1131:
            need_second_column = False
        elif isinstance(validator, CrossColumnEqualityValidator) or isinstance(validator, CrossColumnEqualityCheckValidator):
            need_second_column = True

        if rule_code == "RCCONF_119.2":
            print(f"      [DEBUG] RCCONF_119.2: need_second_column={need_second_column}")
        if is_recon_1131:
            print(f"      [DEBUG] RCCONF_113.1: need_second_column={need_second_column}")
        
        if need_second_column:
            technical_def = rule.get("technical_definition_RU", "")
            if isinstance(technical_def, list):
                technical_def = " ".join(str(x) for x in technical_def)
            value_checked = rule.get("value_checked", "")
            second_column_candidate = None

            # Явная привязка второй колонки для правил BUT000 (сравнение NAME_ORG*)
            if table_name == "BUT000" and rule_code in ("RCCONF_15.2.1", "RCCONF_15.2.2", "RCCONF_15.2.3", "RCCONF_14.1.1", "RCCONF_14.1.2", "RCCONF_13.2"):
                second_by_rule = {
                    "RCCONF_15.2.1": "NAME_ORG1",  # NAME_ORG4 cannot be the same as Name 1
                    "RCCONF_15.2.2": "NAME_ORG2",  # NAME_ORG4 cannot be the same as Name 2
                    "RCCONF_15.2.3": "NAME_ORG3",  # NAME_ORG4 cannot be the same as Name 3
                    "RCCONF_14.1.1": "NAME_ORG1",  # NAME_ORG3 cannot be the same as Name 1
                    "RCCONF_14.1.2": "NAME_ORG2",  # NAME_ORG3 cannot be the same as Name 2
                    "RCCONF_13.2": "NAME_ORG1",   # NAME_ORG2 cannot be the same as Name 1
                }
                wanted = second_by_rule.get(rule_code)
                if wanted:
                    found = self._resolve_column_for_rule(df, wanted, table_name) or (wanted if wanted in df.columns else None)
                    if not found:
                        for c in df.columns:
                            if c.upper() == wanted.upper():
                                found = c
                                break
                    if found:
                        params['second_column'] = found
                        print(f"      [COL] Вторая колонка по коду правила {rule_code}: '{found}'")
            
            # Правила TAXNUM «same value as other»: одна колонка (tax_X) сравнивается со ВСЕМИ остальными tax_0..tax_6 — передаём other_columns
            if not params.get('second_column') and (str(table_name or "").strip().upper().startswith("DFKKBPTAXNUM")):
                is_tax_same_as_other = (
                    rule_code in ("RCCONF_50.11", "RCCONF_52.11", "RCCONF_54.9", "RCCONF_63.7")
                    or (technical_def and ("tax_0_value OR" in technical_def or "= tax_0_value OR" in technical_def or "= tax_1_value OR" in technical_def))
                )
                if is_tax_same_as_other and matched_column:
                    # Собираем все колонки, похожие на tax_* (разные схемы: tax_1_value, TAXNUM1 и т.д.)
                    tax_like = [c for c in df.columns if c != matched_column and (
                        re.match(r'^tax_\d+_value$', str(c), re.I) or
                        re.match(r'^tax_\d+$', str(c), re.I) or
                        (str(c).upper().startswith('TAXNUM') and str(c).upper() != matched_column.upper())
                    )]
                    # Fallback: явно ищем tax_0_value..tax_6_value (какие есть в таблице), кроме проверяемой — чтобы логика была одинаковой для всех TAXNUM
                    if not tax_like:
                        col_upper = {str(c).strip().upper(): c for c in df.columns}
                        for i in range(0, 7):
                            for cand in (f'tax_{i}_value', f'tax_{i}', f'TAXNUM{i}', f'TAXNUM_{i}'):
                                if cand.upper() in col_upper:
                                    c = col_upper[cand.upper()]
                                    if c != matched_column and c not in tax_like:
                                        tax_like.append(c)
                                    break
                            else:
                                for c in df.columns:
                                    cu = str(c).upper().replace(' ', '').replace('_', '')
                                    if (f'TAX{i}VALUE' in cu or f'TAXNUM{i}' in cu) and c != matched_column and c not in tax_like:
                                        tax_like.append(c)
                                        break
                    if tax_like:
                        params['other_columns'] = tax_like
                        print(f"      [COL] TAXNUM «same as other»: проверяем '{matched_column}' против {len(tax_like)} колонок: {[str(c) for c in tax_like[:5]]}{'...' if len(tax_like) > 5 else ''}")
            
            first_num = None
            if value_checked:
                first_match = re.search(r'organization_(\d+)_name', value_checked)
                if first_match:
                    first_num = first_match.group(1)
            
            if not params.get('second_column') and technical_def:
                # Сначала проверяем прямые имена колонок NAME_ORG1, NAME_ORG2 и т.д.
                name_org_pattern = r'NAME_ORG(\d+)'
                name_org_matches = re.findall(name_org_pattern, technical_def, re.IGNORECASE)
                
                if name_org_matches:
                    # Ищем сравнение вида NAME_ORG1 = NAME_ORG2
                    name_org_comparison = re.search(r'NAME_ORG(\d+)\s*=\s*NAME_ORG(\d+)', technical_def, re.IGNORECASE)
                    if name_org_comparison:
                        col1_num = name_org_comparison.group(1)
                        col2_num = name_org_comparison.group(2)
                        # Определяем, какая колонка соответствует matched_column
                        matched_num = None
                        if matched_column and 'ORG' in matched_column.upper():
                            matched_num_match = re.search(r'ORG(\d+)', matched_column.upper())
                            if matched_num_match:
                                matched_num = matched_num_match.group(1)
                        
                        if matched_num == col1_num:
                            second_col_name = f"NAME_ORG{col2_num}"
                        elif matched_num == col2_num:
                            second_col_name = f"NAME_ORG{col1_num}"
                        else:
                            # Если не нашли соответствие, берём ту, что не равна matched_column
                            if matched_column and matched_column.upper() == f"NAME_ORG{col1_num}":
                                second_col_name = f"NAME_ORG{col2_num}"
                            else:
                                second_col_name = f"NAME_ORG{col1_num}"
                        
                        if second_col_name in df.columns:
                            params['second_column'] = second_col_name
                            print(f"      [COL] Вторая колонка из technical_definition (прямое имя): '{second_col_name}'")
                    
                    # Если не нашли через сравнение, ищем по списку найденных
                    if not params.get('second_column'):
                        matched_num = None
                        if matched_column and 'ORG' in matched_column.upper():
                            matched_num_match = re.search(r'ORG(\d+)', matched_column.upper())
                            if matched_num_match:
                                matched_num = matched_num_match.group(1)
                        
                        for num in name_org_matches:
                            if matched_num and num != matched_num:
                                candidate = f"NAME_ORG{num}"
                                if candidate in df.columns:
                                    params['second_column'] = candidate
                                    print(f"      [COL] Вторая колонка из technical_definition (прямое имя): '{candidate}'")
                                    break
                
                # Если не нашли через NAME_ORG, пробуем старый паттерн organization_X_name
                if not params.get('second_column'):
                    pattern = r'organization_(\d+)_name'
                    matches = re.findall(pattern, technical_def)
                    
                    if matches:
                        comparison_pattern = r'organization_(\d+)_name\s*=\s*organization_(\d+)_name'
                        comparison_match = re.search(comparison_pattern, technical_def)
                        
                        if comparison_match:
                            col1_num = comparison_match.group(1)
                            col2_num = comparison_match.group(2)
                            
                            if col1_num == first_num:
                                second_num = col2_num
                            elif col2_num == first_num:
                                second_num = col1_num
                            else:
                                second_num = col2_num
                            
                            second_logical = f"organization_{second_num}_name"
                            second_column_candidate = self._get_mapped_column_name(table_name, second_logical)
                            resolved = self._resolve_column_for_rule(df, second_column_candidate, table_name)
                            if not resolved and second_column_candidate:
                                for c in df.columns:
                                    if c.upper() == second_column_candidate.upper():
                                        resolved = c
                                        break
                            if resolved:
                                params['second_column'] = resolved
                                print(f"      [COL] Вторая колонка из сравнения в technical_definition: '{second_logical}' -> '{resolved}'")
                        
                        if not params.get('second_column'):
                            for match_num in matches:
                                if match_num != first_num:
                                    second_logical = f"organization_{match_num}_name"
                                    second_column_candidate = self._get_mapped_column_name(table_name, second_logical)
                                    resolved = self._resolve_column_for_rule(df, second_column_candidate, table_name)
                                    if not resolved and second_column_candidate:
                                        for c in df.columns:
                                            if c.upper() == second_column_candidate.upper():
                                                resolved = c
                                                break
                                    if resolved:
                                        params['second_column'] = resolved
                                        print(f"      [COL] Вторая колонка из technical_definition: '{second_logical}' -> '{resolved}'")
                                        break
                    
                    if not params.get('second_column'):
                        desc_lower = rule_description.lower()
                        second_logical = None
                        
                        # Сначала проверяем паттерны "cannot be the same as Name X"
                        if "cannot be the same as name 2" in desc_lower or "cannot be the same as name2" in desc_lower:
                            second_logical = "organization_2_name"
                        elif "cannot be the same as name 3" in desc_lower or "cannot be the same as name3" in desc_lower:
                            second_logical = "organization_3_name"
                        elif "cannot be the same as name 4" in desc_lower or "cannot be the same as name4" in desc_lower:
                            second_logical = "organization_4_name"
                        elif "cannot be the same as name 1" in desc_lower or "cannot be the same as name1" in desc_lower:
                            second_logical = "organization_1_name"
                        # Затем проверяем паттерны "equals name X"
                        elif "equals name 2" in desc_lower or "equals name2" in desc_lower:
                            second_logical = "organization_2_name"
                        elif "equals name 3" in desc_lower or "equals name3" in desc_lower:
                            second_logical = "organization_3_name"
                        elif "equals name 4" in desc_lower or "equals name4" in desc_lower:
                            second_logical = "organization_4_name"
                        elif "equals name 1" in desc_lower or "equals name1" in desc_lower:
                            second_logical = "organization_1_name"
                        # Затем общие паттерны "name X"
                        elif ("name 2" in desc_lower or "name2" in desc_lower) and first_num != "2":
                            second_logical = "organization_2_name"
                        elif ("name 3" in desc_lower or "name3" in desc_lower) and first_num != "3":
                            second_logical = "organization_3_name"
                        elif ("name 4" in desc_lower or "name4" in desc_lower) and first_num != "4":
                            second_logical = "organization_4_name"
                        elif ("name 1" in desc_lower or "name1" in desc_lower) and first_num != "1":
                            second_logical = "organization_1_name"
                        
                        if second_logical:
                            second_column_candidate = self._get_mapped_column_name(table_name, second_logical)
                            resolved = self._resolve_column_for_rule(df, second_column_candidate, table_name)
                            if not resolved and second_column_candidate:
                                for c in df.columns:
                                    if c.upper() == second_column_candidate.upper():
                                        resolved = c
                                        break
                            if resolved:
                                params['second_column'] = resolved
                                print(f"      [COL] Вторая колонка из описания правила: '{second_logical}' -> '{resolved}'")
            
            if not params.get('second_column'):
                second_column = self._extract_second_column_from_description(
                    rule_code, rule_description, df.columns, matched_column, table_name
                )
                if second_column and second_column in df.columns:
                    params['second_column'] = second_column
                    print(f"      [COL] Вторая колонка из описания: '{second_column}'")
            
            if not params.get('second_column') and not params.get('other_columns'):
                self._log_failed_rule(rule, table_name, 
                    f"Не найдена вторая колонка для сравнения в правиле '{rule_description}'", 
                    timestamp)
                return 0, 0
        
        # Для правил с условиями фильтруем данные (используем view если возможно, копию только при необходимости)
        # Для AUSP df уже отфильтрован по ATINN (поле для проверки — ATWRT, например ZGLOBAL_CUSTOMER → ATINN=148)
        df_to_validate = df
        # ADRC: скипаем строки с NAME1=RESERVED так же, как скипаем незаполненные ячейки — они не входят в оценку правила
        if (table_name or "").strip().upper() == "ADRC":
            name1_col = None
            for c in df_to_validate.columns:
                if str(c).strip().upper() == "NAME1":
                    name1_col = c
                    break
            if name1_col is None:
                name1_col = self._find_column_alternative(df_to_validate.columns, "NAME1", table_name)
            if name1_col and name1_col in df_to_validate.columns:
                reserved_mask = df_to_validate[name1_col].astype(str).str.strip().str.upper() == "RESERVED"
                if reserved_mask.any():
                    before_adrc = len(df_to_validate)
                    df_to_validate = df_to_validate[~reserved_mask].copy()
                    print(f"      [FILTER] ADRC: скип строк с NAME1=RESERVED (как пустые) — {reserved_mask.sum():,} исключено, к оценке {len(df_to_validate):,} из {before_adrc:,}")
                if df_to_validate.empty:
                    self._log_skipped_rule(rule, table_name, "Нет данных после исключения NAME1=RESERVED", timestamp)
                    return 0, 0
        technical_def = rule.get("technical_definition_RU", "")
        if isinstance(technical_def, list):
            technical_def = " ".join(str(x) for x in technical_def)
        
        # Применяем фильтрацию для правил 38.3, 38.5, 39.3, а также для правил с contact_medium_type или source
        # Также для правил BUT050 (RCCOMP_369.1, RCCONF_369.1) и ADR2 (RCCOMP_375.1)
        if technical_def and ("contact_medium_type" in technical_def or "source" in technical_def or 
                              rule_code in ["RCCONF_38.3", "RCCONF_38.5", "RCCONF_39.3", "RCCONF_39.3.2", "RCCONF_39.5", "RCCONF_39.5.2", "RCCOMP_369.1", "RCCONF_369.1", "RCCOMP_375.1", "RCCOMP_375.1.2"]):
            df_to_validate = self._apply_conditional_filter(df_to_validate, technical_def, rule_code, table_name)
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, "Нет данных, соответствующих условиям правила", timestamp)
                return 0, 0

        # RCCOMP_180.1 (BUT0BK):
        # BANKS проверяем только на RU и только для партнёров из KNA1 (KUNNR).
        if rule_code == "RCCOMP_180.1" and str(table_name or "").strip().upper() == "BUT0BK":
            before_cnt = len(df_to_validate)
            df_to_validate = self._scope_but0bk_to_kna1_partners(df_to_validate, table_name, rule_code)
            print(f"      [FILTER] {rule_code}: scope партнёры из KNA1 -> {len(df_to_validate):,} из {before_cnt:,}")
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, "Нет строк BUT0BK для партнёров из KNA1", timestamp)
                return 0, 0
            # Бизнес-правило: учитываем только BANKS='RU'.
            # Все не-RU исключаем из scope и не считаем ошибками.
            if matched_column in df_to_validate.columns:
                banks_norm = df_to_validate[matched_column].astype(str).str.strip().str.upper()
                ru_mask = banks_norm == "RU"
                ru_cnt = int(ru_mask.sum())
                before_ru = len(df_to_validate)
                df_to_validate = df_to_validate[ru_mask].copy()
                print(f"      [FILTER] {rule_code}: учитываем только BANKS='RU' -> {ru_cnt:,} из {before_ru:,}")
                if df_to_validate.empty:
                    print(f"      [FILTER] {rule_code}: нет строк с BANKS='RU' -> ошибок = 0")
                    return 0, 0
            params["allowed_values"] = ["RU"]
        
        # Добавляем account_group_code из KNA1 для правил, которые его используют (если таблица не KNA1)
        # Проверяем все варианты: account_group_code, b.account_group_code, a.account_group_code и т.д.
        needs_account_group_code = False
        if technical_def and table_name != "KNA1":
            technical_def_lower = technical_def.lower()
            # Проверяем различные варианты использования account_group_code / ktokd
            if (
                "account_group_code" in technical_def_lower
                or "accountgroupcode" in technical_def_lower.replace(" ", "").replace("_", "")
                or "ktokd" in technical_def_lower
            ):
                needs_account_group_code = True
        
        if needs_account_group_code:
            df_to_validate = self._add_account_group_code_from_kna1(df_to_validate, table_name, rule_code)
            if str(rule_code).strip().upper() in self.RULES_KTOKD_ONLY_9038_SCOPE:
                before_scope = len(df_to_validate)
                df_to_validate = self._filter_rows_only_ktokd_9038(df_to_validate, rule_code)
                skipped_non_9038 = before_scope - len(df_to_validate)
                print(
                    f"      [FILTER] {rule_code}: только account_group_code='9038' -> "
                    f"{len(df_to_validate):,} из {before_scope:,} "
                    f"(пропущено не 9038: {skipped_non_9038:,})"
                )
                if df_to_validate.empty:
                    st = getattr(self, "_last_kna1_join_stats", {}) or {}
                    self._log_skipped_rule(
                        rule,
                        table_name,
                        (
                            "Нет строк KNB1 с KTOKD=9038 из KNA1. "
                            f"После JOIN: {st.get('rows_after_join', before_scope):,} строк; "
                            f"с заполненным KTOKD: {st.get('filled_ktokd', '?')}; "
                            f"с KTOKD=9038: {st.get('n9038', 0):,}. "
                            "Проверьте JOIN KNB1.Customer=KNA1.Customer и поле Group_1."
                        ),
                        timestamp,
                    )
                    return 0, 0
                # KTOKD из KNA1 в каждой строке error_df (для выгрузки ошибок)
                df_to_validate = self._attach_kna1_ktokd_export_columns(df_to_validate, rule_code)
            elif self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
                df_to_validate = self._attach_kna1_ktokd_export_columns(df_to_validate, rule_code)
            if is_recon_1131:
                has_account_group = any(
                    str(c).strip().lower() in ("account_group_code", "b.account_group_code", "ktokd")
                    for c in df_to_validate.columns
                )
                if not has_account_group:
                    print("      [SKIP] RCCONF_113.1: account_group_code не доступен после JOIN (memory/SQLite), правило пропущено")
                    self._log_skipped_rule(
                        rule,
                        table_name,
                        "RCCONF_113.1 skipped: account_group_code (KTOKD) not available after KNA1 join",
                        timestamp,
                    )
                    return 0, 0

        # RCCONF_371.2 (BUT051):
        # Проверяем PAFKT только для контактов с central_order_block_code='M' (KNA1.AUFSD).
        if rule_code == "RCCONF_371.2" and str(table_name or "").strip().upper() == "BUT051":
            df_to_validate = self._add_central_order_block_code_from_kna1(df_to_validate, table_name, rule_code)
            block_col = next(
                (c for c in df_to_validate.columns if str(c).strip().lower() == "central_order_block_code"),
                None,
            )
            if not block_col:
                self._log_skipped_rule(
                    rule,
                    table_name,
                    "Для RCCONF_371.2 не удалось получить central_order_block_code (KNA1.AUFSD)",
                    timestamp,
                )
                return 0, 0

            pafkt_norm = df_to_validate[matched_column].astype(str).str.strip()
            block_norm = df_to_validate[block_col].astype(str).str.strip().str.upper()
            eval_mask = (pafkt_norm != "") & (block_norm == "M")
            before_count = len(df_to_validate)
            df_to_validate = df_to_validate[eval_mask].copy()
            print(
                f"      [FILTER] RCCONF_371.2 scope: PAFKT not null AND central_order_block_code='M' -> "
                f"{len(df_to_validate):,} из {before_count:,}"
            )
            if df_to_validate.empty:
                self._log_skipped_rule(
                    rule,
                    table_name,
                    "Нет строк для оценки RCCONF_371.2 после фильтра PAFKT + central_order_block_code='M'",
                    timestamp,
                )
                return 0, 0

            params["allowed_values"] = [
                "0001", "0002", "0003", "0004", "0005", "0006", "0007", "0008", "0009", "0010",
                "DF", "F1", "F2", "F3", "F4", "F5", "FX", "Z1", "Z2", "Z3", "Z4", "Z5", "Z6", "Z7",
            ]
        
        # RCCONF_39.5 / RCCONF_39.5.2: проверяем только формат ЗАПОЛНЕННЫХ номеров. NULL/пустые исключаем до валидатора
        # (в technical_definition_RU "IF TEL_NUMBER IS NULL THEN ''" = пропустить запись; здесь делаем это явно).
        if rule_code in ["RCCONF_39.5", "RCCONF_39.5.2"] and matched_column and matched_column in df_to_validate.columns:
            before_filter_count = len(df_to_validate)
            col_series = df_to_validate[matched_column]
            null_or_empty = (
                col_series.isna()
                | (col_series.astype(str).str.strip() == '')
                | (col_series.astype(str).str.strip().str.lower().isin(['none', 'null', 'nan', 'na']))
            )
            after_filter_count = (~null_or_empty).sum()
            dropped = before_filter_count - after_filter_count
            
            if self.debug and before_filter_count > 0:
                null_count = col_series.isna().sum()
                empty_str = (col_series.astype(str).str.strip() == '').sum()
                null_like = (col_series.astype(str).str.strip().str.lower().isin(['null', 'nan', 'none', ''])).sum()
                print(f"      [DEBUG] {rule_code}: до фильтра — всего строк={before_filter_count}, isna()={null_count}, empty_str={empty_str}, null_like={null_like}")
                if null_or_empty.any():
                    sample_null = col_series[null_or_empty].head(3).tolist()
                    print(f"      [DEBUG] {rule_code}: примеры отброшенных (NULL/пустых): {sample_null}")
            
            df_to_validate = df_to_validate[~null_or_empty].copy()
            print(f"      [FILTER] {rule_code}: только непустые {matched_column} — {len(df_to_validate):,} из {before_filter_count:,} (исключено {dropped:,})")
            
            if self.debug and len(df_to_validate) > 0:
                sample_filled = df_to_validate[matched_column].head(5).tolist()
                print(f"      [DEBUG] {rule_code}: в валидатор передано строк: {len(df_to_validate):,}, первые 5 TEL_NUMBER: {sample_filled}")
            
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, "Нет записей с заполненным TEL_NUMBER для проверки формата", timestamp)
                return 0, 0
        
        try:
            # RCCOMP_180.1 (BUT0BK):
            # После фильтра scope + BANKS='RU' в df_to_validate остаются только валидные строки.
            # Поэтому ошибок быть не может и файл ошибок не должен создаваться.
            if rule_code == "RCCOMP_180.1" and str(table_name or "").strip().upper() == "BUT0BK":
                total_rows = len(df_to_validate)
                result = {
                    "rule_code": rule_code,
                    "rule_description": rule_description,
                    "quality_category": quality_category,
                    "table_name": table_name,
                    "column_checked": column_to_check,
                    "matched_column": matched_column,
                    "total_records": total_rows,
                    "passed": total_rows,
                    "failed": 0,
                    "success_rate_%": 100.0 if total_rows > 0 else 0,
                    "execution_time_sec": 0,
                    "check_date": timestamp,
                    "status": "УСПЕШНО",
                    "status_color": "green",
                    "error_file": "Нет",
                    "comments": "",
                }
                self.results.append(result)
                return 0, total_rows

            # Специальная обработка для правила RCCONF_143.7 (как в technical_definition_RU):
            #   IF sales_group_code IS NULL THEN ''
            #   ELSE IF is_sales_group_office_allowed = '1' THEN '1' ELSE '0'
            # где:
            #   sales_group_code = KNVV.VKGRP
            #   is_sales_group_office_allowed = IF NOT NULL records found IN a8 (TVBVK) THEN '1' ELSE '0'
            if rule_code == "RCCONF_143.7":
                ref_table_name = self._get_reference_table_for_rule(rule_code, "RCCONF_143.7_reference_table") or "TVBVK"
                print(f"      [REF] Правило {rule_code}: данные в KNVV, справочник допустимых VKGRP: '{ref_table_name}'")
                
                def _norm_vkgrp_code(v):
                    # VKGRP в данных/справочниках часто приходит как '001' vs '1' или как 1.0 из SQLite/Excel.
                    s = self._norm_lookup_value(v).upper()
                    if not s:
                        return ""
                    # если чисто число — убираем ведущие нули
                    if re.fullmatch(r"\d+", s):
                        return s.lstrip("0") or "0"
                    return s

                valid_vkgrp = None
                ref_df = self.memory_manager.get_table(ref_table_name)
                if (ref_df is None or ref_df.empty) and getattr(self, "db_path", None):
                    try:
                        self.memory_manager.load_selected_tables_to_ram([ref_table_name], add_reference_tables=False)
                        ref_df = self.memory_manager.get_table(ref_table_name)
                    except Exception:
                        ref_df = self.memory_manager.get_table(ref_table_name)
                if ref_df is not None and not ref_df.empty:
                    ref_filtered = ref_df.copy()

                    # Ищем колонку VKGRP (sales_group_code) в справочнике
                    vkgrp_col = None
                    for col in ref_filtered.columns:
                        col_lower = col.lower()
                        if col_lower == 'vkgrp' or 'sales_group' in col_lower or 'vkgrp' in col_lower:
                            vkgrp_col = col
                            print(f"      [REF] Найдена колонка VKGRP в {ref_table_name}: {col}")
                            break

                    if vkgrp_col:
                        valid_vkgrp = set()
                        for _, row in ref_filtered.iterrows():
                            vkgrp_val = _norm_vkgrp_code(row[vkgrp_col])
                            if vkgrp_val:
                                valid_vkgrp.add(vkgrp_val)
                        print(f"      [REF] Загружено {len(valid_vkgrp)} допустимых значений VKGRP из '{ref_table_name}'")
                    else:
                        msg = (
                            f"RCCONF_143.7 skipped: в справочнике '{ref_table_name}' нет колонки VKGRP "
                            "(для текущего правила ожидается TVBVK с полем VKGRP)."
                        )
                        print(f"      [SKIP] {msg}")
                        self._log_skipped_rule(rule, table_name, msg, timestamp)
                        return 0, 0
                
                if not valid_vkgrp:
                    print(f"      [WARN] Справочник '{ref_table_name}' не найден или пуст; нет допустимых VKGRP для правила {rule_code}")
                    self._log_skipped_rule(
                        rule, table_name,
                        f"Справочник {ref_table_name} не найден или в нём нет колонки VKGRP.",
                        timestamp
                    )
                    return 0, 0

                # Ищем колонку VKGRP в KNVV (таблица с данными для проверки)
                sales_group_col = None
                for col in df_to_validate.columns:
                    col_lower = col.lower()
                    if col_lower == 'vkgrp' or col_lower == 'sales_group_code' or 'sales_group' in col_lower:
                        sales_group_col = col
                        break
                
                if not sales_group_col:
                    # Пробуем найти через matched_column (value_checked = sales_group_code)
                    if matched_column and ('group' in matched_column.lower() or 'vkgrp' in matched_column.lower()):
                        sales_group_col = matched_column
                
                if not sales_group_col:
                    print(f"      [WARN] Колонка Sales Group (VKGRP) не найдена в KNVV для правила {rule_code}")
                    print(f"      [WARN] Искали: sales_group_col={sales_group_col}")
                    print(f"      [WARN] Доступные колонки: {list(df_to_validate.columns)[:10]}...")
                    self._log_skipped_rule(rule, table_name, "Колонка Sales Group (VKGRP) не найдена в KNVV", timestamp)
                    return 0, 0

                print(f"      [REF] Используем колонку VKGRP='{sales_group_col}'")
                # Проверяем VKGRP: если значение найдено в справочнике, то is_sales_group_office_allowed = '1'
                # Логика: IF VKGRP IS NULL THEN '' ELSE IF NOT NULL records found VKGRP IN ref_table THEN '1' ELSE '0'
                error_mask = pd.Series([False] * len(df_to_validate), index=df_to_validate.index)

                # Проверяем только строки, где VKGRP не NULL и не "0" (0 считаем пустым значением)
                non_null_mask = (
                    df_to_validate[sales_group_col].notna()
                    & (df_to_validate[sales_group_col].astype(str).str.strip() != '')
                    & (df_to_validate[sales_group_col].astype(str).str.strip().str.lower() != 'null')
                    & (df_to_validate[sales_group_col].astype(str).str.strip() != '0')
                )

                if non_null_mask.any():
                    # Нормализуем VKGRP как в справочнике (убрать .0, кавычки, пробелы, null-like и т.п.)
                    sales_group_norm = (
                        df_to_validate.loc[non_null_mask, sales_group_col]
                        .apply(_norm_vkgrp_code)
                    )

                    # Векторизованная проверка
                    missing = ~sales_group_norm.isin(valid_vkgrp)
                    if missing.any():
                        error_mask.loc[missing.index[missing]] = True

                    if self.debug:
                        try:
                            u = int(sales_group_norm.nunique(dropna=True))
                            sample = sales_group_norm[missing].head(5).tolist()
                            hit = int((~missing).sum())
                            tot = int(len(sales_group_norm))
                            print(f"      [DEBUG] RCCONF_143.7: VKGRP unique={u}, match={hit:,}/{tot:,} ({(hit/max(tot,1))*100:.1f}%), missing_sample={sample}")
                        except Exception:
                            pass
                
                error_count = int(error_mask.sum())
                # Всего записей = только оценённые (поле не пустое — скипы не входят)
                total_rows = int(non_null_mask.sum())
                
                error_description = (
                    f'Sales Group value in {sales_group_col} not found in reference table {ref_table_name} '
                    f'(SPRAS=\'E\' if applicable). VKGRP must exist in the reference table to be valid.'
                )
                
                if error_count > 0:
                    error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description)
                else:
                    error_df = None
                
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {
                        "rule_code": rule_code,
                        "rule_description": rule.get("rule_description", "Unknown rule"),
                        "quality_category": rule.get("quality_category", "Unknown"),
                        "table_name": table_name,
                        "original_column": column_to_check,
                        "matched_column": matched_column,
                    }
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                else:
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return error_count, total_rows
            
            if rule_code == "RCCONF_372.1":
                print(
                    f"      [REF] {rule_code}: ADR2.COUNTRY (ISO 3166) vs справочник T005 "
                    f"(SAP LAND1 = 2-симв. код, в выгрузке часто C/R)..."
                )

                t005_df = self._get_table_for_rules("T005")
                if t005_df is None or t005_df.empty:
                    print(f"      [WARN] Справочник T005 не найден или пуст для правила {rule_code}")
                    self._log_skipped_rule(
                        rule, table_name,
                        "RCCONF_372.1: справочник T005 не найден (нужен для ADR2.COUNTRY)",
                        timestamp,
                    )
                    return 0, 0

                land1_col = None
                for ref_candidate in ("LAND1", "C/R", "ISO_Code"):
                    land1_col = self._resolve_column_for_rule(t005_df, ref_candidate, "T005")
                    if land1_col:
                        break
                if land1_col:
                    print(f"      [REF] T005 — колонка кодов стран (LAND1): {land1_col}")
                else:
                    print(
                        f"      [WARN] В T005 не найдена колонка кодов стран "
                        f"(LAND1 / C/R / ISO_Code). Колонки: {list(t005_df.columns)[:12]}..."
                    )
                    self._log_skipped_rule(
                        rule, table_name,
                        "RCCONF_372.1: в справочнике T005 нет колонки кодов стран "
                        "(LAND1, C/R, ISO_Code) для сравнения с ADR2.COUNTRY",
                        timestamp,
                    )
                    return 0, 0
                
                # Получаем список валидных кодов стран из T005
                valid_countries = set()
                for val in t005_df[land1_col].dropna():
                    val_str = str(val).strip().upper()
                    if val_str:
                        valid_countries.add(val_str)
                
                print(f"      [REF] Загружено {len(valid_countries)} валидных кодов стран из T005")
                
                # Проверяем значения COUNTRY в ADR2
                # Дополнительный поиск колонки COUNTRY (может быть написана маленькими буквами)
                country_col = matched_column
                if country_col not in df_to_validate.columns:
                    # Ищем альтернативные варианты названия колонки
                    for col in df_to_validate.columns:
                        col_lower = col.lower()
                        if col_lower in ['country', 'country_code', 'co', 'cntry'] or 'country' in col_lower:
                            country_col = col
                            print(f"      [REF] Найдена альтернативная колонка COUNTRY: {col}")
                            break
                
                if country_col not in df_to_validate.columns:
                    print(f"      [WARN] Колонка COUNTRY не найдена в данных для правила {rule_code}")
                    print(f"      [WARN] Искали: {matched_column}, доступные колонки: {list(df_to_validate.columns)[:10]}...")
                    self._log_skipped_rule(rule, table_name, f"Колонка COUNTRY не найдена", timestamp)
                    return 0, 0
                
                # Пропускаем NULL значения (согласно правилу: IF country_code IS NULL THEN '')
                null_mask = df_to_validate[country_col].isna() | (df_to_validate[country_col].astype(str).str.strip() == '')
                
                # Для не-NULL значений проверяем наличие в справочнике (векторизованно)
                non_null_mask = ~null_mask
                
                if non_null_mask.any():
                    # Векторизованная проверка: преобразуем все значения в верхний регистр и проверяем через isin
                    country_values_upper = df_to_validate.loc[non_null_mask, country_col].astype(str).str.strip().str.upper()
                    error_mask = non_null_mask & ~country_values_upper.isin(valid_countries)
                else:
                    error_mask = pd.Series([False] * len(df_to_validate), index=df_to_validate.index)
                
                error_count = error_mask.sum()
                # Всего записей = только оценённые (country не пустой — скипы не входят)
                total_rows = int(non_null_mask.sum())
                
                error_description = f'Country code in {country_col} not found in reference table T005 (ISO 3166 standard). Valid codes must exist in T005.LAND1.'
                
                if error_count > 0:
                    error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description)
                else:
                    error_df = None
                
                # Сохраняем результат и ошибки (rule_errors заполняет _save_rule_error_with_limit с error_count)
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {
                        'rule_code': rule_code,
                        'rule_description': rule.get("rule_description", "Unknown rule"),
                        'quality_category': rule.get("quality_category", "Unknown"),
                        'table_name': table_name,
                        'original_column': column_to_check,
                        'matched_column': matched_column
                    }
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            if error_df is not None and not error_df.empty:
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    else:
                        if error_df is not None and not error_df.empty:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                elif error_df is not None and not error_df.empty:
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return error_count, total_rows

            # RCCONF_383.1 / RCCONF_384.1:
            # Формат координат должен быть строго (-)x.xxxxxx / (-)xx.xxxxxx / (-)xxx.xxxxxx
            # (разделитель — точка, ровно 6 знаков после точки).
            if rule_code in ("RCCONF_383.1", "RCCONF_384.1"):
                # Подбираем колонку координаты (rules могут содержать _LOT_GC_..., а в БД /LOT/GC_...)
                coord_col = matched_column if matched_column in df_to_validate.columns else None
                if coord_col is None:
                    is_long = rule_code == "RCCONF_383.1"
                    candidates = (
                        ["/LOT/GC_LONGITUD", "_LOT_GC_LONGITUD", "LONGITUDE", "longitude"]
                        if is_long else
                        ["/LOT/GC_LATITUDE", "_LOT_GC_LATITUDE", "LATITUDE", "latitude"]
                    )
                    for cand in candidates:
                        c = self._find_column_alternative(df_to_validate.columns, cand, table_name)
                        if c and c in df_to_validate.columns:
                            coord_col = c
                            break
                if coord_col is None:
                    self._log_skipped_rule(rule, table_name, f"Для {rule_code} не найдена колонка координаты", timestamp)
                    return 0, 0

                # account_group_code LIKE '7%' — skip, если колонка доступна
                account_group_col = next(
                    (
                        c for c in df_to_validate.columns
                        if str(c).strip().lower() in ("account_group_code", "b.account_group_code", "ktokd", "b.ktokd", "kna.ktokd", "kna.ktokd")
                    ),
                    None
                )
                account_group_skip = pd.Series(False, index=df_to_validate.index)
                if account_group_col is not None:
                    account_group_skip = df_to_validate[account_group_col].astype(str).str.strip().str.startswith("7")

                s = df_to_validate[coord_col].astype(str).str.strip()
                null_like = (
                    df_to_validate[coord_col].isna()
                    | (s == "")
                    | (s.str.lower().isin(["none", "null", "nan", "na"]))
                )
                integer_part = (
                    s.str.replace(",", ".", regex=False)
                     .str.extract(r"^\s*([+-]?\d+)", expand=False)
                     .fillna("")
                )
                integer_zero = integer_part.str.lstrip("+-").str.lstrip("0").eq("")
                zero_skip = integer_zero & (~null_like)  # TO_INTEGER(value)=0 -> skip

                skip_mask = null_like | zero_skip | account_group_skip
                evaluated_mask = ~skip_mask
                total_rows = int(evaluated_mask.sum())
                print(
                    f"      [DEBUG] {rule_code}: total={len(df_to_validate):,}, "
                    f"null_like={int(null_like.sum()):,}, to_integer_zero={int(zero_skip.sum()):,}, "
                    f"account_group_7xx={int(account_group_skip.sum()):,}, evaluated={total_rows:,}"
                )
                if total_rows == 0:
                    return 0, 0

                fmt_ok = s.str.match(r"^-?\d{1,3}\.\d{6}$", na=False)
                error_mask = evaluated_mask & (~fmt_ok)
                error_count = int(error_mask.sum())
                print(f"      [DEBUG] {rule_code}: coord_col={coord_col}, evaluated={total_rows:,}, errors={error_count:,}")

                error_description = (
                    f"Invalid coordinate format in {coord_col}. Expected (-)x.xxxxxx / (-)xx.xxxxxx / (-)xxx.xxxxxx with dot as decimal separator."
                )
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, "CONFORMITY", error_description) if error_count > 0 else None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {
                        "rule_code": rule_code,
                        "rule_description": rule.get("rule_description", "Unknown rule"),
                        "quality_category": rule.get("quality_category", "Unknown"),
                        "table_name": table_name,
                        "original_column": column_to_check,
                        "matched_column": coord_col,
                    }
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            if error_df is not None and not error_df.empty:
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    else:
                        if error_df is not None and not error_df.empty:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                elif error_df is not None and not error_df.empty:
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return error_count, total_rows

            # RCCONF_372.2 (ADR2): COUNTRY обязателен, если TEL_NUMBER заполнен.
            # technical_definition_RU в rules.json содержит условия по contact_medium_type/source,
            # но в наших данных ADR2 фактически определяет область по наличию TEL_NUMBER.
            if rule_code == "RCCONF_372.2":
                # Ищем колонки TEL_NUMBER и COUNTRY
                tel_col = None
                for c in df_to_validate.columns:
                    if str(c).strip().upper() == "TEL_NUMBER":
                        tel_col = c
                        break
                if tel_col is None:
                    tel_col = self._find_column_alternative(df_to_validate.columns, "TEL_NUMBER", table_name)
                country_col = None
                for c in df_to_validate.columns:
                    if str(c).strip().upper() in ("COUNTRY", "COUNTRY_CODE"):
                        country_col = c
                        break
                if country_col is None:
                    country_col = matched_column if matched_column in df_to_validate.columns else None
                if country_col is None:
                    for c in df_to_validate.columns:
                        if "COUNTRY" in str(c).strip().upper():
                            country_col = c
                            break

                if not tel_col or tel_col not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, "Для RCCONF_372.2 не найдена колонка TEL_NUMBER", timestamp)
                    return 0, 0
                if not country_col or country_col not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, "Для RCCONF_372.2 не найдена колонка COUNTRY", timestamp)
                    return 0, 0

                tel_s = df_to_validate[tel_col].astype(str).str.strip()
                tel_filled = (
                    df_to_validate[tel_col].notna()
                    & (tel_s != "")
                    & (~tel_s.str.lower().isin(["none", "null", "nan", "na"]))
                )
                if not tel_filled.any():
                    # Нет строк для оценки (как IF contact_medium_value IS NULL THEN '')
                    return 0, 0

                country_s = df_to_validate[country_col].astype(str).str.strip()
                country_missing = (
                    df_to_validate[country_col].isna()
                    | (country_s == "")
                    | (country_s.str.lower().isin(["none", "null", "nan", "na"]))
                )
                error_mask = tel_filled & country_missing
                error_count = int(error_mask.sum())
                total_rows = int(tel_filled.sum())

                error_description = (
                    f"{country_col} is required when {tel_col} has a value (ADR2 scope)."
                )
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, "CONFORMITY", error_description) if error_count > 0 else None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {
                        "rule_code": rule_code,
                        "rule_description": rule.get("rule_description", "Unknown rule"),
                        "quality_category": rule.get("quality_category", "Unknown"),
                        "table_name": table_name,
                        "original_column": column_to_check,
                        "matched_column": country_col,
                    }
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            if error_df is not None and not error_df.empty:
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    else:
                        if error_df is not None and not error_df.empty:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                elif error_df is not None and not error_df.empty:
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return error_count, total_rows
            
            # RCCONF_170.7 (KNVV): Consistency between Customer Group 4 and Order Block
            # IF customer_group_4_code IS NULL THEN '' (пропустить)
            # ELSE IF combination found in matrix with flag 0 THEN '0' ELSE '1'
            if rule_code == "RCCONF_170.7":
                print(f"      [FILTER] Обработка правила {rule_code}: пропускаем записи с пустым Customer Group 4...")
                
                # Ищем колонку Customer Group 4 (KVGR4)
                customer_group_4_col = None
                for col in df_to_validate.columns:
                    col_lower = col.lower()
                    if col_lower == 'kvgr4' or col_lower == 'customer_group_4_code' or 'customer_group_4' in col_lower:
                        customer_group_4_col = col
                        print(f"      [FILTER] Найдена колонка Customer Group 4: {col}")
                        break
                
                if customer_group_4_col:
                    # Фильтруем: оставляем только записи, где customer_group_4_code НЕ NULL и не пустое
                    # Также учитываем '-1' как пустое значение (как указано в других правилах)
                    before_filter = len(df_to_validate)
                    non_null_mask = (
                        df_to_validate[customer_group_4_col].notna() &
                        (df_to_validate[customer_group_4_col].astype(str).str.strip() != '') &
                        (df_to_validate[customer_group_4_col].astype(str).str.strip() != '-1') &
                        (df_to_validate[customer_group_4_col].astype(str).str.strip().str.lower() != 'null')
                    )
                    df_to_validate = df_to_validate[non_null_mask].copy()
                    filtered_count = before_filter - len(df_to_validate)
                    
                    if filtered_count > 0:
                        print(f"      [FILTER] Пропущено записей с пустым Customer Group 4: {filtered_count} (осталось {len(df_to_validate)} для проверки)")
                    
                    if df_to_validate.empty:
                        print(f"      [INFO] После фильтрации все записи пропущены (Customer Group 4 пустое)")
                        self._log_skipped_rule(rule, table_name, "Все записи имеют пустое Customer Group 4", timestamp)
                        return 0, 0
                else:
                    print(f"      [WARN] Колонка Customer Group 4 (KVGR4) не найдена для правила {rule_code}")
            
            # Передаем дополнительные параметры валидатору
            if rule_code in ["RCCONF_38.3", "RCCONF_38.5", "RCCONF_39.3", "RCCONF_39.3.2", "RCCONF_39.5", "RCCONF_39.5.2"]:
                params['technical_definition'] = technical_def
                params['rule_code'] = rule_code
            elif rule_code in ['RCCONF_18.2', 'RCCONF_22.2']:
                params['rule_code'] = rule_code
                params['technical_definition'] = technical_def
            else:
                # Унифицированно передаём rule_code в валидаторы (безопасно через **kwargs)
                params['rule_code'] = rule_code
            
            # Отладочный вывод для правил BUT000
            if table_name == "BUT000" and isinstance(validator, (CrossColumnEqualityValidator, CrossColumnEqualityCheckValidator)):
                second_col = params.get('second_column', 'НЕ НАЙДЕНА')
                print(f"      [DEBUG] Правило {rule_code}: сравниваем '{matched_column}' с '{second_col}'")
                if second_col in df_to_validate.columns:
                    sample_vals1 = df_to_validate[matched_column].head(3).tolist() if matched_column in df_to_validate.columns else []
                    sample_vals2 = df_to_validate[second_col].head(3).tolist()
                    print(f"      [DEBUG] Примеры значений: {matched_column}={sample_vals1}, {second_col}={sample_vals2}")
            
            if self.debug and rule_code in ["RCCONF_39.5", "RCCONF_39.5.2"]:
                print(f"      [DEBUG] Вызов validator.validate() для {rule_code}: строк={len(df_to_validate):,}, колонка={matched_column}")
            
            # RCCONF_119.2: сначала подготавливаем источник KNVV через БД, затем сравнение внутри валидатора
            if rule_code == "RCCONF_119.2" and isinstance(validator, PaymentTermsConsistencyValidator):
                # KNB1: KUNNR в текущем df
                knb1_kunnr_col = self._find_kunnr_column(df_to_validate) or next(
                    (c for c in df_to_validate.columns if str(c).strip().upper() == "KUNNR"),
                    None,
                )

                # KNVV можно взять из памяти, но основной путь — SQL внутри валидатора (db_path)
                knvv_df = self.memory_manager.get_table("KNVV")
                knvv_kunnr_col = None
                knvv_zterm_col = None
                if knvv_df is not None and not knvv_df.empty:
                    knvv_kunnr_col = next((c for c in knvv_df.columns if str(c).strip().upper() == "KUNNR"), None) or self._find_kunnr_column(knvv_df)
                    knvv_zterm_col = next((c for c in knvv_df.columns if str(c).strip().upper() == "ZTERM"), None) or self._find_column_alternative(knvv_df.columns, "ZTERM", "KNVV")

                params.update(
                    {
                        "knb1_kunnr_col": knb1_kunnr_col,
                        "db_path": getattr(self, "db_path", None),
                        "knvv_df": knvv_df,
                        "knvv_kunnr_col": knvv_kunnr_col,
                        "knvv_zterm_col": knvv_zterm_col,
                    }
                )

            # RCCONF_113.1: account_group_code + reconciliation_account в conf_recon_accounts.json
            if is_recon_1131 and isinstance(validator, ReconAccountConsistencyValidator):
                account_group_col = None
                for c in df_to_validate.columns:
                    cu = str(c).strip().lower()
                    if cu in ("account_group_code", "b.account_group_code", "ktokd", "b.ktokd"):
                        account_group_col = c
                        break
                if not account_group_col:
                    account_group_col = self._find_column_alternative(df_to_validate.columns, "account_group_code", table_name)

                recon_ref_path = os.path.join(parent_dir, "json files", "conf_recon_accounts.json")
                if not os.path.isfile(recon_ref_path):
                    self._log_skipped_rule(
                        rule,
                        table_name,
                        f"RCCONF_113.1: не найден {recon_ref_path}",
                        timestamp,
                    )
                    return 0, 0
                params.update(
                    {
                        "account_group_col": account_group_col,
                        "reference_path": recon_ref_path,
                    }
                )
                print(
                    f"      [DEBUG] RCCONF_113.1: recon_col={matched_column}, "
                    f"account_group_col={account_group_col}, ref={recon_ref_path}"
                )

            # RCCONF_115.11:
            # Проверка только terms_of_payment_code (KNB1.ZTERM) + planning_group_code (KNB1.FDGRV)
            # по conf_planning_group_matrix, где terms_of_payment_code='*' — wildcard.
            # customer_hierarchy_4_code (KVGR4) в этой реализации НЕ используется.
            # Фильтры по KNA1.KTOKD (account_group_code): см. technical_definition_RU в rules.json.
            if rule_code == "RCCONF_115.11":
                matrix_rules, matrix_path = self._load_planning_group_matrix()
                if not matrix_rules:
                    self._log_skipped_rule(
                        rule,
                        table_name,
                        "conf_planning_group_matrix.json не найден или пуст",
                        timestamp,
                    )
                    return 0, 0

                # account_group_code: предпочитаем колонку после JOIN из KNA1 (kna.KTOKD / account_group_code)
                account_group_col = next(
                    (c for c in df_to_validate.columns if str(c).strip() in ("kna.KTOKD", "kna.ktokd")),
                    None,
                ) or next(
                    (c for c in df_to_validate.columns if str(c).strip().lower() in ("account_group_code", "b.account_group_code", "ktokd", "b.ktokd")),
                    None,
                ) or self._find_column_alternative(df_to_validate.columns, "account_group_code", table_name)

                zterm_col = next(
                    (c for c in df_to_validate.columns if str(c).strip().upper() == "ZTERM"),
                    None,
                ) or self._find_column_alternative(df_to_validate.columns, "ZTERM", table_name)

                if not account_group_col:
                    self._log_skipped_rule(
                        rule,
                        table_name,
                        "Для RCCONF_115.11 не найден account_group_code (KTOKD из KNA1 после JOIN)",
                        timestamp,
                    )
                    return 0, 0
                if not zterm_col:
                    self._log_skipped_rule(rule, table_name, "Для RCCONF_115.11 не найдена колонка ZTERM", timestamp)
                    return 0, 0

                print(
                    f"      [DEBUG] RCCONF_115.11: fdgrv_col={matched_column}, "
                    f"account_group_col={account_group_col} (KNA1.KTOKD), zterm_col={zterm_col}, matrix={matrix_path}"
                )
                print("      [DEBUG] RCCONF_115.11 logic: compare ZTERM + FDGRV only; '*' in ZTERM is wildcard")

                # Матрица: terms_of_payment_code -> {planning_group_code}
                # '*' в terms_of_payment_code = wildcard для любого ZTERM.
                term_to_planning = {}
                for item in matrix_rules:
                    if not isinstance(item, dict):
                        continue
                    t = self._norm_lookup_value(item.get("terms_of_payment_code")).upper()
                    p = self._norm_lookup_value(item.get("planning_group_code")).upper()
                    if not p:
                        continue
                    if not t:
                        t = "*"
                    term_to_planning.setdefault(t, set()).add(p)

                if not term_to_planning:
                    self._log_skipped_rule(
                        rule,
                        table_name,
                        "RCCONF_115.11: conf_planning_group_matrix не содержит валидных строк rules[]",
                        timestamp,
                    )
                    return 0, 0

                fdgrv_norm = df_to_validate[matched_column].apply(self._norm_lookup_value)
                account_group_norm = df_to_validate[account_group_col].apply(self._norm_lookup_value)
                zterm_norm = df_to_validate[zterm_col].apply(self._norm_lookup_value)

                # Для RCCONF_115.11 пустой FDGRV должен трактоваться как "нет значения" и пропускаться,
                # даже если в источнике это не SQL NULL, а пустая строка/пробелы/null-like строка.
                skip_mask = (
                    (fdgrv_norm == "")
                    | account_group_norm.str.startswith("7")
                    | account_group_norm.isin({"9096", "9022", "9023", "9095"})
                )
                evaluated_mask = ~skip_mask
                total_rows = int(evaluated_mask.sum())
                if total_rows == 0:
                    self._log_skipped_rule(rule, table_name, "Нет записей для оценки RCCONF_115.11 после skip-условий", timestamp)
                    return 0, 0

                fdgrv_upper = fdgrv_norm.str.upper()
                zterm_upper = zterm_norm.str.upper()

                # Для каждой строки допустимые коды = term_to_planning[ZTERM] ∪ term_to_planning['*']
                allowed_series = pd.Series([set()] * len(df_to_validate), index=df_to_validate.index, dtype=object)
                for idx in df_to_validate.index[evaluated_mask]:
                    t = zterm_upper.loc[idx] if zterm_upper.loc[idx] else "*"
                    allowed = set()
                    allowed |= term_to_planning.get(t, set())
                    allowed |= term_to_planning.get("*", set())
                    allowed_series.loc[idx] = allowed

                # Ошибка: FDGRV не входит в допустимое множество для этой строки
                error_mask = evaluated_mask & ~df_to_validate.index.to_series().apply(lambda i: fdgrv_upper.loc[i] in (allowed_series.loc[i] or set()))
                error_count = int(error_mask.sum())
                error_description = (
                    "Invalid value in column FDGRV: planning_group_code not found in conf_planning_group_matrix "
                    "for terms_of_payment_code (ZTERM) with wildcard '*'. "
                    "Skip if FDGRV empty or KNA1 account_group_code (KTOKD) starts with 7 or IN (9096,9022,9023,9095)."
                )
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, "CONFORMITY", error_description) if error_count > 0 else None
                if error_df is not None and not error_df.empty:
                    error_df["FDGRV_ACTUAL"] = fdgrv_norm.loc[error_df.index].values
                    error_df["LOOKUP_ACCOUNT_GROUP_KTOKD"] = account_group_norm.loc[error_df.index].values
                    error_df["LOOKUP_TERMS_OF_PAYMENT"] = zterm_norm.loc[error_df.index].values
                    error_df["ALLOWED_PLANNING_GROUPS"] = error_df.index.to_series().apply(
                        lambda i: ", ".join(sorted(list(allowed_series.loc[i] or set())))[:2000]
                    ).values
                    # Какие поля участвуют в правиле (имена колонок в выгрузке = SAP / логические имена)
                    error_df["DQ_RULE_CHECK_COLUMNS"] = (
                        f"KNB1 planning_group_code [{matched_column}] (FDGRV) сверяется с conf_planning_group_matrix "
                        f"по terms_of_payment_code [{zterm_col}] (ZTERM) с wildcard '*'; "
                        f"фильтр по KNA1 account_group_code [{account_group_col}] (не «7…», не 9096/9022/9023/9095); "
                        f"пустой FDGRV — не оценивается"
                    )
                    error_df["DQ_ERROR_DESCRIPTION"] = error_df.apply(
                        lambda row: (
                            f"FDGRV='{row['FDGRV_ACTUAL']}' не найден в матрице для "
                            f"ZTERM='{row['LOOKUP_TERMS_OF_PAYMENT']}' "
                            f"(допустимо: {row['ALLOWED_PLANNING_GROUPS']}). "
                            f"KTOKD='{row['LOOKUP_ACCOUNT_GROUP_KTOKD']}'"
                        ),
                        axis=1,
                    )
                    # Дополнительная страховка: пустой FDGRV не должен попадать в экспорт ошибок
                    # для RCCONF_115.11, даже если маска выше дала ложноположительное срабатывание.
                    empty_fdgrv_mask = error_df["FDGRV_ACTUAL"].apply(lambda v: self._norm_lookup_value(v) == "")
                    if empty_fdgrv_mask.any():
                        before_cnt = len(error_df)
                        error_df = error_df[~empty_fdgrv_mask].copy()
                        error_count = len(error_df)
                        print(
                            f"      [FILTER] RCCONF_115.11: исключены строки с пустым FDGRV из error_df: "
                            f"{before_cnt} -> {error_count}"
                        )
                    if error_df.empty:
                        error_df = None
                        error_count = 0

                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {
                        "rule_code": rule_code,
                        "rule_description": rule.get("rule_description", "Unknown rule"),
                        "quality_category": rule.get("quality_category", "Unknown"),
                        "table_name": table_name,
                        "original_column": column_to_check,
                        "matched_column": matched_column,
                    }
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                            if error_df is not None and not error_df.empty:
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                        if error_df is not None and not error_df.empty:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                elif error_df is not None and not error_df.empty:
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return error_count, total_rows

            if rule_code == "RCCONF_15.1" and str(table_name or "").strip().upper() == "BUT000":
                org3_res = self._find_column_alternative(df_to_validate.columns, "NAME_ORG3", table_name)
                if org3_res:
                    params["org3_column_resolved"] = org3_res

            # RCCOMP_375.1.2: TEL_NUMBER или e-mail в ADR6 (колонка E-Mail Address / SMTP_ADDR)
            if rule_code == "RCCOMP_375.1.2":
                adr6_df = self._get_adr6_df()
                if adr6_df is not None and not adr6_df.empty:
                    params["adr6_df"] = adr6_df
                    print(f"      [ADR6] RCCOMP_375.1.2: подключена таблица ADR6 ({len(adr6_df):,} строк) для проверки e-mail")
                else:
                    print(
                        "      [WARN] RCCOMP_375.1.2: таблица ADR6 не найдена — "
                        "проверяется только TEL_NUMBER"
                    )

            total_rows, error_count, error_df = validator.validate(df_to_validate, matched_column, **params)

            if is_recon_1131 and total_rows == 0 and error_count == 0:
                st = getattr(self, "_last_kna1_join_stats", {}) or {}
                ag_col = params.get("account_group_col") or self._find_account_group_column(df_to_validate)
                skip_reason = (
                    "RCCONF_113.1: нет строк для оценки (нужны заполненные AKONT и KTOKD из KNA1). "
                    f"Строк KNB1 в срезе: {len(df_to_validate):,}"
                )
                if ag_col and ag_col in df_to_validate.columns:
                    from utils.sap_account_keys import norm_sap_account_group, norm_sap_recon_account
                    has_k = df_to_validate[ag_col].apply(norm_sap_account_group) != ""
                    has_a = df_to_validate[matched_column].apply(norm_sap_recon_account) != ""
                    skip_reason += (
                        f"; с KTOKD: {int(has_k.sum()):,}; с AKONT: {int(has_a.sum()):,}; "
                        f"с обоими: {int((has_k & has_a).sum()):,}"
                    )
                self._log_skipped_rule(rule, table_name, skip_reason, timestamp)
                return 0, 0
            
            # Для правил с «оценёнными» строками (например CrossColumnEquality: только обе колонки заполнены) — сохраняем всего в таблице
            
            # Для правила RCCONF_38.3 дополнительно фильтруем error_df по R3_USER = '1'
            # Это гарантирует, что в ошибки попадут только стационарные телефоны
            if rule_code == "RCCONF_38.3" and error_df is not None and not error_df.empty:
                r3_user_col = None
                for col in error_df.columns:
                    col_lower = col.lower()
                    if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or 'r3user' in col_lower:
                        r3_user_col = col
                        break
                
                if r3_user_col:
                    # Фильтруем только записи с R3_USER = '1'
                    before_count = len(error_df)
                    error_df = error_df[error_df[r3_user_col].astype(str).str.strip() == '1'].copy()
                    after_count = len(error_df)
                    if before_count != after_count:
                        print(f"      [FILTER] Дополнительная фильтрация error_df: {before_count} → {after_count} записей (только R3_USER='1')")
                        error_count = after_count
                else:
                    # Если R3_USER не найден в error_df, проверяем исходный DataFrame
                    # Это может означать, что валидатор вернул неправильные данные
                    print(f"      [WARN] Колонка R3_USER не найдена в error_df для правила {rule_code}")
                    # Пытаемся найти R3_USER в исходном df
                    for col in df.columns:
                        col_lower = col.lower()
                        if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or 'r3user' in col_lower:
                            # Если error_df имеет индексы из исходного df, фильтруем по индексам
                            if error_df.index.isin(df.index).any():
                                original_mask = df[col].astype(str).str.strip() == '1'
                                error_df = error_df[error_df.index.isin(df[original_mask].index)].copy()
                                error_count = len(error_df)
                                print(f"      [FILTER] Отфильтровано error_df по индексам исходного DataFrame: {len(error_df)} записей")
                            break
            
            # Проверяем корректность error_df
            if error_df is not None and not error_df.empty:
                if len(error_df) > total_rows:
                    print(f"      [ERROR] Валидатор вернул error_df с {len(error_df)} строками, но total_rows={total_rows}")
                    error_df = pd.DataFrame()
                    error_count = 0
                elif len(error_df) > error_count * 1.1:
                    print(f"      [WARN] error_df содержит {len(error_df)} строк, но error_count={error_count}")
                    error_df = error_df.head(error_count)
            
            total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
            is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
            
            # ========== ПАПКА TOTAL (ADR2): выгрузка партнёров по RCCOMP_375.1.2 отключена — используется отдельный скрипт adr2_knvv_exporter.py ==========
            filtered_adr2_addr_partner_df = None
            
            # Сохраняем ошибки и результат (общее состояние — под блокировкой при параллелизме)
            if self._parallel_lock:
                with self._parallel_lock:
                    if error_df is not None and not error_df.empty:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows)
                    if save_result:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious, filtered_adr2_addr_partner_df=filtered_adr2_addr_partner_df)
            else:
                if error_df is not None and not error_df.empty:
                    self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows)
                if save_result:
                    self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious, filtered_adr2_addr_partner_df=filtered_adr2_addr_partner_df)
            
            return error_count, total_rows
        
        except Exception as e:
            self._log_failed_rule(rule, table_name, str(e), timestamp)
            return 0, 0
    
    def _save_rule_error_with_limit(self, rule_code, table_name, error_df, error_count, is_suspicious, total_rows):
        if error_df is None or error_df.empty:
            return

        # RCCOMP_113.1: в файле ошибок обязательно показываем AKONT
        # (даже если исходная колонка была под другим заголовком и попала в DQ_COLUMN_CHECKED).
        if self._normalize_rule_code(rule_code) == "RCCOMP_113.1" and "AKONT" not in error_df.columns:
            src_col = None
            if "DQ_COLUMN_CHECKED" in error_df.columns:
                try:
                    cn = str(error_df["DQ_COLUMN_CHECKED"].iloc[0]).strip()
                    if cn and cn in error_df.columns:
                        src_col = cn
                except Exception:
                    src_col = None
            if src_col is None:
                src_col = self._find_column_alternative(error_df.columns, "AKONT", table_name)
            if src_col and src_col in error_df.columns:
                error_df = error_df.copy()
                error_df["AKONT"] = error_df[src_col]
            else:
                error_df = error_df.copy()
                error_df["AKONT"] = ""

        if self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
            error_df = self._enrich_error_df_kna1_ktokd(error_df, table_name, rule_code)

        # ADRC: в ошибки не попадают строки с NAME1=RESERVED (исключаем перед сохранением)
        if (str(table_name or "").strip().upper() == "ADRC"):
            name1_col = None
            for c in error_df.columns:
                if str(c).strip().upper() == "NAME1":
                    name1_col = c
                    break
            if name1_col is None:
                name1_col = self._find_column_alternative(error_df.columns, "NAME1", table_name)
            if name1_col is None:
                best_col, best_count = None, 0
                for c in error_df.columns:
                    try:
                        cnt = (error_df[c].astype(str).str.strip().str.upper() == "RESERVED").sum()
                        if cnt > best_count:
                            best_count, best_col = cnt, c
                    except Exception:
                        pass
                if best_col and best_count > 0:
                    name1_col = best_col
            if name1_col and name1_col in error_df.columns:
                val_str = error_df[name1_col].astype(str).str.strip().str.upper()
                error_df = error_df[val_str != "RESERVED"].copy()
                error_count = len(error_df)
                if error_df.empty:
                    return
        
        # RCCONF_39.5: в rule_errors только строки, где PERSNUMBER именно пустой
        if rule_code == "RCCONF_39.5":
            def _norm(s):
                return str(s).strip().upper().replace('_', '').replace(' ', '')
            persnumber_col = next(
                (c for c in error_df.columns if _norm(c) in ('PERSNUMBER', 'PERSONNUMBER') or 'PERSNUMBER' in _norm(c)),
                None
            )
            if persnumber_col is not None:
                s = error_df[persnumber_col].astype(str).str.strip().str.upper()
                empty_mask = (
                    error_df[persnumber_col].isna()
                    | (s == '')
                    | (s.isin(['NONE', 'NAN', 'NULL', '-', '.']))
                )
                error_df = error_df.loc[empty_mask].copy()
                error_count = len(error_df)
                if error_df.empty:
                    return
        
        # RCCONF_39.5 и RCCONF_39.5.2: жёстко выкидываем из ошибок номера 10 цифр с первой 9 (они валидны).
        # Важно: этот блок меняет только error_df для правил 39.5/39.5.2; ошибки остальных правил (RCCOMP_375.1 и т.д.) не трогаем.
        if rule_code in ["RCCONF_39.5", "RCCONF_39.5.2"]:
            def _digits(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return ''
                s = str(v).strip().replace('\ufeff', '')
                s = re.sub(r'\s+', '', s)  # все пробелы: space, tab, nbsp, newline
                try:
                    if isinstance(v, (int, float)) and not pd.isna(v) and v == int(v):
                        return str(int(v))
                except (ValueError, TypeError):
                    pass
                if re.match(r'^\d+\.0+$', s):
                    return str(int(float(s)))
                return re.sub(r'\D', '', s)
            def _is_valid_10_9(v):
                d = _digits(v)
                return len(d) == 10 and d.startswith('9')

            def _is_valid_39_5_format(v):
                d = _digits(v)
                if len(d) == 10 and d.startswith('9'):
                    return True
                if len(d) == 11 and (d.startswith('89') or d.startswith('79')):
                    return True
                if len(d) == 11 and d.startswith('8') and d[1] != '9':
                    return True
                return False
            # 1) Колонка из метаданных (та, что реально проверялась)
            tel_col = None
            if 'DQ_COLUMN_CHECKED' in error_df.columns:
                try:
                    cn = error_df['DQ_COLUMN_CHECKED'].iloc[0]
                    if cn and str(cn).strip() and str(cn).strip() in error_df.columns:
                        tel_col = str(cn).strip()
                except Exception:
                    pass
            # 2) По имени
            if tel_col is None:
                for c in error_df.columns:
                    if c in ('DQ_ERROR_TYPE', 'DQ_RULE_CODE', 'DQ_COLUMN_CHECKED', 'DQ_ERROR_DESCRIPTION', 'DQ_TIMESTAMP', 'DQ_RULE_DESCRIPTION'):
                        continue
                    cn = str(c).upper().replace('_', '').replace(' ', '')
                    if 'TELNUMBER' in cn or 'TELNR' in cn or ('TEL' in cn and 'NUM' in cn):
                        tel_col = c
                        break
            # 3) Перебор всех колонок: если хоть в одной есть 10 цифр с 9 — считаем строку валидной и выкидываем
            _fmt_ok = _is_valid_39_5_format if rule_code == "RCCONF_39.5.2" else _is_valid_10_9
            if tel_col is not None:
                drop_mask = error_df[tel_col].apply(_fmt_ok)
            else:
                drop_mask = pd.Series(False, index=error_df.index)
                for c in error_df.columns:
                    if c in ('DQ_ERROR_TYPE', 'DQ_RULE_CODE', 'DQ_COLUMN_CHECKED', 'DQ_ERROR_DESCRIPTION', 'DQ_TIMESTAMP', 'DQ_RULE_DESCRIPTION'):
                        continue
                    drop_mask = drop_mask | error_df[c].apply(_fmt_ok)
            if drop_mask.any():
                before = len(error_df)
                error_df = error_df.loc[~drop_mask].copy()
                error_count = len(error_df)
                if before > error_count:
                    print(f"      [{rule_code}] Убраны из ошибок номера с валидным форматом: {before - error_count} строк, осталось {error_count}")
            if error_df.empty:
                return
        
        key = f"{rule_code}_{table_name}"
        # Для отдельных правил/таблиц ограничение 100k снято — лимит по размеру Excel
        tbl = str(table_name or "").strip().upper()
        rule_u = str(rule_code or "").strip().upper()
        save_all_errors = (
            tbl in ("ADR2", "BUT000")
            or rule_u == "RCCONF_18.2"
            or rule_u == "RCCONF_63.1"
            or tbl.startswith("DFKKBPTAXNUM")
        )
        limit_errors = self.EXCEL_MAX_ROWS if save_all_errors else self.MAX_ERRORS_TO_SAVE
        
        # Проверяем, что error_df содержит только ошибки
        if len(error_df) > error_count * 1.1:
            print(f"      [WARN] Для {rule_code} ({table_name}): error_df содержит {len(error_df)} строк, но error_count={error_count}")
            if error_count > 0:
                error_df = error_df.head(error_count)
            else:
                print(f"      [ERROR] error_count=0, но error_df не пустой. Пропускаем сохранение.")
                return
        
        # Ограничиваем ошибки до limit_errors (для ADR2/BUT000 — до лимита Excel)
        error_df_to_save = error_df.copy()
        is_truncated = False
        
        if key in self.rule_errors:
            existing_count = self.rule_errors[key].get('error_count', 0)
            existing_df = self.rule_errors[key]['error_df']
            
            if existing_count >= limit_errors:
                print(f"      [WARN] {rule_code} ({table_name}): уже накоплено {existing_count:,} ошибок (лимит {limit_errors:,}), новые ошибки не добавляются")
                return
            
            combined_df = pd.concat([existing_df, error_df_to_save], ignore_index=True)
            total_combined = len(combined_df)
            # Для 39.5/39.5.2 убираем из объединённого датафрейма номера 10 цифр с 9
            if rule_code in ["RCCONF_39.5", "RCCONF_39.5.2"]:
                def _d(v):
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return ''
                    s = str(v).strip().replace('\ufeff', '')
                    s = re.sub(r'\s+', '', s)
                    try:
                        if isinstance(v, (int, float)) and v == int(v):
                            return str(int(v))
                    except (ValueError, TypeError):
                        pass
                    if re.match(r'^\d+\.0+$', s):
                        return str(int(float(s)))
                    return re.sub(r'\D', '', s)
                def _ok(v):
                    d = _d(v)
                    if len(d) == 10 and d.startswith('9'):
                        return True
                    if rule_code == "RCCONF_39.5.2":
                        if len(d) == 11 and (d.startswith('89') or d.startswith('79')):
                            return True
                        if len(d) == 11 and d.startswith('8') and d[1] != '9':
                            return True
                    return False
                tel_col = None
                if 'DQ_COLUMN_CHECKED' in combined_df.columns:
                    try:
                        cn = combined_df['DQ_COLUMN_CHECKED'].iloc[0]
                        if cn and str(cn).strip() in combined_df.columns:
                            tel_col = str(cn).strip()
                    except Exception:
                        pass
                if tel_col is None:
                    tel_col = next((c for c in combined_df.columns if 'TEL' in str(c).upper() and ('NUMBER' in str(c).upper() or 'NR' in str(c).upper() or 'NUM' in str(c).upper())), None)
                if tel_col is not None:
                    drop = combined_df[tel_col].apply(_ok)
                else:
                    drop = pd.Series(False, index=combined_df.index)
                    for c in combined_df.columns:
                        if 'DQ_' in str(c):
                            continue
                        drop = drop | combined_df[c].apply(_ok)
                if drop.any():
                    combined_df = combined_df.loc[~drop].copy()
                    total_combined = len(combined_df)
            
            if total_combined > limit_errors:
                combined_df = combined_df.head(limit_errors)
                is_truncated = True
                print(f"      [WARN] {rule_code} ({table_name}): ошибок {total_combined:,}, сохранено только {limit_errors:,} (первые {limit_errors:,})")
            
            self.rule_errors[key] = {
                'rule_code': rule_code,
                'table_name': table_name,
                'error_df': combined_df,
                'error_count': min(existing_count + error_count, limit_errors),
                'is_suspicious': is_suspicious,
                'total_rows': total_rows,
                'is_truncated': is_truncated or self.rule_errors[key].get('is_truncated', False)
            }
        else:
            if len(error_df_to_save) > limit_errors:
                error_df_to_save = error_df_to_save.head(limit_errors)
                is_truncated = True
                print(f"      [WARN] {rule_code} ({table_name}): ошибок {error_count:,}, сохранено только {limit_errors:,} (первые {limit_errors:,})")
            
            self.rule_errors[key] = {
                'rule_code': rule_code,
                'table_name': table_name,
                'error_df': error_df_to_save,
                'error_count': min(error_count, limit_errors),
                'is_suspicious': is_suspicious,
                'total_rows': total_rows,
                'is_truncated': is_truncated
            }
    
    def _check_if_suspicious(self, rule_code, error_count, total_rows):
        if rule_code in ['RCCONF_12.2', 'RCCONF_12.3']:
            return True
        if error_count > 1000000:
            return True
        if total_rows > 0 and (error_count / total_rows) > self.MASS_ERROR_THRESHOLD:
            return True
        if rule_code == 'RCCONF_15.1' and total_rows > 0 and (error_count / total_rows) > 0.3:
            return True
        return False
    
    def _save_rule_result(self, rule_info, total_rows, error_count, execution_time, timestamp, is_suspicious, filtered_adr2_addr_partner_df=None):
        # total_rows здесь = только оценённые строки (passed + failed), без "пустых по правилам"
        # passed = количество строк с отработкой "1" по правилу (сумма единиц)
        passed_count = total_rows - error_count
        # Всего записей = passed + failed (без строк, которые по правилам пустые). TOTAL в отчёте = passed.
        total_records_evaluated = passed_count + error_count
        success_rate = (passed_count / total_records_evaluated * 100) if total_records_evaluated > 0 else 0
        
        # Особый случай: правило формально отработало, но ни одной строки не было оценено
        # (total_rows=0 и error_count=0). Считаем это ошибкой выполнения правила, а не успешной проверкой.
        if total_records_evaluated == 0:
            status = "ОШИБКА ВЫПОЛНЕНИЯ"
            status_color = "red"
        elif error_count == 0:
            status = "УСПЕШНО"
            status_color = "green"
        elif is_suspicious:
            if error_count > self.MAX_ERRORS_TO_SAVE:
                status = "МАССОВЫЕ ОШИБКИ"
            else:
                status = "ПОДОЗРИТЕЛЬНО"
            status_color = "orange"
        else:
            status = "ОШИБКИ"
            status_color = "red"
        
        # Проверяем, есть ли ошибки в словаре rule_errors
        rule_code = rule_info['rule_code']
        table_name = rule_info['table_name']
        key = f"{rule_code}_{table_name}"
        
        has_errors_saved = key in self.rule_errors and self.rule_errors[key].get('error_df') is not None and not self.rule_errors[key]['error_df'].empty
        
        if error_count > self.MAX_ERRORS_TO_SAVE:
            error_file_status = f"[!] Частично ({self.MAX_ERRORS_TO_SAVE:,} из {error_count:,})" if has_errors_saved else "Нет (не сохранено)"
        elif error_count > 0:
            error_file_status = "Есть" if has_errors_saved else "Нет (не сохранено)"
        else:
            error_file_status = "Нет"
        
        comments = ""
        if total_records_evaluated == 0:
            comments = "Правило не смогло оценить ни одной строки (total_rows=0, errors=0). Проверьте техническое условие и входные данные."
        if error_count > self.MAX_ERRORS_TO_SAVE:
            comments = f"[!] ВНИМАНИЕ: Всего ошибок {error_count:,}, сохранено только первые {self.MAX_ERRORS_TO_SAVE:,}! Обратите внимание!"
        elif is_suspicious and total_records_evaluated > 0:
            error_percent = (error_count / total_records_evaluated * 100)
            comments = f"ПОДОЗРИТЕЛЬНО: {error_percent:.1f}% ДАННЫХ С ОШИБКАМИ - ПРОВЕРИТЬ ЛОГИКУ ПРАВИЛА"
        
        result = {
            "rule_code": rule_info['rule_code'],
            "rule_description": rule_info['rule_description'],
            "quality_category": rule_info['quality_category'],
            "table_name": rule_info['table_name'],
            "column_checked": rule_info.get('original_column', ''),
            "matched_column": rule_info['matched_column'],
            "total_records": total_records_evaluated,  # Всего записей = passed + failed по правилу
            "passed": passed_count,   # Успешно = сумма единиц (строки с "1"), TOTAL в отчёте = это же значение
            "failed": error_count,
            "total_evaluated": total_records_evaluated,  # всего записей (для расчёта % успеха)
            "success_rate_%": round(success_rate, 2),
            "execution_time_sec": round(execution_time, 2),
            "check_date": timestamp,
            "status": status,
            "status_color": status_color,
            "error_file": error_file_status,
            "comments": comments
        }
        # ADR2: для RCCOMP_375.1.2 сохраняем только счётчик в отчёт (файлы/БД — через отдельный скрипт adr2_knvv_exporter.py)
        rule_code = str(rule_info.get('rule_code', '')).strip()
        if rule_code == "RCCOMP_375.1.2":
            result["filtered_adr2_count"] = total_rows
            result["filtered_adr2_file"] = ""
        elif filtered_adr2_addr_partner_df is not None and not filtered_adr2_addr_partner_df.empty:
            result["filtered_adr2_count"] = len(filtered_adr2_addr_partner_df)
            result["filtered_adr2_file"] = ""

        self.results.append(result)
    
    # Имя таблицы в БД для списка партнёров по правилам ADR2 («всего записей» = конкретные ADDRNUMBER+PARTNER)
    ADR2_RULE_PARTNERS_TABLE = "adr2_rule_partners"
    # Имя таблицы в БД для строк ADR2 с ошибками по правилам (копируем исходные строки ADR2 + DQ_ metadata)
    ADR2_RULE_ERRORS_TABLE = "adr2_rule_errors"

    def _insert_adr2_partners_batch(self, rows: list, rule_code: str = None, run_ts: str = None):
        """Вставляет пачку строк в adr2_rule_partners. Каждая строка: (rule_code, run_ts, ADDRNUMBER, PARTNER) или (rule_code, run_ts, ADDRNUMBER, PARTNER, AUFSD). При первой вставке для пары (rule_code, run_ts) создаёт таблицу и удаляет старые строки."""
        if not rows or not getattr(self, 'db_path', None):
            return
        try:
            import sqlite3
            r0 = rows[0]
            rule_code = rule_code or (r0[0] if len(r0) >= 4 else None)
            run_ts = run_ts or (r0[1] if len(r0) >= 4 else None)
            if not rule_code or not run_ts:
                return
            has_aufsd = len(r0) >= 5
            key = (rule_code, run_ts)
            cleared = getattr(self, '_adr2_partners_cleared', None)
            if cleared is None:
                self._adr2_partners_cleared = set()
                cleared = self._adr2_partners_cleared
            conn = connect_sqlite(self.db_path)
            if key not in cleared:
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS """ + self.ADR2_RULE_PARTNERS_TABLE + """ (
                        rule_code TEXT NOT NULL,
                        run_ts TEXT NOT NULL,
                        ADDRNUMBER TEXT,
                        PARTNER TEXT,
                        AUFSD TEXT
                    )
                """)
                # Миграция: если таблица была создана без AUFSD — добавить колонку
                try:
                    info = conn.execute("PRAGMA table_info(" + self.ADR2_RULE_PARTNERS_TABLE + ")").fetchall()
                    col_names = [c[1] for c in info]
                    if 'AUFSD' not in col_names:
                        conn.execute("ALTER TABLE " + self.ADR2_RULE_PARTNERS_TABLE + " ADD COLUMN AUFSD TEXT")
                except Exception:
                    pass
                conn.execute(
                    "DELETE FROM " + self.ADR2_RULE_PARTNERS_TABLE + " WHERE rule_code = ? AND run_ts = ?",
                    (rule_code, run_ts)
                )
                cleared.add(key)
            if has_aufsd:
                conn.executemany(
                    "INSERT INTO " + self.ADR2_RULE_PARTNERS_TABLE + " (rule_code, run_ts, ADDRNUMBER, PARTNER, AUFSD) VALUES (?, ?, ?, ?, ?)",
                    rows
                )
            else:
                conn.executemany(
                    "INSERT INTO " + self.ADR2_RULE_PARTNERS_TABLE + " (rule_code, run_ts, ADDRNUMBER, PARTNER, AUFSD) VALUES (?, ?, ?, ?, ?)",
                    [(r[0], r[1], r[2], r[3], "") for r in rows]
                )
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"      [WARN] Вставка пачки в {self.ADR2_RULE_PARTNERS_TABLE}: {e}")
    
    def _save_adr2_partners_to_db(self, rule_code: str, run_ts: str, addr_partner_df: pd.DataFrame):
        """
        Записывает список партнёров (ADDRNUMBER, PARTNER) по правилу ADR2 в таблицу БД.
        Таблица создаётся при первом вызове. По одному правилу и одному run_ts — одна выборка (старые строки для этой пары заменяются).
        """
        if addr_partner_df is None or addr_partner_df.empty or not rule_code:
            return
        try:
            import sqlite3
            conn = connect_sqlite(self.db_path)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS """ + self.ADR2_RULE_PARTNERS_TABLE + """ (
                    rule_code TEXT NOT NULL,
                    run_ts TEXT NOT NULL,
                    ADDRNUMBER TEXT,
                    PARTNER TEXT
                )
            """)
            conn.execute(
                "DELETE FROM " + self.ADR2_RULE_PARTNERS_TABLE + " WHERE rule_code = ? AND run_ts = ?",
                (rule_code, run_ts)
            )
            rows = [
                (rule_code, run_ts, str(row.get('ADDRNUMBER', '') or '').strip(), str(row.get('PARTNER', '') or '').strip())
                for _, row in addr_partner_df.iterrows()
            ]
            conn.executemany(
                "INSERT INTO " + self.ADR2_RULE_PARTNERS_TABLE + " (rule_code, run_ts, ADDRNUMBER, PARTNER) VALUES (?, ?, ?, ?)",
                rows
            )
            conn.commit()
            conn.close()
            print(f"      [DB] Записано {len(addr_partner_df):,} партнёров в таблицу {self.ADR2_RULE_PARTNERS_TABLE} для правила {rule_code}")
        except Exception as e:
            print(f"      [WARN] Запись в таблицу {self.ADR2_RULE_PARTNERS_TABLE}: {e}")

    def _save_adr2_rule_errors_to_db(self, error_df: pd.DataFrame, rule_code: str, run_ts: str):
        """
        Сохраняет строки ошибок по ADR2 в sqlite.
        Требование пользователя: копируем строку из таблицы ADR2 "как есть" (все колонки ADR2),
        а также добавляем metadata об ошибке (колонки DQ_* из error_df).
        """
        if error_df is None or error_df.empty:
            return
        if not getattr(self, "db_path", None):
            return
        if not run_ts or not rule_code:
            return

        try:
            import sqlite3

            conn = connect_sqlite(self.db_path)

            # 1) Получаем схему ADR2 (точные имена/типы колонок).
            adr2_info = conn.execute('PRAGMA table_info("ADR2")').fetchall()
            if not adr2_info:
                conn.close()
                print(f"      [WARN] ADR2: таблица не найдена в БД, запись ошибок пропущена")
                return

            adr2_cols = [r[1] for r in adr2_info]
            adr2_col_types = {r[1]: (r[2] or "TEXT") for r in adr2_info}

            # 2) Приводим имена колонок error_df к регистру ADR2 (если есть различия).
            #    Это важно, потому что PRAGMA table_info возвращает имена колонок в "реальном" регистре.
            rename_map = {}
            for df_col in error_df.columns:
                for schema_col in adr2_cols:
                    if str(df_col).upper() == str(schema_col).upper() and df_col != schema_col:
                        rename_map[df_col] = schema_col
                        break
            df = error_df.rename(columns=rename_map).copy()

            # 3) Формируем итоговый набор колонок "данные" = колонки ADR2 + дополнительные из error_df.
            extra_cols = [c for c in df.columns if c not in adr2_cols]
            data_cols = adr2_cols + extra_cols

            # 4) Создаём таблицу при необходимости и добавляем отсутствующие колонки.
            table = self.ADR2_RULE_ERRORS_TABLE

            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (table,),
            ).fetchone()

            desired_cols = ["run_ts", "rule_code"] + data_cols

            if not exists:
                col_defs = [
                    '"run_ts" TEXT NOT NULL',
                    '"rule_code" TEXT NOT NULL',
                ]
                for c in data_cols:
                    c_type = adr2_col_types.get(c, "TEXT") or "TEXT"
                    col_defs.append(f'"{c}" {c_type}')

                conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({", ".join(col_defs)})')
            # Миграция: добавляем новые колонки при необходимости
            existing_cols = {
                r[1] for r in conn.execute(f'PRAGMA table_info("{table}")').fetchall()
            }
            for c in desired_cols:
                if c not in existing_cols:
                    conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{c}" TEXT')

            # 5) Перед записью удаляем предыдущие ошибки этого run_ts/rule_code (идемпотентность).
            conn.execute(
                f'DELETE FROM "{table}" WHERE run_ts = ? AND rule_code = ?',
                (run_ts, rule_code),
            )

            # 6) Вставляем строки.
            df_insert = df.copy()
            df_insert["run_ts"] = run_ts
            df_insert["rule_code"] = rule_code
            # Гарантируем наличие всех колонок.
            for c in desired_cols:
                if c not in df_insert.columns:
                    df_insert[c] = None
            df_insert = df_insert[desired_cols]
            df_insert = df_insert.where(pd.notnull(df_insert), None)

            cols_sql = ", ".join([f'"{c}"' for c in desired_cols])
            placeholders = ", ".join(["?"] * len(desired_cols))
            insert_sql = f'INSERT INTO "{table}" ({cols_sql}) VALUES ({placeholders})'

            values = df_insert.to_numpy(dtype=object)
            chunk_size = 5000
            for start in range(0, len(values), chunk_size):
                chunk = values[start:start + chunk_size]
                conn.executemany(insert_sql, [tuple(r) for r in chunk])

            conn.commit()
            conn.close()

            print(
                f"   [DB] ADR2 errors saved: table={table}, rule={rule_code}, run_ts={run_ts}, rows={len(df_insert):,}"
            )
        except Exception as e:
            # Ошибки в БД не должны ломать основной экспорт в файлы.
            print(f"   [WARN] Не удалось сохранить ADR2 ошибки в БД: {e}")
    
    def run(self, specific_table: str = None, table_list: list = None, only_rule_codes: set = None):
        """
        Основной метод запуска проверки.
        specific_table — одна таблица; table_list — список таблиц; оба None — все таблицы.
        only_rule_codes — если задано, выполняются только указанные правила (изолированно).
        """
        print(f"\n" + "="*100)
        print(f"\033[1mЗАПУСК СИСТЕМЫ ПРОВЕРКИ КАЧЕСТВА\033[0m")
        print(f"="*100)
        
        if table_list:
            print(f"[INFO] Проверяем выбранные таблицы: {len(table_list)} шт.")
        elif specific_table:
            print(f"[INFO] Проверяем только таблицу: {specific_table}")
        else:
            print(f"[INFO] Проверяем все таблицы")
        if only_rule_codes:
            print(f"[INFO] Изолированный режим: только правила {sorted(only_rule_codes)}")
        
        print(f"[INFO] Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило")
        print(f"="*100)
        
        results_df = self.run_quality_checks_fast(specific_table=specific_table, table_list=table_list, only_rule_codes=only_rule_codes)
        
        if not results_df.empty:
            print(f"\n" + "="*100)
            print(f"\033[1mПРОВЕРКА ЗАВЕРШЕНА УСПЕШНО!\033[0m")
            print(f"="*100)
            
            if specific_table:
                report_name = f"quality_check_report_{self._safe_filename_token(specific_table)}.xlsx"
            elif table_list and len(table_list) == 1:
                report_name = f"quality_check_report_{self._safe_filename_token(table_list[0])}.xlsx"
            else:
                report_name = "quality_check_report.xlsx"
                
            print(f"   Отчет: {os.path.join(self.output_dir, report_name)}")
            
            if self.rule_errors:
                print(f"   Ошибки: {os.path.join(self.output_dir, 'errors')} ({len(self.rule_errors)} файлов)")
            # Напоминание: по ADR2 конкретные партнёры — только в БД
            adr2_with_data = [r for r in (self.results or []) if str(r.get("table_name", "")).strip().upper() == "ADR2" and r.get("filtered_adr2_count")]
            if adr2_with_data:
                print(f"   ADR2 — список партнёров по правилам: таблица «{self.ADR2_RULE_PARTNERS_TABLE}» в БД (rule_code, run_ts, ADDRNUMBER, PARTNER)")
        else:
            print(f"\n[INFO] Нет результатов проверки")
    
    def list_available_tables(self):
        """Выводит список таблиц, доступных для проверки."""
        rules_config = self.load_configuration()
        if not rules_config:
            print(f"\n[ERROR] Не удалось загрузить конфигурацию правил")
            return []
        tables = list(rules_config.keys())
        print(f"\n[INFO] Доступные таблицы для проверки:")
        print(f"="*50)
        for i, table in enumerate(tables, 1):
            rule_count = len(rules_config[table])
            print(f"{i:3}. {table:25} - {rule_count:3} правил")
        print(f"="*50)
        print(f"[INFO] Всего таблиц: {len(tables)}")
        return tables
    
    # Постоянные таблицы по taxtype (DFKKBPTAXNUM1, 2, 3, 5). Общие правила 52.x в rules.json — один раз на DFKKBPTAXNUM1.
    DFKKBPTAXNUM_ALIASES = ("DFKKBPTAXNUM1", "DFKKBPTAXNUM2", "DFKKBPTAXNUM3", "DFKKBPTAXNUM5")
    DFKKBPTAXNUM_SHARED_RULE_CODES = frozenset({
        "RCCONF_52.4", "RCCONF_52.3", "RCCONF_52.2", "RCCOMP_52.2",
    })
    # Таблицы, для которых «всего записей» = уникальные PARTNER (один клиент считается один раз, дубли исключаются)
    TABLE_UNIQUE_PARTNER = ("ZBUT0000P3VVI9", "ZBUT0000P", "ZBUT0000P3VV19")

    # AUSP в интерфейсе показывается как одна таблица; внутри — производные AUSP_143, AUSP_604, AUSP_148, AUSP_151
    AUSP_TABLE_GROUP = ("AUSP_143", "AUSP_604", "AUSP_148", "AUSP_151")

    def _expand_ausp_for_load(self, table_names):
        """Если в списке есть AUSP — под его загрузку подставляем производные (memory_manager загрузит AUSP)."""
        if not table_names:
            return table_names
        out = []
        for t in table_names:
            if t == "AUSP":
                out.extend(self.AUSP_TABLE_GROUP)
            else:
                out.append(t)
        # Правила BUT0BK/BUT051/KNB1/... джойнятся к KNA1.KUNNR (в выгрузке — Customer).
        kna1_dependent = {
            "BUT0BK", "BUT051", "KNB1", "KNVV", "KNVP", "KNVH", "ADR2", "ADRC", "BUT050",
        }
        if any(str(t).strip().upper() in kna1_dependent for t in out) and "KNA1" not in out:
            out.append("KNA1")
        return out

    def get_table_rules(self, table_name: str):
        """Получает правила для конкретной таблицы. Для DFKKBPTAXNUM — объединённые RU1/RU2/RU3/RU5. Для AUSP — объединённые AUSP_143/604/148/151."""
        rules_config = self.load_configuration()
        if table_name in rules_config:
            rules = list(rules_config[table_name])
            # Общие tax-правила (52.x) в rules.json один раз на DFKKBPTAXNUM1 — подмешиваем для 2/3/5
            if (
                table_name in self.DFKKBPTAXNUM_ALIASES
                and table_name != "DFKKBPTAXNUM1"
            ):
                have = {str(r.get("rule_code") or "").strip() for r in rules}
                for shared in rules_config.get("DFKKBPTAXNUM1", []):
                    code = str(shared.get("rule_code") or "").strip()
                    if code in self.DFKKBPTAXNUM_SHARED_RULE_CODES and code not in have:
                        rules.append(shared)
                        have.add(code)
            return rules
        if table_name == "DFKKBPTAXNUM":
            combined = []
            for alias in self.DFKKBPTAXNUM_ALIASES:
                combined.extend(rules_config.get(alias, []))
            return combined
        if table_name == "AUSP":
            combined = []
            for t in self.AUSP_TABLE_GROUP:
                combined.extend(rules_config.get(t, []))
            return combined if combined else []
        print(f"\n[ERROR] Таблица '{table_name}' не найдена в конфигурации")
        return []

    def _get_validator_for_rule(self, rule_description, quality_category, rule_info):
        rule_desc_lower = rule_description.lower()
        rule_code_raw = str(rule_info.get('rule_code', ''))
        rule_code = re.sub(r"[^A-Za-z0-9._-]", "", rule_code_raw).upper()
        
        # RCCONF_119.2: согласованность Terms of Payment между KNB1 и KNVV
        # Защита: даже при нестандартном коде/пробелах выбираем спец-валидатор по смыслу правила.
        if (
            rule_code == "RCCONF_119.2"
            or (
                "payment terms" in rule_desc_lower
                and "knb1" in rule_desc_lower
                and "knvv" in rule_desc_lower
            )
        ):
            # Ошибки сохраняет сам checker в errors/<timestamp>/..., поэтому error_saver не передаём
            return PaymentTermsConsistencyValidator(rule_info)

        # RCCONF_113.1: согласованность Recon. Account и Account Group по conf_recon_accounts
        if (
            rule_code == "RCCONF_113.1"
            or ("recon" in rule_desc_lower and "account group" in rule_desc_lower)
        ):
            return ReconAccountConsistencyValidator(rule_info)
        
        if rule_code == "RCCONF_63.1":
            return ConformityValidator(rule_info)

        if rule_code == 'RCCONF_15.1':
            return LogicalValidator(rule_info, self.error_manager)
        
        # Правила полноты (заполненность поля): только проверка «есть значение», без проверки формата/корректности
        if rule_code in ["RCCOMP_375.1", "RCCOMP_375.1.2"]:
            return CompletenessValidator(rule_info)
        if rule_code in ['RCCONF_18.2', 'RCCONF_22.2']:
            return AdvancedSpecialCharactersValidator(rule_info)
        
        # Правила проверки равенства
        if ("check if" in rule_desc_lower and "equals" in rule_desc_lower):
            return CrossColumnEqualityCheckValidator(rule_info)
        
        # Правила сравнения колонок
        elif ("cannot be the same" in rule_desc_lower or "cannot be a the same" in rule_desc_lower):
            return CrossColumnEqualityValidator(rule_info)
        
        # Русские правила равенства
        elif ("равен" in rule_desc_lower or "равны" in rule_desc_lower):
            return CrossColumnEqualityValidator(rule_info)
        
        # Правила для consecutive spaces
        elif ("недопустимые пробелы" in rule_desc_lower or 
              "consecutive space" in rule_desc_lower or 
              "two or more consecutive" in rule_desc_lower):
            return ConsecutiveSpacesValidator(rule_info)
        
        elif "специальные символы" in rule_desc_lower or "special character" in rule_desc_lower:
            return SpecialCharactersValidator(rule_info)
        
        elif "верхний регистр" in rule_desc_lower or "uppercase" in rule_desc_lower:
            return UppercaseValidator(rule_info)
        
        elif "отсутствует" in rule_desc_lower or "missing" in rule_desc_lower:
            return CompletenessValidator(rule_info)
        
        else:
            return ConformityValidator(rule_info)
    
    def _find_kunnr_column(self, df):
        """Возвращает имя колонки KUNNR (или аналога: Customer, MC_KUNNR и т.д.) в df или None."""
        if df is None or df.empty or not hasattr(df, 'columns'):
            return None
        col_upper_map = {str(col).strip().upper(): col for col in df.columns}
        # Точное совпадение
        if 'KUNNR' in col_upper_map:
            return col_upper_map['KUNNR']
        # Содержит KUNNR (MC_KUNNR, KUNNR_KNA1 и т.д.)
        for cu, col in col_upper_map.items():
            if 'KUNNR' in cu:
                return col
        # Часто используемые маппинги
        for candidate in ('CUSTOMER', 'CUSTOMER_CODE', 'MC_CUSTOMER', 'KUNNR_KNA1', 'KUNN'):
            if candidate in col_upper_map:
                return col_upper_map[candidate]
        return None

    def _norm_customer_partner_key(self, v):
        """Ключ для JOIN KNA1.KUNNR <-> BUT0BK.PARTNER / Business_Partner (нормализация Excel/SQLite)."""
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).replace("\ufeff", "").replace("\u00a0", " ").strip().strip("'").strip('"').strip()
        if s.lower() in {"", "none", "null", "nan", "<na>", "nat", "-", ".", "n/a", "na"}:
            return ""
        if re.fullmatch(r"\d+\.0+", s):
            s = s.split(".")[0]
        digits = re.sub(r"\D", "", s)
        if not digits:
            return ""
        return digits.zfill(10)
    
    def _resolve_but0bk_partner_column(self, df, table_name="BUT0BK"):
        """BUT0BK: Business_Partner (выгрузка) = SAP PARTNER = KNA1.KUNNR (Customer)."""
        if df is None or df.empty:
            return None
        try:
            from utils.column_map_resolver import resolve_column_in_df
            for sap in ("PARTNER",):
                col = resolve_column_in_df(df, sap, table_name, self.column_map, parent_dir)
                if col:
                    return col
        except ImportError:
            pass
        for name in ("PARTNER", "Business_Partner", "BUSINESS_PARTNER"):
            for c in df.columns:
                if str(c).strip().upper() == name:
                    return c
        return self._find_partner_column(df, table_name=table_name)

    def _find_account_group_column(self, df):
        """KTOKD / account_group_code после JOIN с KNA1."""
        if df is None or df.empty:
            return None
        for c in df.columns:
            cu = str(c).strip().lower()
            if cu in (
                "account_group_code",
                "b.account_group_code",
                "ktokd",
                "b.ktokd",
                "kna.ktokd",
                "group_1",
            ):
                return c
        return None

    def _filter_rows_only_ktokd_9038(self, df, rule_code):
        """
        RCCOMP_113.1 / RCCOMP_115.1:
        IF account_group_code != '9038' THEN '' (пропуск).
        Источник KTOKD: KNA1.Group_1 после JOIN KNB1.Customer = KNA1.KUNNR.
        """
        from utils.sap_account_keys import norm_sap_account_group

        if df is None or df.empty:
            return df
        ag_col = self._find_account_group_column(df)
        if not ag_col:
            print(f"      [WARN] {rule_code}: account_group_code/KTOKD не найден — scope «только 9038» невозможен")
            return df.iloc[0:0].copy()
        ag_norm = df[ag_col].apply(norm_sap_account_group)
        mask = ag_norm == "9038"
        matched = int(mask.sum())
        filled_ktokd = int((ag_norm != "").sum())
        self._last_kna1_join_stats = {
            "rows_after_join": len(df),
            "filled_ktokd": filled_ktokd,
            "n9038": matched,
        }
        print(
            f"      [FILTER] {rule_code}: KTOKD=9038 (KNA1) -> {matched:,} из {len(df):,} "
            f"(с заполненным KTOKD: {filled_ktokd:,})"
        )
        if matched == 0 and len(df) > 0:
            top = ag_norm[ag_norm != ""].value_counts().head(8)
            if not top.empty:
                print(f"      [FILTER] топ KTOKD после JOIN KNA1: {top.to_dict()}")
            else:
                print(
                    f"      [WARN] {rule_code}: после JOIN KNA1 колонка {ag_col} пуста у всех "
                    f"{len(df):,} строк — проверьте ключ Customer/KUNNR"
                )
        return df[mask].copy()

    def _resolve_kna1_kunnr_column(self, df):
        """KNA1: Customer / KUNN (выгрузка) = SAP KUNNR."""
        if df is None or df.empty:
            return None
        try:
            from utils.column_map_resolver import resolve_column_in_df
            col = resolve_column_in_df(df, "KUNNR", "KNA1", self.column_map, parent_dir)
            if col:
                return col
        except ImportError:
            pass
        return self._find_kunnr_column(df) or next(
            (c for c in df.columns if str(c).strip().upper() in ("KUNNR", "CUSTOMER", "KUNN")),
            None,
        )
    
    def _find_partner_column(self, df, table_name=None):
        """Возвращает имя колонки с партнёром для ZBUT0000P3VVI9: всего записей = количество уникальных PARTNER (дубли удаляются)."""
        if df is None or df.empty or not hasattr(df, 'columns'):
            return None
        cols = list(df.columns)
        # 1) column_map для этой таблицы
        if table_name and getattr(self, 'column_map', None) and table_name in self.column_map:
            for key in ('partner', 'partner_column', 'PARTNER', 'partners'):
                if key in self.column_map[table_name]:
                    phys = self.column_map[table_name][key]
                    if phys and str(phys).strip() in cols:
                        return str(phys).strip()
                    for c in cols:
                        if str(c).strip().upper() == str(phys).strip().upper():
                            return c
        # 2) Конфиг conf_zbut0000p_partner.json
        if table_name and table_name in getattr(self, 'TABLE_UNIQUE_PARTNER', ()):
            config_col = self._load_partner_column_config(table_name)
            if config_col:
                config_upper = config_col.strip().upper()
                for c in cols:
                    if str(c).strip().upper() == config_upper:
                        return c
                if config_col in cols:
                    return config_col
        col_upper = {str(c).strip().upper(): c for c in cols}
        tn = str(table_name or "").strip().upper()
        # KNVP: ключ клиента — Customer/KUNNR, не колонка Partner (часто 0)
        if tn == "KNVP":
            try:
                from utils.column_map_resolver import resolve_column_in_df
                for sap in ("KUNNR", "Customer"):
                    col = resolve_column_in_df(df, sap, table_name, self.column_map, parent_dir)
                    if col:
                        return col
            except ImportError:
                pass
            for name in ("KUNNR", "CUSTOMER", "CUSTOMER_1"):
                if name in col_upper:
                    return col_upper[name]
        if tn == "BUT0BK":
            for name in ("PARTNER", "BUSINESS_PARTNER", "BUSINESS PARTNER"):
                if name in col_upper:
                    return col_upper[name]
        for name in ('PARTNER', 'PARTNERS', 'PARTNER_ID', 'PARTNER_NUM', 'BP', 'CUSTOMER', 'KUNNR', 'PARTNER_CODE', 'CUSTOMER_ID', 'BP_NUMBER'):
            if name in col_upper:
                return col_upper[name]
        if tn != "BUT0BK":
            for name in ('CLIENT',):
                if name in col_upper:
                    return col_upper[name]
        for cu, col in col_upper.items():
            if 'PARTNER' in cu or cu.startswith('PARTNER') or 'CLIENT' in cu or 'KUNNR' in cu or cu == 'BP':
                return col
        for col in cols:
            cu = str(col).upper()
            if 'PARTNER' in cu or 'CUSTOMER' in cu or 'KUNNR' in cu or 'BP' in cu or 'CUST' in cu or 'CLIENT' in cu:
                return col
        # 3) Для ZBUT0000P3VVI9: fallback — колонка с макс. числом уникальных значений (обычно ключ/партнёр)
        if table_name and table_name in getattr(self, 'TABLE_UNIQUE_PARTNER', ()) and len(cols) > 0:
            try:
                best_col = None
                best_nunique = 0
                for c in cols:
                    n = df[c].nunique()
                    if n > best_nunique and n <= len(df) and n > 1:
                        best_nunique = n
                        best_col = c
                if best_col:
                    return best_col
            except Exception:
                pass
        return None

    def _scope_but0bk_to_kna1_partners(self, df, table_name, rule_code):
        """
        BUT0BK scope: оставляем только строки, где партнёр существует в KNA1 (KUNNR).
        JOIN: BUT0BK.PARTNER (Business_Partner в выгрузке) = KNA1.KUNNR (Customer в выгрузке).
        """
        try:
            if df is None or df.empty:
                return df

            partner_col = self._resolve_but0bk_partner_column(df, table_name=table_name)
            if not partner_col:
                print(f"      [WARN] {rule_code}: в {table_name} не найдена колонка партнёра (PARTNER/Business_Partner)")
                return df.iloc[0:0].copy()

            try:
                kna1_df = self._get_table_for_rules("KNA1")
            except Exception:
                kna1_df = None
            if kna1_df is None or kna1_df.empty:
                try:
                    if hasattr(self.memory_manager, "load_selected_tables_to_ram"):
                        self.memory_manager.load_selected_tables_to_ram(["KNA1"], add_reference_tables=False)
                    kna1_df = self._get_table_for_rules("KNA1")
                except Exception:
                    kna1_df = None

            if kna1_df is None or kna1_df.empty:
                try:
                    conn = connect_sqlite(self.db_path)
                    try:
                        kna1_df = pd.read_sql_query(
                            'SELECT "Customer" AS "KUNNR" FROM "KNA1"', conn
                        )
                    except Exception:
                        try:
                            kna1_df = pd.read_sql_query('SELECT "KUNNR" FROM "KNA1"', conn)
                        except Exception:
                            kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                    kna1_df = self._apply_rule_time_column_map(kna1_df, "KNA1")
                except Exception as e:
                    print(f"      [WARN] {rule_code}: не удалось загрузить KNA1 для scope: {e}")
                    return df.iloc[0:0].copy()

            if kna1_df is None or kna1_df.empty:
                print(f"      [WARN] {rule_code}: KNA1 пуста, scope невозможен")
                return df.iloc[0:0].copy()

            kna1_kunnr_col = self._resolve_kna1_kunnr_column(kna1_df)
            if not kna1_kunnr_col:
                print(f"      [WARN] {rule_code}: в KNA1 не найдена колонка KUNNR/Customer для scope")
                return df.iloc[0:0].copy()

            left = df[partner_col].apply(self._norm_customer_partner_key)
            right = kna1_df[kna1_kunnr_col].apply(self._norm_customer_partner_key)
            kna1_keys = set(right[right != ""].unique().tolist())
            if not kna1_keys:
                return df.iloc[0:0].copy()

            mask = left.isin(kna1_keys)
            matched = int(mask.sum())
            print(
                f"      [JOIN] {rule_code}: {table_name}.{partner_col} -> KNA1.{kna1_kunnr_col}, "
                f"совпало строк: {matched:,}"
            )
            return df[mask].copy()

        except Exception as e:
            print(f"      [WARN] {rule_code}: ошибка scope BUT0BK по KNA1 партнёрам: {e}")
            return df.iloc[0:0].copy()
    
    def _get_reference_table_for_rule(self, rule_code, config_key):
        """Читает conf_sales_group_office.json и возвращает имя справочной таблицы для правила (например RCCONF_143.7)."""
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [
            os.path.join(root, 'json files', 'conf_sales_group_office.json'),
            os.path.join(os.getcwd(), 'json files', 'conf_sales_group_office.json'),
            os.path.join(root, 'config', 'conf_sales_group_office.json'),
        ]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    name = cfg.get(config_key) or cfg.get('reference_table')
                    if name and isinstance(name, str) and name.strip():
                        return name.strip()
                except Exception:
                    pass
        return None
    
    def _load_allowed_vkgrp_vkbur_from_json(self):
        """Загружает допустимые комбинации (VKGRP, VKBUR) из conf_sales_group_office.json.
        Ключи в JSON: allowed_combinations (список объектов с полями VKGRP/VKBUR или sales_group_code/sales_office_code).
        Возвращает set of (vkgrp, vkbur) или None если файла/данных нет."""
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [
            os.path.join(root, 'json files', 'conf_sales_group_office.json'),
            os.path.join(os.getcwd(), 'json files', 'conf_sales_group_office.json'),
            os.path.join(root, 'config', 'conf_sales_group_office.json'),
        ]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    arr = cfg.get('allowed_combinations') or cfg.get('combinations')
                    if not arr or not isinstance(arr, list):
                        return None
                    result = set()
                    for item in arr:
                        if isinstance(item, dict):
                            vg = item.get('VKGRP') or item.get('vkgrp') or item.get('sales_group_code') or item.get('sales_group')
                            vb = item.get('VKBUR') or item.get('vkbur') or item.get('sales_office_code') or item.get('sales_office')
                            if vg is not None and vb is not None:
                                result.add((str(vg).strip(), str(vb).strip()))
                        elif isinstance(item, (list, tuple)) and len(item) >= 2:
                            result.add((str(item[0]).strip(), str(item[1]).strip()))
                    return result if result else None
                except Exception:
                    pass
        return None

    def _load_vkorg_cluster_scope_from_json(self):
        """Опциональный scope для RCCONF_143.7: допустимые пары (VKORG, Cluster) из conf_sales_group_office.json.
        Поддерживаемые ключи массива: vkorg_cluster_scope | allowed_vkorg_cluster | vkorg_cluster_combinations.
        Поля элементов: VKORG/vkorg/sales_org и CLUSTER/cluster/KVGR4/customer_group_4_code.
        Возвращает set[(vkorg, cluster)] или None.
        """
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [
            os.path.join(root, 'json files', 'conf_sales_group_office.json'),
            os.path.join(os.getcwd(), 'json files', 'conf_sales_group_office.json'),
            os.path.join(root, 'config', 'conf_sales_group_office.json'),
        ]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    arr = (
                        cfg.get('vkorg_cluster_scope')
                        or cfg.get('allowed_vkorg_cluster')
                        or cfg.get('vkorg_cluster_combinations')
                    )
                    if not arr or not isinstance(arr, list):
                        return None
                    out = set()
                    for item in arr:
                        if not isinstance(item, dict):
                            continue
                        vg = item.get('VKORG') or item.get('vkorg') or item.get('sales_org')
                        cl = item.get('CLUSTER') or item.get('cluster') or item.get('KVGR4') or item.get('customer_group_4_code')
                        vg = self._norm_lookup_value(vg).upper() if vg is not None else ""
                        cl = self._norm_lookup_value(cl).upper() if cl is not None else ""
                        if vg and cl:
                            out.add((vg, cl))
                    return out if out else None
                except Exception:
                    pass
        return None

    def _norm_lookup_value(self, value):
        """Нормализация значения для правил lookup/matrix."""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        s = str(value).replace("\u00a0", " ").replace("\ufeff", "").strip().strip("'").strip('"').strip()
        if s.lower() in {"", "none", "null", "nan", "<na>", "nat", "-", ".", "n/a", "na"}:
            return ""
        if s.endswith(".0"):
            s = re.sub(r"\.0+$", "", s)
        return s

    def _load_planning_group_matrix(self):
        """Загружает conf_planning_group_matrix.json и возвращает список правил матрицы."""
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [
            os.path.join(root, "json files", "conf_planning_group_matrix.json"),
            os.path.join(os.getcwd(), "json files", "conf_planning_group_matrix.json"),
            os.path.join(root, "config", "conf_planning_group_matrix.json"),
        ]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        cfg = json.load(f)
                    rules = cfg.get("rules") if isinstance(cfg, dict) else None
                    if isinstance(rules, list) and rules:
                        return rules, path
                except Exception:
                    pass
        return None, None

    def _add_customer_group_4_from_knvv(self, df, table_name, rule_code):
        """
        Добавляет customer_group_4_code (KVGR4 из KNVV) в DataFrame через JOIN по KUNNR.
        Для одного клиента берём первую непустую KVGR4 из KNVV.
        """
        if df is None or df.empty:
            return df
        try:
            existing = next(
                (c for c in df.columns if str(c).strip().lower() in ("customer_group_4_code", "kvgr4")),
                None,
            )
            if existing:
                return df

            kunnr_col = self._find_kunnr_column(df) or next(
                (c for c in df.columns if str(c).strip().upper() == "KUNNR"),
                None,
            )
            if not kunnr_col:
                print(f"      [WARN] Колонка KUNNR не найдена в {table_name} для правила {rule_code}")
                return df

            knvv_df = self.memory_manager.get_table("KNVV")
            if (knvv_df is None or knvv_df.empty) and getattr(self, "db_path", None):
                try:
                    self.memory_manager.load_selected_tables_to_ram(["KNVV"], add_reference_tables=False)
                    knvv_df = self.memory_manager.get_table("KNVV")
                except Exception:
                    knvv_df = self.memory_manager.get_table("KNVV")

            if knvv_df is None or knvv_df.empty:
                print(f"      [WARN] Таблица KNVV не найдена или пуста для правила {rule_code}")
                return df

            knvv_kunnr_col = next((c for c in knvv_df.columns if str(c).strip().upper() == "KUNNR"), None) or self._find_kunnr_column(knvv_df)
            kvgr4_col = next(
                (c for c in knvv_df.columns if str(c).strip().upper() in ("KVGR4", "CUSTOMER_GROUP_4_CODE")),
                None,
            ) or self._find_column_alternative(knvv_df.columns, "KVGR4", "KNVV")

            if not knvv_kunnr_col or not kvgr4_col:
                print(f"      [WARN] В KNVV не найдены KUNNR/KVGR4 для правила {rule_code}")
                return df

            def _norm_kunnr(series):
                s = series.astype(str).str.strip()
                s = s.str.replace(r"\.0$", "", regex=True)
                s = s.str.replace(r"\D+", "", regex=True)
                return s.str.zfill(10)

            knvv_join = knvv_df[[knvv_kunnr_col, kvgr4_col]].copy()
            kvgr4_norm = knvv_join[kvgr4_col].apply(self._norm_lookup_value)
            knvv_join = knvv_join[kvgr4_norm != ""].copy()
            if knvv_join.empty:
                print(f"      [WARN] В KNVV нет заполненных KVGR4 для правила {rule_code}")
                return df

            knvv_join["_kunnr_key"] = _norm_kunnr(knvv_join[knvv_kunnr_col])
            knvv_join["customer_group_4_code"] = knvv_join[kvgr4_col].apply(self._norm_lookup_value)
            knvv_join = knvv_join.drop_duplicates(subset=["_kunnr_key"], keep="first")

            df_joined = df.copy()
            df_joined["_kunnr_key"] = _norm_kunnr(df_joined[kunnr_col])
            df_joined = df_joined.merge(
                knvv_join[["_kunnr_key", "customer_group_4_code"]],
                on="_kunnr_key",
                how="left",
            ).drop(columns=["_kunnr_key"], errors="ignore")

            filled = int(df_joined["customer_group_4_code"].apply(lambda v: self._norm_lookup_value(v) != "").sum())
            print(
                f"      [JOIN] Добавлен customer_group_4_code из KNVV (KVGR4) для правила {rule_code}. "
                f"Строк: {len(df_joined)} (было {len(df)}), заполнено customer_group_4_code: {filled} ({(filled / max(len(df_joined), 1)) * 100:.1f}%)"
            )
            return df_joined
        except Exception as e:
            print(f"      [WARN] Ошибка добавления customer_group_4_code из KNVV для {rule_code}: {e}")
            return df
    
    def _load_partner_column_config(self, table_name):
        """Читает conf_zbut0000p_partner.json и возвращает имя колонки партнёра для подсчёта без дублей."""
        if table_name not in getattr(self, 'TABLE_UNIQUE_PARTNER', ()):
            return None
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [
            os.path.join(root, 'json files', 'conf_zbut0000p_partner.json'),
            os.path.join(os.getcwd(), 'json files', 'conf_zbut0000p_partner.json'),
            os.path.join(root, 'config', 'conf_zbut0000p_partner.json'),
        ]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    col = cfg.get('partner_column') or cfg.get('partner_column_name')
                    if col and isinstance(col, str) and col.strip():
                        return col.strip()
                except Exception:
                    pass
        return None
    
    def _apply_unique_partner_counts_if_needed(self, table_name, df, error_df, total_rows, error_count):
        """Для ZBUT0000P*: «всего записей» = уникальные PARTNER (не для KNVP — там много строк на клиента)."""
        if table_name not in self.TABLE_UNIQUE_PARTNER:
            return total_rows, error_count
        partner_col = self._find_partner_column(df, table_name=table_name)
        if not partner_col:
            if not hasattr(self, '_partner_warned_tables'):
                self._partner_warned_tables = set()
            if table_name not in self._partner_warned_tables:
                self._partner_warned_tables.add(table_name)
                print(f"      [WARN] [{table_name}] Колонка PARTNER не найдена — подсчёт без удаления дублей (будут миллионы строк). Укажите имя колонки в conf_zbut0000p_partner.json (partner_column). Колонки: {list(df.columns)[:30]}")
            return total_rows, error_count
        # Всего записей = только уникальные PARTNER (дубли не считаем)
        total_rows = int(df[partner_col].nunique())
        if error_df is None or error_df.empty:
            error_count = 0
        elif partner_col in error_df.columns:
            error_count = int(error_df[partner_col].nunique())
        else:
            try:
                common = error_df.index.intersection(df.index)
                if len(common) >= len(error_df) * 0.99:
                    error_count = int(df.loc[error_df.index, partner_col].nunique())
                else:
                    error_count = int(error_count)
            except Exception:
                error_count = int(error_count)
        print(f"      [{table_name}] Подсчёт по уникальным {partner_col} (дубли исключены): всего {total_rows:,} клиентов, с ошибками: {error_count:,}")
        return total_rows, error_count
    
    def _find_column_alternative(self, columns, column_name, table_name):
        try:
            from utils.column_map_resolver import resolve_column, map_logical_to_sap
            sap = map_logical_to_sap(table_name, column_name, self.column_map, parent_dir)
            for target in (sap, column_name):
                found = resolve_column(
                    columns,
                    target,
                    table_name,
                    self.column_map,
                    parent_dir,
                )
                if found:
                    return found
        except ImportError:
            pass
        return None
    
    def _find_most_similar_column(self, columns, target_column):
        target_upper = target_column.upper().replace('_', '')
        
        best_match = None
        best_score = 0
        
        for col in columns:
            col_upper = col.upper().replace('_', '')
            if len(col_upper) < 4:
                continue
            
            score = 0
            
            if col_upper == target_upper:
                score += 100
            
            if len(target_upper) >= 4 and (target_upper in col_upper or col_upper in target_upper):
                score += 50
            
            common = set(target_upper) & set(col_upper)
            if len(common) > 0:
                score += len(common) * 2
            
            if score > best_score:
                best_score = score
                best_match = col
        
        return best_match if best_score > 10 else None
    
    def _extract_second_column_from_description(self, rule_code, rule_description, columns, first_column, table_name=None):
        """Извлекает вторую колонку из описания правила с учетом маппинга"""
        desc_lower = rule_description.lower()
        
        patterns = {
            'name 1': ['name_org1', 'name1', 'name_1', 'organization_1_name'],
            'name 2': ['name_org2', 'name2', 'name_2', 'organization_2_name'],
            'name 3': ['name_org3', 'name3', 'name_3', 'organization_3_name'],
            'name 4': ['name_org4', 'name4', 'name_4', 'organization_4_name'],
            'tax 1': ['taxnum1', 'taxnum', 'tax_1_value'],
            'tax 2': ['taxnum2', 'taxnum', 'tax_2_value'],
            'tax 3': ['taxnum3', 'taxnum', 'tax_3_value'],
            'tax 4': ['taxnum4', 'taxnum', 'tax_4_value'],
            'tax 5': ['taxnum5', 'taxnum', 'tax_5_value'],
            'tax 6': ['taxnum6', 'taxnum', 'tax_6_value'],
        }
        
        second_column_candidate = None
        
        for key, variations in patterns.items():
            for variation in variations:
                if variation in desc_lower:
                    if variation not in first_column.lower():
                        second_column_candidate = variation
                        break
            if second_column_candidate:
                break
        
        if not second_column_candidate:
            if "Tax Number 1 и Tax Number 2" in rule_description:
                if "TAXNUM" in first_column.upper():
                    second_column_candidate = "TAXNUM1" if "TAXNUM2" in first_column.upper() else "TAXNUM2"
            elif "Name 3 и Name 4" in rule_description:
                if "NAME" in first_column.upper():
                    second_column_candidate = "NAME4" if "NAME3" in first_column.upper() else "NAME3"
            elif "Street и House Number" in rule_description:
                if "STRAS" in first_column.upper():
                    second_column_candidate = "ORT01"
                elif "ORT01" in first_column.upper():
                    second_column_candidate = "STRAS"
        
        if not second_column_candidate:
            return None
    
        # Применяем маппинг
        if table_name:
            mapped_second = self._get_mapped_column_name(table_name, second_column_candidate)
            second_column_candidate = mapped_second
        
        # Ищем колонку среди доступных
        for col in columns:
            if col.upper() == second_column_candidate.upper():
                return col
        
        # Пробуем альтернативные варианты
        matched = self._find_column_alternative(columns, second_column_candidate, table_name)
        if matched:
            return matched
        
        return None

    def _apply_conditional_filter(self, df, technical_def, rule_code, table_name=None):
        """Применяет фильтрацию данных на основе условий"""
        try:
            print(f"      [FILTER] Анализ условий для {rule_code}...")
            
            # Для правил RCCONF_39.5.2 и RCCONF_39.3.2 фильтруем по PERSNUMBER (должно быть ЗАПОЛНЕНО), затем по KNVV AUFSD F/M для ADR2
            if rule_code in ["RCCONF_39.5.2", "RCCONF_39.3.2"]:
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                        break
                if person_col:
                    mask = (
                        df[person_col].notna() &
                        (df[person_col].astype(str).str.strip() != '') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'none') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'null')
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL (заполнено) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации (PERSNUMBER заполнено) данных нет!")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                else:
                    print(f"      [WARN] Колонка PERSNUMBER не найдена для правила {rule_code}")
            
            # Для правила RCCONF_39.5 (ADR2): только пустой PERSNUMBER, затем фильтр по KNVV AUFSD F/M
            if rule_code == "RCCONF_39.5" and table_name and str(table_name).strip().upper() == "ADR2":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк")
                if not df.empty:
                    return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)
                return df

            # RCCONF_39.3: только PERSNUMBER IS NULL (без R3_USER / KNVV — см. rules.json)
            if rule_code == "RCCONF_39.3":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower in ("persnumber", "pers_number") or col_lower == "person" or "persnumber" in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                        break
                if person_col:
                    mask = (
                        df[person_col].isna()
                        | (df[person_col].astype(str).str.strip() == "")
                        | (df[person_col].astype(str).str.strip().str.lower().isin(("none", "null", "nan")))
                    )
                    filtered_df = df[mask].copy()
                    print(
                        f"      [FILTER] RCCONF_39.3: PERSNUMBER IS NULL — "
                        f"{len(filtered_df)} из {len(df)} строк"
                    )
                    return filtered_df
                print(f"      [WARN] RCCONF_39.3: колонка PERSNUMBER не найдена")
                return df
            
            # Для правила RCCONF_38.3: R3_USER = '1' (стационарный телефон) AND PERSNUMBER IS NULL
            # contact_medium_type = 'fixed_tel_number' когда R3_USER = '1'
            if rule_code == "RCCONF_38.3":
                r3_user_col = None
                person_col = None
                contact_medium_col = None
                
                # Ищем колонку R3_USER (определяет тип телефона: '1' = стационарный, '3' = мобильный)
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or 'r3user' in col_lower:
                        r3_user_col = col
                        print(f"      [FILTER] Найдена колонка R3_USER: {col}")
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                    if 'contact_medium' in col_lower or 'medium_type' in col_lower or 'contactmedium' in col_lower:
                        contact_medium_col = col
                        print(f"      [FILTER] Найдена колонка contact_medium_type: {col}")
                
                # Приоритет: используем R3_USER = '1' для определения стационарного телефона
                if r3_user_col and person_col:
                    mask = (
                        (df[r3_user_col].astype(str).str.strip() == '1') &  # R3_USER = '1' означает стационарный телефон
                        (df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none'))
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (R3_USER='1' AND PERSNUMBER IS NULL) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации данных нет!")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif r3_user_col:
                    # Если нет PERSNUMBER, фильтруем только по R3_USER = '1'
                    mask = (df[r3_user_col].astype(str).str.strip() == '1')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только R3_USER='1') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif contact_medium_col and person_col:
                    # Fallback: используем contact_medium_type если R3_USER не найден
                    mask = (
                        (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number') &
                        (df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none'))
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (contact_medium_type='fixed_tel_number' AND PERSNUMBER IS NULL) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации данных нет!")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif contact_medium_col:
                    # Если нет PERSNUMBER, фильтруем только по contact_medium_type
                    mask = (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только contact_medium_type='fixed_tel_number') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif person_col:
                    mask = (
                        df[person_col].isna()
                        | (df[person_col].astype(str).str.strip() == "")
                        | (df[person_col].astype(str).str.strip().str.lower().isin(("none", "null", "nan")))
                    )
                    filtered_df = df[mask].copy()
                    print(
                        f"      [FILTER] RCCONF_38.3: PERSNUMBER IS NULL (fallback без R3_USER) — "
                        f"{len(filtered_df)} из {len(df)} строк"
                    )
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                else:
                    print(f"      [WARN] RCCONF_38.3: колонки R3_USER / PERSNUMBER / contact_medium_type не найдены")
            
            # Для правила RCCONF_38.5 фильтруем по PERSNUMBER (должно быть пустым)
            elif rule_code == "RCCONF_38.5":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                        break
                
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации (PERSNUMBER IS NULL) данных нет!")
                    return filtered_df
                else:
                    print(f"      [WARN] Колонка PERSNUMBER не найдена для правила {rule_code}")
            
            # Для правил BUT050: RCCOMP_369.1 и RCCONF_369.1
            if rule_code in ["RCCOMP_369.1", "RCCONF_369.1"]:
                source_col = None
                
                # Ищем колонку source (может быть source_file или source)
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'source' or col_lower == 'source_file' or 'source' in col_lower:
                        source_col = col
                        print(f"      [FILTER] Найдена колонка source: {col}")
                        break
                
                # Для RCCOMP_369.1: фильтруем по source = 's4'
                if rule_code == "RCCOMP_369.1":
                    if source_col:
                        mask = df[source_col].astype(str).str.strip().str.lower() == 's4'
                        filtered_df = df[mask].copy()
                        print(f"      [FILTER] Применен фильтр (source='s4') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                        if len(filtered_df) == 0:
                            print(f"      [WARN] После фильтрации (source='s4') данных нет!")
                        return filtered_df
                    else:
                        print(f"      [WARN] Колонка source не найдена для правила {rule_code}")
                        # Если source не найден, работаем со всеми данными
                        return df
                
                # Для RCCONF_369.1: фильтрация не требуется по source, но может быть по другим условиям
                # Пока возвращаем все данные, фильтрация будет в валидаторе
                elif rule_code == "RCCONF_369.1":
                    print(f"      [FILTER] Правило {rule_code} не требует фильтрации по source, работаем со всеми данными")
                    return df
            
            # Для правила RCCOMP_375.1 (ADR2): только пустой PERSNUMBER, затем фильтр по KNVV AUFSD F/M (BUT020->PARTNER->KNVV)
            elif rule_code == "RCCOMP_375.1":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    pers_str = df[person_col].astype(str).str.strip()
                    pers_low = pers_str.str.lower()
                    # "битые нули" в Excel иногда попадают как 0 / 0.0 / 0,0 / 000...
                    pers_str_for_zero = pers_str.str.replace(",", ".", regex=False)
                    pers_is_zeroish = pers_str_for_zero.str.match(r"^-?0+(?:[.][0]+)?$", na=False)
                    mask = (
                        df[person_col].isna()
                        | (pers_str == '')
                        | pers_low.isin(['none', 'null'])
                        | pers_is_zeroish
                    )
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк")
                    if len(df) == 0:
                        return df
                # В scope для 375.1 применяем фильтрацию по KNA1.AUFSD (blocked codes),
                # а не по KNVV AUFSD F/M.
                return self._filter_adr2_rccomp_375_1_scope_by_kna1_aufsd(df, rule_code, table_name)

            # Для правила RCCOMP_375.1.2 (ADR2): только заполненный PERSNUMBER, затем фильтр по KNVV AUFSD F/M
            elif rule_code == "RCCOMP_375.1.2":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    pers_str = df[person_col].astype(str).str.strip()
                    pers_low = pers_str.str.lower()
                    pers_is_zeroish = pers_str.str.match(r"^0+(\.0+)?$", na=False)
                    mask = (
                        df[person_col].notna() &
                        (pers_str != '') &
                        (~pers_low.isin(['none', 'null'])) &
                        (~pers_is_zeroish)
                    )
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL для {rule_code}: {len(df)} строк")
                    if len(df) == 0:
                        return df
                return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)
            
            # RCCONF_39.5 — проверка формата TEL_NUMBER только для тех же записей, что и RCCOMP_375.1 (пустой PERSNUMBER, KNVV F/M).
            # Пустые TEL_NUMBER не считаются ошибкой формата (их ловит RCCOMP_375.1).
            elif rule_code == "RCCONF_39.5":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк")
                    if len(df) == 0:
                        return df
                return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)
            
            # RCCONF_39.5.2 и RCCONF_39.3.2 — проверка формата/символов только для записей с заполненным PERSNUMBER (как RCCOMP_375.1.2).
            elif rule_code in ["RCCONF_39.5.2", "RCCONF_39.3.2"]:
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    mask = (
                        df[person_col].notna() &
                        (df[person_col].astype(str).str.strip() != '') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'none') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'null')
                    )
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL для {rule_code}: {len(df)} строк")
                    if len(df) == 0:
                        return df
                return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)
            
            return df
        
        except Exception as e:
            print(f"      [WARN] Ошибка в _apply_conditional_filter для {rule_code}: {e}")
            import traceback
            traceback.print_exc()
            return df
    
    def _get_adr6_df(self):
        """Загружает ADR6 из кэша или SQLite (для RCCOMP_375.1.2 — e-mail)."""
        try:
            adr6 = self.memory_manager.get_table("ADR6")
            if (adr6 is None or adr6.empty) and getattr(self, "db_path", None):
                conn = connect_sqlite(self.db_path)
                tables = pd.read_sql_query(
                    "SELECT name FROM sqlite_master WHERE type='table'", conn
                )
                adr6_name = next(
                    (r[0] for r in tables.values if str(r[0]).strip().upper() == "ADR6"),
                    None,
                )
                if adr6_name:
                    adr6 = pd.read_sql_query(f'SELECT * FROM "{adr6_name}"', conn)
                conn.close()
            if adr6 is None or adr6.empty:
                return None
            return self._apply_rule_time_column_map(adr6.copy(), "ADR6")
        except Exception as e:
            print(f"      [WARN] _get_adr6_df: {e}")
            return None

    def _ensure_adr2_has_partner(self, df, rule_code):
        """
        Для ADR2: если в df нет колонки PARTNER, добавляет её из BUT020 по ADDRNUMBER.
        Нужно, чтобы по каждому правилу ADR2 можно было выгрузить файл «всего записей» (ADDRNUMBER+PARTNER).
        """
        if df is None or df.empty:
            return df
        addr_col = next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)
        partner_col = next((c for c in df.columns if str(c).upper() == 'PARTNER'), None)
        if not addr_col or partner_col:
            return df
        try:
            but020 = self.memory_manager.get_table("BUT020")
            if (but020 is None or but020.empty) and getattr(self, 'db_path', None):
                import sqlite3
                conn = connect_sqlite(self.db_path)
                tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                but_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'BUT020'), None)
                if but_name:
                    but020 = pd.read_sql_query(f'SELECT * FROM "{but_name}"', conn)
                conn.close()
            if but020 is None or but020.empty:
                return df
            addr_but = next((c for c in but020.columns if 'ADDRNUMBER' in str(c).upper()), None)
            partner_but = next((c for c in but020.columns if str(c).upper() == 'PARTNER'), None)
            if not addr_but or not partner_but:
                return df
            j = lambda x: (str(x).strip().lstrip('0') or '0')
            df = df.copy()
            df['_ak'] = df[addr_col].apply(j)
            but_join = but020[[addr_but, partner_but]].copy()
            but_join['_ak'] = but_join[addr_but].apply(j)
            but_join = but_join.drop_duplicates(subset=['_ak'])
            df = df.merge(but_join[['_ak', partner_but]], on='_ak', how='left')
            df = df.drop(columns=['_ak'], errors='ignore')
            if partner_but not in df.columns:
                return df
            if partner_but != 'PARTNER':
                df = df.rename(columns={partner_but: 'PARTNER'})
            return df
        except Exception as e:
            print(f"      [WARN] _ensure_adr2_has_partner для {rule_code}: {e}")
            return df

    def _build_adr2_rccomp_37512_export_df(self, df_adr2, rule_code='RCCOMP_375.1.2'):
        """
        Строит выгрузку для RCCOMP_375.1.2 по цепочке в процессе:
        1) Фильтр: PERSNUMBER заполнен
        2) ADR2 → BUT020 → PARTNER, затем KNVV: только AUFSD in ('F','M')
        Возвращает DataFrame с колонками ADDRNUMBER, PARTNER, AUFSD (гарантированно).
        """
        if df_adr2 is None or df_adr2.empty:
            return None
        try:
            person_col = None
            for col in df_adr2.columns:
                c = str(col).lower()
                if c == 'persnumber' or c == 'pers_number' or 'persnumber' in c:
                    person_col = col
                    break
            if person_col is not None:
                mask = (
                    df_adr2[person_col].notna()
                    & (df_adr2[person_col].astype(str).str.strip() != '')
                    & (df_adr2[person_col].astype(str).str.strip().str.lower() != 'none')
                    & (df_adr2[person_col].astype(str).str.strip().str.lower() != 'null')
                )
                df_adr2 = df_adr2[mask].copy()
                if df_adr2.empty:
                    return None
            df = self._filter_adr2_by_knvv_aufsd_fm(df_adr2, rule_code, 'ADR2')
            if df is None or df.empty:
                return None
            addr_col = next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)
            partner_col = next((c for c in df.columns if str(c).upper() == 'PARTNER'), None)
            if not addr_col:
                partner_col = next((c for c in df.columns if 'PARTNER' in str(c).upper()), None)
            aufsd_col = next((c for c in df.columns if 'AUFSD' in str(c).upper()), None)
            if not addr_col or not partner_col:
                return None
            cols = [addr_col, partner_col]
            if aufsd_col:
                cols.append(aufsd_col)
            out = df[cols].copy()
            out.columns = ['ADDRNUMBER', 'PARTNER', 'AUFSD'] if aufsd_col else ['ADDRNUMBER', 'PARTNER']
            if not aufsd_col:
                out['AUFSD'] = ''
            return out
        except Exception as e:
            print(f"      [WARN] _build_adr2_rccomp_37512_export_df: {e}")
            return None
    
    def _filter_adr2_by_knvv_aufsd_fm(self, df, rule_code, table_name=None):
        """
        Для правил ADR2 (RCCOMP_375.1, RCCONF_39.5, RCCONF_39.5.2):
        ADR2.ADDRNUMBER -> BUT020 -> PARTNER; KNVV: KUNNR=PARTNER, AUFSD in ('F','M').
        Оставляет только строки ADR2, у которых PARTNER в KNVV с блоками F или M. Добавляет колонку AUFSD.
        """
        if table_name and str(table_name).strip().upper() != "ADR2":
            return df
        try:
            def _norm_key(series: pd.Series) -> pd.Series:
                """
                Нормализация ключа JOIN в стиле telbd:
                - trim
                - убрать хвост '.0'
                - оставить только цифры
                - zfill(10)
                """
                s = series.astype(str).str.strip()
                s = s.str.replace(r"\.0$", "", regex=True)
                s = s.str.replace(r"\D+", "", regex=True)
                return s.str.zfill(10)

            addr_col = next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)
            if not addr_col:
                print(f"      [FILTER] Колонка ADDRNUMBER не найдена в ADR2 для {rule_code}")
                return df
            but020 = self.memory_manager.get_table("BUT020")
            if but020 is None or but020.empty:
                print(f"      [WARN] Таблица BUT020 не найдена или пуста для {rule_code}")
                return df
            addr_but = next((c for c in but020.columns if 'ADDRNUMBER' in str(c).upper()), None)
            partner_but = next((c for c in but020.columns if str(c).upper() == 'PARTNER'), None)
            if not addr_but or not partner_but:
                print(f"      [WARN] В BUT020 не найдены ADDRNUMBER или PARTNER для {rule_code}")
                return df
            df = df.copy()
            df['_ak'] = _norm_key(df[addr_col])
            but_join = but020[[addr_but, partner_but]].copy()
            but_join['_ak'] = _norm_key(but_join[addr_but])
            but_join = but_join.drop_duplicates(subset=['_ak'], keep='first')
            df = df.merge(but_join[['_ak', partner_but]], on='_ak', how='left')
            df = df.drop(columns=['_ak'], errors='ignore')
            partner_col = partner_but
            if partner_col not in df.columns:
                print(f"      [WARN] PARTNER не добавлен из BUT020 для {rule_code}")
                return df
            knvv = self.memory_manager.get_table("KNVV")
            if knvv is None or knvv.empty:
                # Подгружаем KNVV из БД, если не загружена (например, при проверке только ADR2)
                if getattr(self, 'db_path', None):
                    import sqlite3
                    conn = connect_sqlite(self.db_path)
                    tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                    knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                    if knvv_name:
                        knvv = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                    conn.close()
            if knvv is None or knvv.empty:
                print(f"      [WARN] Таблица KNVV не найдена или пуста для {rule_code}")
                return df
            kunnr_knvv = next((c for c in knvv.columns if str(c).upper() in ('KUNNR', 'KUNNR_KNVV')), None)
            aufsd_knvv = next((c for c in knvv.columns if 'AUFSD' in str(c).upper()), None)
            if not kunnr_knvv or not aufsd_knvv:
                print(f"      [WARN] В KNVV не найдены KUNNR или AUFSD для {rule_code}")
                return df
            knvv_fm = knvv[knvv[aufsd_knvv].astype(str).str.strip().str.upper().isin(('F', 'M'))][[kunnr_knvv, aufsd_knvv]].drop_duplicates(subset=[kunnr_knvv], keep='first')
            knvv_fm = knvv_fm.rename(columns={kunnr_knvv: '_partner', aufsd_knvv: 'AUFSD'})
            df['_partner'] = _norm_key(df[partner_col])
            knvv_fm['_partner'] = _norm_key(knvv_fm['_partner'])
            before = len(df)
            df = df.merge(knvv_fm, on='_partner', how='inner')
            df = df.drop(columns=['_partner'], errors='ignore')
            print(f"      [FILTER] Фильтр ADR2 по KNVV AUFSD in (F,M): {len(df):,} из {before:,} строк для {rule_code}")
            if df.empty:
                print(f"      [WARN] После фильтра по блокам F/M данных нет для {rule_code}")
            return df
        except Exception as e:
            print(f"      [WARN] Ошибка _filter_adr2_by_knvv_aufsd_fm для {rule_code}: {e}")
            import traceback
            traceback.print_exc()
            return df

    def _filter_adr2_rccomp_375_1_scope_by_kna1_aufsd(self, df, rule_code, table_name=None):
        """
        Scope для RCCOMP_375.1 (как telbd export):
          - ADR2.ADDRNUMBER -> BUT020.ADDRNUMBER -> PARTNER
          - PARTNER -> KNA1.KUNNR -> AUFSD (central_order_block_code)
          - исключаем AUFSD IN ('E','G','SP','R','U')

        ВАЖНО: этот scope НЕ использует KNVV AUFSD F/M.
        """
        if table_name and str(table_name).strip().upper() != "ADR2":
            return df

        try:
            def _norm_key(series: pd.Series) -> pd.Series:
                s = series.astype(str).str.strip()
                s = s.str.replace(r"\.0$", "", regex=True)
                s = s.str.replace(r"\D+", "", regex=True)
                return s.str.zfill(10)

            addr_col = next((c for c in df.columns if "ADDRNUMBER" in str(c).upper()), None)
            if not addr_col:
                print(f"      [FILTER] Колонка ADDRNUMBER не найдена в ADR2 для {rule_code}")
                return df

            but020 = self.memory_manager.get_table("BUT020")
            if but020 is None or but020.empty:
                print(f"      [WARN] Таблица BUT020 не найдена или пуста для {rule_code}")
                return df

            addr_but = next((c for c in but020.columns if "ADDRNUMBER" in str(c).upper()), None)
            partner_but = next((c for c in but020.columns if str(c).upper() == "PARTNER"), None)
            if not addr_but or not partner_but:
                print(f"      [WARN] В BUT020 не найдены ADDRNUMBER или PARTNER для {rule_code}")
                return df

            df = df.copy()
            df["_ak"] = _norm_key(df[addr_col])
            but_join = but020[[addr_but, partner_but]].copy()
            but_join["_ak"] = _norm_key(but_join[addr_but])
            but_join = but_join.drop_duplicates(subset=["_ak"], keep="first")
            df = (
                df.merge(but_join[["_ak", partner_but]], on="_ak", how="left")
                .drop(columns=["_ak"], errors="ignore")
            )
            if partner_but not in df.columns:
                print(f"      [WARN] PARTNER не добавлен из BUT020 для {rule_code}")
                return df

            kna1 = self._get_table_for_rules("KNA1")
            if kna1 is None or kna1.empty:
                print(f"      [WARN] Таблица KNA1 не найдена или пуста для {rule_code}")
                return df

            kunnr_col = next((c for c in kna1.columns if str(c).upper() == "KUNNR"), None)
            aufsd_col = next((c for c in kna1.columns if str(c).upper() == "AUFSD"), None)
            if not kunnr_col or not aufsd_col:
                print(f"      [WARN] В KNA1 не найдены KUNNR или AUFSD для {rule_code}")
                return df

            # JOIN ключ партнёра
            df["_partner_key"] = _norm_key(df[partner_but])
            kna1_join = kna1[[kunnr_col, aufsd_col]].copy()
            kna1_join["_partner_key"] = _norm_key(kna1_join[kunnr_col])
            kna1_join = kna1_join.drop_duplicates(subset=["_partner_key"], keep="first")

            df = df.merge(
                kna1_join[["_partner_key", aufsd_col]].rename(columns={aufsd_col: "AUFSD"}),
                on="_partner_key",
                how="left",
            ).drop(columns=["_partner_key"], errors="ignore")

            blocked_codes = {"E", "G", "SP", "R", "U"}
            aufsd_raw = df["AUFSD"]
            aufsd_raw = aufsd_raw.where(aufsd_raw.notna(), "")
            aufsd_norm = aufsd_raw.astype(str).str.strip().str.upper()
            is_blocked = aufsd_norm.isin(blocked_codes)
            return df.loc[~is_blocked].copy()

        except Exception as e:
            print(f"      [WARN] Ошибка _filter_adr2_rccomp_375_1_scope_by_kna1_aufsd для {rule_code}: {e}")
            return df
    
    def _non_empty_key_count(self, series) -> int:
        """Число непустых значений ключа клиента (для выбора Customer vs пустой KUNNR)."""
        if series is None:
            return 0
        filled, _distinct = self._kunnr_join_key_stats(series)
        return filled

    def _kunnr_join_key_stats(self, series) -> tuple:
        """
        (число нормализованных ключей, число уникальных ключей) для выбора колонки JOIN.
        KNA1/KNB1: Customer и KUNNR — одно поле (номер клиента); Cl_ не используется.
        """
        if series is None:
            return 0, 0
        try:
            norms = series.apply(self._norm_customer_partner_key)
        except Exception:
            return 0, 0
        mask = norms.ne("")
        filled = int(mask.sum())
        distinct = int(norms[mask].nunique()) if filled else 0
        return filled, distinct

    def _pick_best_kunnr_column(self, df, table_name: str = "KNB1"):
        """
        Номер клиента (SAP KUNNR): в KNB1 выгрузка — Customer, в KNA1 — Customer или KUNNR
        (одно и то же поле). Cl_ в KNA1 — не KUNNR, в JOIN не участвует.
        """
        if df is None or df.empty:
            return None
        tn = str(table_name or "").strip().upper()
        col_upper = {str(c).strip().upper(): c for c in df.columns}
        # Явный приоритет под вашу выгрузку:
        # KNB1 -> Customer, KNA1 -> KUNNR (fallback на Customer)
        if tn == "KNB1":
            name_order = ("CUSTOMER", "KUNNR", "CUSTOMER_CODE", "KUNNR_KNB1", "CLIENT")
        elif tn == "KNA1":
            name_order = ("KUNNR", "CUSTOMER", "CUSTOMER_CODE", "KUNNR_KNB1", "CLIENT")
        else:
            name_order = ("CUSTOMER", "KUNNR", "CUSTOMER_CODE", "PARTNER", "CLIENT")
        priority = {name: i for i, name in enumerate(name_order)}
        candidates = []
        for name in name_order:
            if name in col_upper:
                candidates.append(col_upper[name])
        if not candidates:
            try:
                from utils.column_map_resolver import resolve_column_in_df
                for sap in ("KUNNR", "Customer"):
                    c = resolve_column_in_df(df, sap, table_name, self.column_map, parent_dir)
                    if c and c not in candidates:
                        candidates.append(c)
            except ImportError:
                pass
        if not candidates:
            return self._find_kunnr_column(df)

        def _score(col):
            filled, distinct = self._kunnr_join_key_stats(df[col])
            prio = priority.get(str(col).strip().upper(), len(name_order))
            return (filled, distinct, -prio)

        best = max(candidates, key=_score)
        filled_b, distinct_b = self._kunnr_join_key_stats(df[best])
        if filled_b == 0:
            print(f"      [WARN] {table_name}: все кандидаты ключа клиента пусты ({candidates})")
        elif len(candidates) > 1:
            alt = [c for c in candidates if c != best]
            if alt:
                c0 = alt[0]
                f0, d0 = self._kunnr_join_key_stats(df[c0])
                if f0 == filled_b and d0 != distinct_b and d0 < max(100, distinct_b // 100):
                    print(
                        f"      [JOIN] {table_name}: ключ KUNNR/Customer [{best}] "
                        f"(заполнено {filled_b:,}, уникальных {distinct_b:,}); "
                        f"отклонён [{c0}] (уникальных {d0:,})"
                    )
        return best

    def _resolve_knb1_kna1_join_column(self, df, table_name: str):
        """
        JOIN KNB1 ↔ KNA1: KNB1.Customer = KNA1.KUNNR (fallback KNA1.Customer).
        """
        return self._pick_best_kunnr_column(df, table_name)

    def _add_account_group_code_from_kna1(self, df, table_name, rule_code):
        """
        Добавляет account_group_code (KTOKD / kna.KTOKD / b.ktokd) из KNA1 через JOIN по KUNNR.

        KNB1: KNB1.Customer = KNA1.KUNNR (= KNA1.Customer в выгрузке) -> KNA1.Group_1 (KTOKD).
        KNVV и др.: аналогично по customer_code / KUNNR.
        """
        try:
            rule_code_u = str(rule_code).strip().upper()
            table_u = str(table_name or "").strip().upper()
            # KNB1: KTOKD только из KNA1 (не KNB1.AuGr и не пустой локальный KTOKD).
            force_rebuild_for_rule = table_u == "KNB1" or rule_code_u in getattr(
                self, "RULES_FORCE_KNA1_KTOKD_JOIN", ("RCCONF_24.1", "RCCONF_115.11")
            )

            if str(rule_code).strip().upper() == "RCCONF_113.1":
                print("      [DEBUG] RCCONF_113.1 JOIN PATH v2 (memory->sqlite fallback)")

            def _norm_join_key(v):
                return self._norm_customer_partner_key(v)

            # Если account_group_code уже есть — обычно не пересчитываем.
            # Но для RCCONF_24.1 принудительно пересобираем (может быть, что колонка
            # присутствует, но нам нужна именно KNA1.KTOKD под фильтр 7038/9038).
            if not force_rebuild_for_rule and ('account_group_code' in df.columns or 'KTOKD' in df.columns):
                print(f"      [JOIN] account_group_code уже присутствует в таблице {table_name}")
                return df

            if force_rebuild_for_rule:
                df = self._drop_kna1_account_group_columns(df)
            
            print(f"      [JOIN] Добавление account_group_code из KNA1 для правила {rule_code} в таблице {table_name}...")
            
            # Ищем колонку для JOIN (customer_code/KUNNR) в текущей таблице.
            # KNB1: обязательно Customer (SAP KUNNR) <-> KNA1.KUNNR / KNA1.Customer.
            # ADRC: ADDRNUMBER -> BUT020 -> PARTNER -> KNA1.KUNNR.
            join_col = None
            is_rule_24_1 = rule_code_u == "RCCONF_24.1"
            if table_u == "KNB1":
                join_col = self._resolve_knb1_kna1_join_column(df, table_name)
                if join_col:
                    print(
                        f"      [JOIN] KNB1: ключ {table_name}.{join_col} (Customer/KUNNR) "
                        f"-> KNA1.Customer/KUNNR -> KNA1.KTOKD (Group_1) как account_group_code"
                    )
            if is_rule_24_1:
                partner_direct = next((c for c in df.columns if str(c).strip().upper() == "PARTNER"), None)
                if partner_direct:
                    join_col = partner_direct
                    print(f"      [JOIN] RCCONF_24.1: принудительный JOIN по PARTNER -> KNA1.KUNNR (колонка: {join_col})")
            if not join_col:
                join_col = self._resolve_knb1_kna1_join_column(df, table_name)
                if join_col:
                    print(f"      [JOIN] Найдена колонка для JOIN в {table_name}: {join_col}")
            if not join_col:
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower in ['kunnr', 'customer_code', 'customer', 'kunnr_knvv', 'kunnr_knb1']:
                        join_col = col
                        print(f"      [JOIN] Найдена колонка для JOIN в {table_name}: {col}")
                        break
            
            # ADRC fallback: сначала получаем PARTNER из BUT020 по адресу, затем JOIN к KNA1 по KUNNR
            join_mode = "kunnr"
            if not join_col and str(table_name or "").strip().upper() == "ADRC":
                addr_col = None
                for col in df.columns:
                    cu = str(col).strip().upper()
                    if cu in ("ADDRNUMBER", "ADRNR", "ADDRNR_ADRC") or ("ADDR" in cu and ("NUMBER" in cu or cu.endswith("NR"))):
                        addr_col = col
                        break
                if addr_col:
                    but020_df = self.memory_manager.get_table("BUT020")
                    if but020_df is None or but020_df.empty:
                        try:
                            self.memory_manager.load_selected_tables_to_ram(["BUT020"], add_reference_tables=False)
                            but020_df = self.memory_manager.get_table("BUT020")
                        except Exception:
                            but020_df = None
                    if but020_df is None or but020_df.empty:
                        try:
                            conn = connect_sqlite(self.db_path)
                            but020_df = pd.read_sql_query('SELECT * FROM "BUT020"', conn)
                            conn.close()
                        except Exception:
                            but020_df = None

                    if but020_df is not None and not but020_df.empty:
                        addr_but = next((c for c in but020_df.columns if "ADDRNUMBER" in str(c).upper() or str(c).upper() == "ADRNR"), None)
                        partner_but = next((c for c in but020_df.columns if str(c).upper() == "PARTNER"), None)
                        if addr_but and partner_but:
                            df = df.copy()
                            but_join = but020_df[[addr_but, partner_but]].copy()
                            # В Python 3.13 / pandas возможен Buffer dtype mismatch при merge int/int64.
                            # Делаем join по нормализованному строковому ключу адреса.
                            def _norm_addr_key(series: pd.Series) -> pd.Series:
                                s = series.astype(str).str.strip()
                                s = s.str.replace(r"\.0$", "", regex=True)
                                s = s.str.replace(r"\D+", "", regex=True)
                                return s.str.zfill(10)
                            df["_addr_key_norm"] = _norm_addr_key(df[addr_col])
                            but_join["_addr_key_norm"] = _norm_addr_key(but_join[addr_but])
                            but_join = but_join.drop_duplicates(subset=["_addr_key_norm"], keep='first')
                            df = df.merge(but_join[["_addr_key_norm", partner_but]], on="_addr_key_norm", how='left')
                            df = df.drop(columns=["_addr_key_norm"], errors='ignore')
                            if addr_but in df.columns and addr_but != addr_col:
                                df = df.drop(columns=[addr_but], errors='ignore')
                            if partner_but in df.columns:
                                if partner_but != "PARTNER":
                                    df = df.rename(columns={partner_but: "PARTNER"})
                                join_col = "PARTNER"
                                join_mode = "kunnr"
                                if is_rule_24_1:
                                    print(f"      [JOIN] RCCONF_24.1: путь JOIN по PARTNER (ADRC.{addr_col} -> BUT020.{addr_but} -> KNA1.KUNNR)")
                                else:
                                    print(f"      [JOIN] ADRC: путь JOIN по PARTNER (ADRC.{addr_col} -> BUT020.{addr_but} -> KNA1.KUNNR)")

            if not join_col:
                print(f"      [WARN] Колонка для JOIN (KUNNR/customer_code или ADRC address) не найдена в {table_name} для правила {rule_code}")
                return df

            # Подгрузка KNA1 в RAM при проверке только KNB1
            if self._get_table_for_rules("KNA1") is None or self._get_table_for_rules("KNA1").empty:
                print(f"      [INFO] KNA1 отсутствует в RAM для {rule_code} — загружаем...")
                try:
                    self.memory_manager.load_selected_tables_to_ram(["KNA1"], add_reference_tables=False)
                    setattr(self, "_kna1_ktokd_lookup_df", None)
                except Exception as e:
                    print(f"      [WARN] Не удалось загрузить KNA1: {e}")

            return self._merge_kna1_account_group_from_lookup(df, table_name, rule_code, join_col)
            
        except Exception as e:
            print(f"      [WARN] Ошибка при добавлении account_group_code из KNA1 для правила {rule_code}: {e}")
            import traceback
            traceback.print_exc()
            return df

    def _find_kna1_ktokd_column_in_df(self, df):
        """Колонка группы счёта из KNA1 в DataFrame (после JOIN или выгрузка Group_1)."""
        if df is None or df.empty:
            return None
        col_lower = {str(c).strip().lower(): c for c in df.columns}
        for name in (
            "ktokd",
            "kna.ktokd",
            "account_group_code",
            "b.account_group_code",
            "b.ktokd",
            "group_1",
            "lookup_account_group_ktokd",
        ):
            if name in col_lower:
                return col_lower[name]
        return None

    def _place_column_after(self, df, col_name, after_col_names):
        """Переставляет col_name сразу после первой найденной колонки из after_col_names."""
        if df is None or col_name not in df.columns:
            return df
        after_col = None
        upper_map = {str(c).strip().upper(): c for c in df.columns}
        for name in after_col_names:
            if name.upper() in upper_map:
                after_col = upper_map[name.upper()]
                break
        if not after_col:
            return df
        cols = [c for c in df.columns if c != col_name]
        if after_col not in cols:
            return df
        ix = cols.index(after_col) + 1
        return df[cols[:ix] + [col_name] + cols[ix:]]

    def _normalize_rule_code(self, rule_code: str) -> str:
        return re.sub(r"[^A-Za-z0-9._-]", "", str(rule_code or "")).strip().upper()

    def _format_ktokd_for_export(self, series):
        """KTOKD / Group_1 из KNA1 для выгрузки ошибок (без .0, пусто = '')."""
        if series is None:
            return pd.Series(dtype=object)
        s = series.astype(str).str.strip().str.replace(r"\.0+$", "", regex=True)
        return s.replace({"nan": "", "None": "", "null": "", "NaN": ""})

    def _find_customer_column_for_kna1_join(self, df, table_name: str = "KNB1"):
        """Колонка клиента в KNB1/ошибках для JOIN к KNA1 (Customer, не пустая SAP-копия KUNNR)."""
        return self._pick_best_kunnr_column(df, table_name or "KNB1")

    def _build_kna1_ktokd_lookup(self, force_reload: bool = False):
        """Справочник KNA1: _join_key (норм. KUNNR) -> KTOKD (Group_1). Кэш на прогон."""
        cache_key = "_kna1_ktokd_lookup_df"
        if not force_reload and getattr(self, cache_key, None) is not None:
            return getattr(self, cache_key)

        kna1_df = None
        if hasattr(self, "memory_manager"):
            kna1_df = self.memory_manager.get_table("KNA1")
        if kna1_df is None or kna1_df.empty:
            if getattr(self, "db_path", None):
                try:
                    conn = connect_sqlite(self.db_path)
                    try:
                        kna1_df = pd.read_sql_query(
                            'SELECT "Customer", "Group_1" FROM "KNA1"',
                            conn,
                        )
                    except Exception:
                        kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                except Exception:
                    kna1_df = None

        lookup = pd.DataFrame(columns=["_join_key", "KTOKD"])
        if kna1_df is None or kna1_df.empty:
            setattr(self, cache_key, lookup)
            return lookup

        kna1_mapped = self._apply_rule_time_column_map(kna1_df.copy(), "KNA1")
        kunnr_col = self._pick_best_kunnr_column(kna1_mapped, "KNA1")
        ktokd_col = None
        for c in kna1_mapped.columns:
            cl = str(c).strip().lower()
            if cl in ("group_1", "ktokd", "account_group_code"):
                if ktokd_col is None or self._non_empty_key_count(kna1_mapped[c]) > self._non_empty_key_count(
                    kna1_mapped[ktokd_col]
                ):
                    ktokd_col = c
        kna1_df = kna1_mapped

        if kunnr_col and ktokd_col:
            lookup = kna1_df[[kunnr_col, ktokd_col]].copy()
            lookup["_join_key"] = lookup[kunnr_col].apply(self._norm_customer_partner_key)
            lookup["KTOKD"] = self._format_ktokd_for_export(lookup[ktokd_col])
            lookup = lookup.drop_duplicates(subset=["_join_key"], keep="first")[["_join_key", "KTOKD"]]

        if kunnr_col and ktokd_col:
            _kf, _kd = self._kunnr_join_key_stats(kna1_df[kunnr_col])
            print(
                f"      [JOIN] справочник KNA1: ключ [{kunnr_col}] "
                f"({_kf:,} ключей, уникальных {_kd:,}), KTOKD из [{ktokd_col}]"
            )
        setattr(self, cache_key, lookup)
        return lookup

    def _drop_kna1_account_group_columns(self, df):
        if df is None or df.empty:
            return df
        drop = []
        for c in df.columns:
            cl = str(c).strip().lower()
            cu = str(c).strip().upper()
            if cl in (
                "account_group_code",
                "b.account_group_code",
                "ktokd",
                "b.ktokd",
                "kna.ktokd",
                "group_1",
            ) or cu == "KTOKD":
                drop.append(c)
        if drop:
            return df.drop(columns=drop, errors="ignore")
        return df

    def _merge_kna1_account_group_from_lookup(self, df, table_name, rule_code, join_col):
        """JOIN KNA1.Group_1 -> account_group_code / KTOKD по нормализованному Customer/KUNNR."""
        if df is None or df.empty or not join_col:
            return df
        lookup = self._build_kna1_ktokd_lookup(force_reload=False)
        if lookup is None or lookup.empty:
            print(f"      [WARN] {rule_code}: справочник KNA1 (Customer->Group_1) пуст")
            return df

        out = self._drop_kna1_account_group_columns(df.copy())
        out["_join_key"] = out[join_col].apply(self._norm_customer_partner_key)
        matched_keys = out["_join_key"].isin(lookup["_join_key"])
        out = out.merge(lookup, on="_join_key", how="left")
        out = out.drop(columns=["_join_key"], errors="ignore")

        if "KTOKD" in out.columns:
            out["account_group_code"] = out["KTOKD"]
            out["b.account_group_code"] = out["KTOKD"]
            out["ktokd"] = out["KTOKD"]
            out["b.ktokd"] = out["KTOKD"]
            out["kna.ktokd"] = out["KTOKD"]

        from utils.sap_account_keys import norm_sap_account_group

        ktokd_norm = out["KTOKD"].apply(norm_sap_account_group) if "KTOKD" in out.columns else pd.Series(dtype=str)
        filled = int((ktokd_norm != "").sum())
        n9038 = int((ktokd_norm == "9038").sum())
        self._last_kna1_join_stats = {
            "rows_after_join": len(out),
            "filled_ktokd": filled,
            "n9038": n9038,
            "join_col": join_col,
            "key_matched": int(matched_keys.sum()),
        }
        print(
            f"      [JOIN] {rule_code}: KNA1.Group_1 -> {table_name} по [{join_col}]: "
            f"ключ совпал {int(matched_keys.sum()):,}/{len(out):,}, "
            f"KTOKD из KNA1: {filled:,}, из них 9038: {n9038:,}"
        )
        if int(matched_keys.sum()) == 0 and filled > 0:
            print(
                f"      [WARN] {rule_code}: KTOKD не пустой, но ключей KNA1=0 — "
                f"проверьте колонку JOIN (нужна Customer, не пустая KUNNR)"
            )
        return out

    def _attach_kna1_ktokd_export_columns(self, df, rule_code=None):
        """
        Колонки KTOKD / KTOKD_SOURCE в срезе до валидации — попадут в error_df как есть.
        """
        if df is None or df.empty:
            return df
        ag_col = self._find_account_group_column(df)
        out = df.copy()
        if ag_col is not None:
            out["KTOKD"] = self._format_ktokd_for_export(out[ag_col])
        else:
            customer_col = self._find_customer_column_for_kna1_join(out, "KNB1")
            if customer_col:
                lookup = self._build_kna1_ktokd_lookup()
                if not lookup.empty:
                    out["_jk"] = out[customer_col].apply(self._norm_customer_partner_key)
                    out = out.merge(lookup, left_on="_jk", right_on="_join_key", how="left")
                    out = out.drop(columns=["_jk", "_join_key"], errors="ignore")
        if "KTOKD" not in out.columns:
            out["KTOKD"] = ""
        out["KTOKD_SOURCE"] = "KNA1"
        rule_u = self._normalize_rule_code(rule_code)
        if rule_u in self.RULES_KTOKD_ONLY_9038_SCOPE:
            out["RULE_SCOPE"] = "only KNA1.KTOKD=9038"
        return self._place_column_after(out, "KTOKD", ("Customer", "KUNNR", "CUSTOMER", "Cl_", "CLIENT"))

    def _enrich_error_df_kna1_ktokd(self, error_df, table_name, rule_code=None):
        """
        В файл ошибок добавляет колонку KTOKD из KNA1 (Group_1) по Customer/KUNNR.
        Всегда перечитывает KNA1 — не полагается на JOIN на этапе проверки.
        """
        if error_df is None or error_df.empty:
            return error_df
        rule_u = self._normalize_rule_code(rule_code)

        # Уже есть KTOKD из этапа валидации — только выровнять порядок колонок
        if "KTOKD" in error_df.columns:
            out = error_df.copy()
            if "KTOKD_SOURCE" not in out.columns:
                out["KTOKD_SOURCE"] = "KNA1"
            if rule_u in self.RULES_KTOKD_ONLY_9038_SCOPE and "RULE_SCOPE" not in out.columns:
                out["RULE_SCOPE"] = "only KNA1.KTOKD=9038"
            return self._place_column_after(out, "KTOKD", ("Customer", "KUNNR", "CUSTOMER", "Cl_", "CLIENT"))

        customer_col = self._find_customer_column_for_kna1_join(error_df, table_name or "KNB1")
        if not customer_col:
            print(
                f"      [WARN] {rule_code}: в error_df нет Customer/KUNNR — "
                f"колонки: {list(error_df.columns)[:12]}..."
            )
            out = error_df.copy()
            out["KTOKD"] = ""
            out["KTOKD_SOURCE"] = "KNA1 (join key not found)"
            return out

        try:
            lookup = self._build_kna1_ktokd_lookup()
            if lookup.empty:
                print(f"      [WARN] {rule_code}: справочник KNA1.KTOKD пуст — KTOKD не добавлен")
                out = error_df.copy()
                out["KTOKD"] = ""
                out["KTOKD_SOURCE"] = "KNA1"
                return out

            out = error_df.copy()
            out["_join_key"] = out[customer_col].apply(self._norm_customer_partner_key)
            out = out.merge(lookup, on="_join_key", how="left")
            out = out.drop(columns=["_join_key"], errors="ignore")
            if "KTOKD" not in out.columns:
                out["KTOKD"] = ""
            out["KTOKD_SOURCE"] = "KNA1"

            filled = int((out["KTOKD"].astype(str).str.strip() != "").sum())
            print(
                f"      [JOIN] {rule_code}: KTOKD из KNA1 (Group_1) в error_df — "
                f"заполнено {filled:,} из {len(out):,} (ключ: {customer_col})"
            )

            if rule_u in self.RULES_KTOKD_ONLY_9038_SCOPE:
                out["RULE_SCOPE"] = "only KNA1.KTOKD=9038"
                ktokd_chk = out["KTOKD"].astype(str).str.strip()
                not_9038 = ~ktokd_chk.isin({"9038", ""}) & ktokd_chk.notna()
                if not_9038.any():
                    print(
                        f"      [WARN] {rule_code}: {int(not_9038.sum()):,} строк в error_df с KTOKD != 9038"
                    )

            out = self._place_column_after(out, "KTOKD", ("Customer", "KUNNR", "CUSTOMER", "Cl_", "CLIENT"))
            if "KTOKD_SOURCE" in out.columns:
                out = self._place_column_after(out, "KTOKD_SOURCE", ("KTOKD",))
            if "RULE_SCOPE" in out.columns:
                out = self._place_column_after(out, "RULE_SCOPE", ("KTOKD", "KTOKD_SOURCE"))
            return out
        except Exception as e:
            print(f"      [WARN] {rule_code}: не удалось подтянуть KTOKD из KNA1 в error_df: {e}")
            traceback.print_exc()
            out = error_df.copy()
            out["KTOKD"] = ""
            out["KTOKD_SOURCE"] = "KNA1 (error)"
            return out

    def _add_central_order_block_code_from_kna1(self, df, table_name, rule_code):
        """
        Добавляет central_order_block_code из KNA1.AUFSD для BUT051 (RCCONF_371.2).
        JOIN: BUT051.PARTNER2 -> KNA1.KUNNR.
        """
        try:
            if df is None or df.empty:
                return df
            if "central_order_block_code" in df.columns:
                return df

            def _norm_partner_key(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return ""
                s = str(v).replace("\ufeff", "").replace("\u00a0", "").strip()
                if re.fullmatch(r"\d+\.0+", s):
                    s = s.split(".")[0]
                digits = re.sub(r"\D", "", s)
                if digits:
                    s = digits
                return s.lstrip("0") or "0"

            # Для RCCONF_371.2 основной ключ — PARTNER1 (контрагент), fallback — PARTNER2.
            partner_col = next((c for c in df.columns if str(c).strip().upper() == "PARTNER1"), None)
            if not partner_col:
                partner_col = next((c for c in df.columns if "PARTNER1" in str(c).strip().upper()), None)
            if not partner_col:
                partner_col = next((c for c in df.columns if str(c).strip().upper() == "PARTNER2"), None)
            if not partner_col:
                partner_col = next((c for c in df.columns if "PARTNER2" in str(c).strip().upper()), None)
            if not partner_col:
                print(f"      [WARN] {rule_code}: в {table_name} не найдены колонки PARTNER1/PARTNER2 для JOIN с KNA1")
                return df

            kna1_df = self._get_table_for_rules("KNA1")
            if kna1_df is None or kna1_df.empty:
                try:
                    self.memory_manager.load_selected_tables_to_ram(["KNA1"], add_reference_tables=False)
                    kna1_df = self._get_table_for_rules("KNA1")
                except Exception:
                    kna1_df = None

            if kna1_df is None or kna1_df.empty:
                try:
                    conn = connect_sqlite(self.db_path)
                    try:
                        kna1_df = pd.read_sql_query('SELECT "KUNNR","AUFSD" FROM "KNA1"', conn)
                    except Exception:
                        kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                except Exception as e:
                    print(f"      [WARN] {rule_code}: не удалось загрузить KNA1 для central_order_block_code: {e}")
                    return df

            if kna1_df is None or kna1_df.empty:
                print(f"      [WARN] {rule_code}: KNA1 пуста, central_order_block_code не добавлен")
                return df

            kna1_df = self._apply_rule_time_column_map(kna1_df, "KNA1")
            try:
                from utils.column_map_resolver import resolve_column_in_df
                kna1_kunnr_col = resolve_column_in_df(
                    kna1_df, "KUNNR", "KNA1", self.column_map, parent_dir
                )
                aufsd_col = resolve_column_in_df(
                    kna1_df, "AUFSD", "KNA1", self.column_map, parent_dir
                ) or resolve_column_in_df(
                    kna1_df, "central_order_block_code", "KNA1", self.column_map, parent_dir
                )
            except ImportError:
                kna1_kunnr_col = None
                aufsd_col = None
            if not kna1_kunnr_col:
                kna1_kunnr_col = next((c for c in kna1_df.columns if str(c).strip().upper() == "KUNNR"), None)
            if not aufsd_col:
                aufsd_col = next((c for c in kna1_df.columns if str(c).strip().upper() == "AUFSD"), None)
            if not kna1_kunnr_col or not aufsd_col:
                print(f"      [WARN] {rule_code}: в KNA1 не найдены KUNNR/AUFSD для JOIN")
                return df

            out = df.copy()
            out["_partner2_key"] = out[partner_col].apply(_norm_partner_key)

            kna1_join = kna1_df[[kna1_kunnr_col, aufsd_col]].copy()
            kna1_join["_partner2_key"] = kna1_join[kna1_kunnr_col].apply(_norm_partner_key)
            kna1_join = (
                kna1_join[["_partner2_key", aufsd_col]]
                .drop_duplicates(subset=["_partner2_key"], keep="first")
                .rename(columns={aufsd_col: "central_order_block_code"})
            )

            out = out.merge(kna1_join, on="_partner2_key", how="left")
            out = out.drop(columns=["_partner2_key"], errors="ignore")
            print(
                f"      [JOIN] {rule_code}: добавлен central_order_block_code из KNA1.AUFSD "
                f"по {table_name}.{partner_col} -> KNA1.{kna1_kunnr_col}"
            )
            return out
        except Exception as e:
            print(f"      [WARN] Ошибка добавления central_order_block_code из KNA1 для {rule_code}: {e}")
            return df
    
    def _apply_conditional_filter(self, df, technical_def, rule_code, table_name=None):
        """Применяет фильтрацию данных на основе условий"""
        try:
            print(f"      [FILTER] Анализ условий для {rule_code}...")
            
            # Для правил RCCONF_39.5.2 и RCCONF_39.3.2 фильтруем по PERSNUMBER (должно быть ЗАПОЛНЕНО), затем по KNVV AUFSD F/M для ADR2
            if rule_code in ["RCCONF_39.5.2", "RCCONF_39.3.2"]:
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                        break
                if person_col:
                    mask = (
                        df[person_col].notna() &
                        (df[person_col].astype(str).str.strip() != '') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'none') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'null')
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL (заполнено) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации (PERSNUMBER заполнено) данных нет!")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                else:
                    print(f"      [WARN] Колонка PERSNUMBER не найдена для правила {rule_code}")
            
            # Для правила RCCONF_39.5 (ADR2): только пустой PERSNUMBER, затем фильтр по KNVV AUFSD F/M
            if rule_code == "RCCONF_39.5" and table_name and str(table_name).strip().upper() == "ADR2":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк")
                if not df.empty:
                    return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)
                return df

            # RCCONF_39.3: только PERSNUMBER IS NULL (без R3_USER / KNVV — см. rules.json)
            if rule_code == "RCCONF_39.3":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower in ("persnumber", "pers_number") or col_lower == "person" or "persnumber" in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                        break
                if person_col:
                    mask = (
                        df[person_col].isna()
                        | (df[person_col].astype(str).str.strip() == "")
                        | (df[person_col].astype(str).str.strip().str.lower().isin(("none", "null", "nan")))
                    )
                    filtered_df = df[mask].copy()
                    print(
                        f"      [FILTER] RCCONF_39.3: PERSNUMBER IS NULL — "
                        f"{len(filtered_df)} из {len(df)} строк"
                    )
                    return filtered_df
                print(f"      [WARN] RCCONF_39.3: колонка PERSNUMBER не найдена")
                return df
            
            # Для правила RCCONF_38.3: R3_USER = '1' (стационарный телефон) AND PERSNUMBER IS NULL
            # contact_medium_type = 'fixed_tel_number' когда R3_USER = '1'
            if rule_code == "RCCONF_38.3":
                r3_user_col = None
                person_col = None
                contact_medium_col = None
                
                # Ищем колонку R3_USER (определяет тип телефона: '1' = стационарный, '3' = мобильный)
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or 'r3user' in col_lower:
                        r3_user_col = col
                        print(f"      [FILTER] Найдена колонка R3_USER: {col}")
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                    if 'contact_medium' in col_lower or 'medium_type' in col_lower or 'contactmedium' in col_lower:
                        contact_medium_col = col
                        print(f"      [FILTER] Найдена колонка contact_medium_type: {col}")
                
                # Приоритет: используем R3_USER = '1' для определения стационарного телефона
                if r3_user_col and person_col:
                    mask = (
                        (df[r3_user_col].astype(str).str.strip() == '1') &  # R3_USER = '1' означает стационарный телефон
                        (df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none'))
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (R3_USER='1' AND PERSNUMBER IS NULL) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации данных нет!")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif r3_user_col:
                    # Если нет PERSNUMBER, фильтруем только по R3_USER = '1'
                    mask = (df[r3_user_col].astype(str).str.strip() == '1')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только R3_USER='1') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif contact_medium_col and person_col:
                    # Fallback: используем contact_medium_type если R3_USER не найден
                    mask = (
                        (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number') &
                        (df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none'))
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (contact_medium_type='fixed_tel_number' AND PERSNUMBER IS NULL) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации данных нет!")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif contact_medium_col:
                    # Если нет PERSNUMBER, фильтруем только по contact_medium_type
                    mask = (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только contact_medium_type='fixed_tel_number') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                elif person_col:
                    mask = (
                        df[person_col].isna()
                        | (df[person_col].astype(str).str.strip() == "")
                        | (df[person_col].astype(str).str.strip().str.lower().isin(("none", "null", "nan")))
                    )
                    filtered_df = df[mask].copy()
                    print(
                        f"      [FILTER] RCCONF_38.3: PERSNUMBER IS NULL (fallback без R3_USER) — "
                        f"{len(filtered_df)} из {len(df)} строк"
                    )
                    return self._filter_adr2_by_knvv_aufsd_fm(filtered_df, rule_code, table_name) if (table_name and str(table_name).strip().upper() == 'ADR2') else filtered_df
                else:
                    print(f"      [WARN] RCCONF_38.3: колонки R3_USER / PERSNUMBER / contact_medium_type не найдены")
            
            # Для правила RCCONF_38.5 фильтруем по PERSNUMBER (должно быть пустым)
            elif rule_code == "RCCONF_38.5":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f"      [FILTER] Найдена колонка PERSNUMBER: {col}")
                        break
                
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f"      [WARN] После фильтрации (PERSNUMBER IS NULL) данных нет!")
                    return filtered_df
                else:
                    print(f"      [WARN] Колонка PERSNUMBER не найдена для правила {rule_code}")
            
            # Для правил BUT050: RCCOMP_369.1 и RCCONF_369.1
            if rule_code in ["RCCOMP_369.1", "RCCONF_369.1"]:
                source_col = None
                
                # Ищем колонку source (может быть source_file или source)
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'source' or col_lower == 'source_file' or 'source' in col_lower:
                        source_col = col
                        print(f"      [FILTER] Найдена колонка source: {col}")
                        break
                
                # Для RCCOMP_369.1: фильтруем по source = 's4'
                if rule_code == "RCCOMP_369.1":
                    if source_col:
                        mask = df[source_col].astype(str).str.strip().str.lower() == 's4'
                        filtered_df = df[mask].copy()
                        print(f"      [FILTER] Применен фильтр (source='s4') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                        if len(filtered_df) == 0:
                            print(f"      [WARN] После фильтрации (source='s4') данных нет!")
                        return filtered_df
                    else:
                        print(f"      [WARN] Колонка source не найдена для правила {rule_code}")
                        # Если source не найден, работаем со всеми данными
                        return df
                
                # Для RCCONF_369.1: фильтрация не требуется по source, но может быть по другим условиям
                # Пока возвращаем все данные, фильтрация будет в валидаторе
                elif rule_code == "RCCONF_369.1":
                    print(f"      [FILTER] Правило {rule_code} не требует фильтрации по source, работаем со всеми данными")
                    return df
            
            # Для правила RCCOMP_375.1 (ADR2): только пустой PERSNUMBER, затем фильтр по KNVV AUFSD F/M (BUT020->PARTNER->KNVV)
            elif rule_code == "RCCOMP_375.1":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк")
                    if len(df) == 0:
                        return df
                return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)

            # Для правила RCCOMP_375.1.2 (ADR2): только заполненный PERSNUMBER, затем фильтр по KNVV AUFSD F/M
            elif rule_code == "RCCOMP_375.1.2":
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        break
                if person_col:
                    mask = (
                        df[person_col].notna() &
                        (df[person_col].astype(str).str.strip() != '') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'none') &
                        (df[person_col].astype(str).str.strip().str.lower() != 'null')
                    )
                    df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL для {rule_code}: {len(df)} строк")
                    if len(df) == 0:
                        return df
                return self._filter_adr2_by_knvv_aufsd_fm(df, rule_code, table_name)
            
            # Для других правил ADR2
            if "contact_medium_type" in technical_def and "source" in technical_def:
                contact_medium_col = None
                source_col = None
                
                for col in df.columns:
                    col_lower = col.lower()
                    if 'contact_medium' in col_lower or 'medium_type' in col_lower or 'contactmedium' in col_lower:
                        contact_medium_col = col
                        print(f"      [FILTER] Найдена колонка contact_medium_type: {col}")
                    if col_lower == 'source' or 'source' in col_lower:
                        source_col = col
                        print(f"      [FILTER] Найдена колонка source: {col}")
                
                if contact_medium_col and source_col:
                    mask = (
                        (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number') &
                        (df[source_col].astype(str).str.strip().str.lower() == 's4')
                    )
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return filtered_df
                elif contact_medium_col:
                    mask = (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number')
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только contact_medium_type) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return filtered_df
                else:
                    print(f"      [WARN] Не найдены колонки для фильтрации.")
            
            # Если условий нет или они не распознаны, возвращаем исходный DataFrame
            print(f"      [FILTER] Условия фильтрации не найдены или не распознаны")
            return df
            
        except Exception as e:
            print(f"      [WARN] Ошибка при применении фильтра для {rule_code}: {e}")
            traceback.print_exc()
            return df
    
    def _log_skipped_rule(self, rule, table_name, reason, timestamp):
        rule_code = rule.get("rule_code", "UNKNOWN")
        self.results.append({
            "rule_code": rule_code,
            "rule_description": rule.get("rule_description", "Unknown rule"),
            "quality_category": rule.get("quality_category", "Unknown"),
            "table_name": table_name,
            "column_checked": rule.get("column_name_checked", ""),
            "matched_column": "",
            "total_records": 0,
            "passed": 0,
            "failed": 0,
            "total_evaluated": 0,
            "success_rate_%": 0,
            "execution_time_sec": 0,
            "check_date": timestamp,
            "status": "ПРОПУЩЕНО",
            "status_color": "gray",
            "error_file": "Нет",
            "comments": f"Пропущено: {reason}"
        })
    
    def _log_failed_rule(self, rule, table_name, error_message, timestamp):
        rule_code = rule.get("rule_code", "UNKNOWN")
        self.results.append({
            "rule_code": rule_code,
            "rule_description": rule.get("rule_description", "Unknown rule"),
            "quality_category": rule.get("quality_category", "Unknown"),
            "table_name": table_name,
            "column_checked": rule.get("column_name_checked", ""),
            "matched_column": "",
            "total_records": 0,
            "passed": 0,
            "failed": 0,
            "total_evaluated": 0,
            "success_rate_%": 0,
            "execution_time_sec": 0,
            "check_date": timestamp,
            "status": "ОШИБКА ВЫПОЛНЕНИЯ",
            "status_color": "dark_red",
            "error_file": "Нет",
            "comments": f"Ошибка: {error_message}"
        })

    def _save_rule_errors(self, timestamp=None):
        """Сохраняет ошибки в файлы Excel с датой и временем в имени папки"""
        if not self.rule_errors:
            print(f"\n[INFO] Нет ошибок для сохранения")
            return
        
        print(f"\n[INFO] Сохранение ошибок по правилам...")
        print(f"   Всего правил с ошибками в памяти: {len(self.rule_errors)}")
        
        # Подсчитываем ошибки по таблицам
        errors_by_table = {}
        for key, error_data in self.rule_errors.items():
            table_name = error_data['table_name']
            if table_name not in errors_by_table:
                errors_by_table[table_name] = 0
            errors_by_table[table_name] += 1
        
        print(f"   Ошибок по таблицам: {errors_by_table}")
        
        # Используем переданный timestamp или создаем новый
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
        # Создаем папку для ошибок с датой и временем
        errors_dir = os.path.join(self.output_dir, f"errors_{timestamp}")
        os.makedirs(errors_dir, exist_ok=True)
        print(f"   [INFO] Создана папка для ошибок: {errors_dir}")
        
        # —— ОБЯЗАТЕЛЬНАЯ ПОДГОТОВКА ФАЙЛА RCCONF_39.5: только пустой PERSNUMBER + колонка PARTNER из BUT020 ——
        key_39_5 = next((k for k in self.rule_errors if self.rule_errors[k].get('rule_code') == "RCCONF_39.5" and str(self.rule_errors[k].get('table_name', '')).strip().upper() == "ADR2"), None)
        if key_39_5 is not None:
            ed = self.rule_errors[key_39_5]
            raw = ed.get('error_df')
            if raw is not None and not raw.empty:
                out = raw.copy()
                # Ищем PERSNUMBER по подстроке (без учёта регистра)
                pcol = next((c for c in out.columns if 'PERSNUMBER' in str(c).upper() or ('PERSON' in str(c).upper() and 'NUMBER' in str(c).upper())), None)
                if pcol is not None:
                    v = out[pcol].astype(str).str.strip().str.upper()
                    empty = out[pcol].isna() | (v == '') | (v.isin(['NONE', 'NAN', 'NULL', '-', '.']))
                    out = out.loc[empty].copy()
                    print(f"   [RCCONF_39.5] Оставлены только строки с пустым PERSNUMBER: {len(out):,} из {len(raw):,}")
                else:
                    print(f"   [RCCONF_39.5] Колонка PERSNUMBER не найдена. Имена колонок: {list(out.columns)}")
                # Жёстко убираем номера 10 цифр с 9 — они не ошибочные
                def _dig(v):
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return ''
                    s = str(v).strip().replace('\ufeff', '')
                    s = re.sub(r'\s+', '', s)
                    try:
                        if isinstance(v, (int, float)) and v == int(v):
                            return str(int(v))
                    except (ValueError, TypeError):
                        pass
                    if re.match(r'^\d+\.0+$', s):
                        return str(int(float(s)))
                    return re.sub(r'\D', '', s)
                def _is_10_9(v):
                    d = _dig(v)
                    return len(d) == 10 and d.startswith('9')
                tcol = None
                if 'DQ_COLUMN_CHECKED' in out.columns:
                    try:
                        cn = out['DQ_COLUMN_CHECKED'].iloc[0]
                        if cn and str(cn).strip() in out.columns:
                            tcol = str(cn).strip()
                    except Exception:
                        pass
                if tcol is None:
                    tcol = next((c for c in out.columns if 'TEL' in str(c).upper() and ('NUMBER' in str(c).upper() or 'NR' in str(c).upper() or 'NUM' in str(c).upper())), None)
                if tcol is not None and not out.empty:
                    drop = out[tcol].apply(_is_10_9)
                else:
                    drop = pd.Series(False, index=out.index)
                    for c in out.columns:
                        if 'DQ_' in str(c):
                            continue
                        drop = drop | out[c].apply(_is_10_9)
                if drop.any():
                    n_before = len(out)
                    out = out.loc[~drop].copy()
                    print(f"   [RCCONF_39.5] Убраны из выгрузки номера 10 цифр с 9: {n_before - len(out):,} строк")
                # Подтягиваем PARTNER из BUT020 по ADDRNUMBER
                acol = next((c for c in out.columns if 'ADDRNUMBER' in str(c).upper()), None)
                if acol is not None and not out.empty:
                    try:
                        import sqlite3
                        conn = connect_sqlite(self.db_path)
                        tables_check = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name='BUT020'", conn)
                        if not tables_check.empty:
                            but020_df = pd.read_sql_query("SELECT ADDRNUMBER, PARTNER FROM BUT020", conn)
                            conn.close()
                            # В BUT020 колонка называется PARTNER (капс) — приводим к капсу
                            _pc = next((c for c in but020_df.columns if str(c).upper() == 'PARTNER'), None)
                            if _pc:
                                but020_df = but020_df.rename(columns={_pc: 'PARTNER'})
                            _ac = next((c for c in but020_df.columns if str(c).upper() == 'ADDRNUMBER'), None)
                            addr_col_but = _ac if _ac else 'ADDRNUMBER'
                            j = lambda x: (str(x).strip().lstrip('0') or '0')
                            out['_ak'] = out[acol].apply(j)
                            but020_df['_ak'] = but020_df[addr_col_but].apply(j)
                            out = out.merge(but020_df[['_ak', 'PARTNER']].drop_duplicates('_ak'), on='_ak', how='left')
                            out = out.drop(columns=['_ak'], errors='ignore')
                            # В файле колонка строго PARTNER (капс)
                            _pout = next((c for c in out.columns if str(c).upper() == 'PARTNER' and c != 'PARTNER'), None)
                            if _pout:
                                out = out.rename(columns={_pout: 'PARTNER'})
                            if 'PARTNER_y' in out.columns:
                                out['PARTNER'] = out['PARTNER_y']
                                out = out.drop(columns=['PARTNER_x', 'PARTNER_y'], errors='ignore')
                            if 'PARTNER' in out.columns and acol in out.columns:
                                c_order = [c for c in out.columns if c != 'PARTNER']
                                ix = c_order.index(acol) + 1
                                out = out[c_order[:ix] + ['PARTNER'] + c_order[ix:]]
                            print(f"   [RCCONF_39.5] Добавлена колонка PARTNER из BUT020 по ADDRNUMBER: {out['PARTNER'].notna().sum():,} из {len(out):,}")
                        else:
                            conn.close()
                    except Exception as e:
                        print(f"   [RCCONF_39.5] Ошибка при добавлении PARTNER: {e}")
                ed['error_df'] = out
        
        saved_count = 0
        
        for key, error_data in self.rule_errors.items():
            try:
                rule_code = error_data['rule_code']
                table_name = error_data['table_name']
                error_df = error_data['error_df']
                
                if error_df is None or error_df.empty:
                    print(f"   [WARN] Пропускаем {key}: error_df пустой")
                    continue
                
                # ADRC: при записи в файл исключаем строки с NAME1=RESERVED (не должны попадать в выгрузку ошибок)
                if (str(table_name or "").strip().upper() == "ADRC"):
                    name1_col = None
                    for c in error_df.columns:
                        if str(c).strip().upper() == "NAME1":
                            name1_col = c
                            break
                    if name1_col is None:
                        name1_col = self._find_column_alternative(error_df.columns, "NAME1", table_name)
                    if name1_col is None:
                        best_col, best_count = None, 0
                        for c in error_df.columns:
                            try:
                                cnt = (error_df[c].astype(str).str.strip().str.upper() == "RESERVED").sum()
                                if cnt > best_count:
                                    best_count, best_col = cnt, c
                            except Exception:
                                pass
                        if best_col and best_count > 0:
                            name1_col = best_col
                    if name1_col and name1_col in error_df.columns:
                        val_str = error_df[name1_col].astype(str).str.strip().str.upper()
                        error_df = error_df[val_str != "RESERVED"].copy()
                        if error_df.empty:
                            print(f"   [ADRC] {key}: после исключения NAME1=RESERVED записей не осталось, файл не создаётся")
                            continue
                
                is_adr2 = (str(table_name or "").strip().upper() == "ADR2")
                # Для 38.5, 39.3, 39.5 (ADR2) подтягиваем PARTNER из BUT020 по ADDRNUMBER (39.5 уже подготовлен выше)
                acol = next((c for c in error_df.columns if 'ADDRNUMBER' in str(c).upper()), None)
                if rule_code in ["RCCONF_38.5", "RCCONF_39.3", "RCCONF_39.3.2", "RCCONF_39.5", "RCCONF_39.5.2"] and is_adr2 and acol is not None:
                    try:
                        import sqlite3
                        conn = connect_sqlite(self.db_path)
                        tc = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name='BUT020'", conn)
                        if not tc.empty:
                            but020_df = pd.read_sql_query("SELECT ADDRNUMBER, PARTNER FROM BUT020", conn)
                            conn.close()
                            _pc = next((c for c in but020_df.columns if str(c).upper() == 'PARTNER'), None)
                            if _pc:
                                but020_df = but020_df.rename(columns={_pc: 'PARTNER'})
                            _ac = next((c for c in but020_df.columns if str(c).upper() == 'ADDRNUMBER'), None)
                            addr_but = _ac if _ac else 'ADDRNUMBER'
                            j = lambda x: (str(x).strip().lstrip('0') or '0')
                            error_df = error_df.assign(_ak=error_df[acol].apply(j))
                            but020_df['_ak'] = but020_df[addr_but].apply(j)
                            error_df = error_df.merge(but020_df[['_ak', 'PARTNER']].drop_duplicates('_ak'), on='_ak', how='left')
                            error_df = error_df.drop(columns=['_ak'], errors='ignore')
                            if 'PARTNER_y' in error_df.columns:
                                error_df['PARTNER'] = error_df['PARTNER_y']
                                error_df = error_df.drop(columns=['PARTNER_x', 'PARTNER_y'], errors='ignore')
                            _pout = next((c for c in error_df.columns if str(c).upper() == 'PARTNER' and c != 'PARTNER'), None)
                            if _pout:
                                error_df = error_df.rename(columns={_pout: 'PARTNER'})
                            if 'PARTNER' in error_df.columns and acol in error_df.columns:
                                cols = [c for c in error_df.columns if c != 'PARTNER']
                                ix = cols.index(acol) + 1
                                error_df = error_df[cols[:ix] + ['PARTNER'] + cols[ix:]]
                            print(f"   [INFO] {rule_code}: добавлена колонка PARTNER из BUT020 по ADDRNUMBER")
                        else:
                            conn.close()
                    except Exception as e:
                        print(f"   [ERROR] {rule_code}: ошибка при добавлении PARTNER: {e}")
                
                # Для таблицы ADR2 в файлах ошибок должна быть колонка AUFSD (из KNVV по PARTNER=KUNNR)
                if is_adr2:
                    need_aufsd = 'AUFSD' not in error_df.columns or error_df['AUFSD'].isna().all()
                    if need_aufsd:
                        partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                        if partner_col is None and acol is not None:
                            try:
                                import sqlite3
                                conn = connect_sqlite(self.db_path)
                                tc = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name='BUT020'", conn)
                                if not tc.empty:
                                    but020_df = pd.read_sql_query("SELECT ADDRNUMBER, PARTNER FROM BUT020", conn)
                                    conn.close()
                                    _ac = next((c for c in but020_df.columns if str(c).upper() == 'ADDRNUMBER'), None)
                                    _pc = next((c for c in but020_df.columns if str(c).upper() == 'PARTNER'), None)
                                    if _ac and _pc:
                                        j = lambda x: (str(x).strip().lstrip('0') or '0')
                                        error_df = error_df.assign(_ak=error_df[acol].apply(j))
                                        but020_df['_ak'] = but020_df[_ac].apply(j)
                                        error_df = error_df.merge(but020_df[['_ak', _pc]].drop_duplicates('_ak'), on='_ak', how='left')
                                        error_df = error_df.drop(columns=['_ak'], errors='ignore')
                                        partner_col = _pc
                                else:
                                    conn.close()
                            except Exception as e:
                                print(f"   [WARN] {rule_code}: не удалось подтянуть PARTNER для AUFSD: {e}")
                        if partner_col is None:
                            error_df['AUFSD'] = None
                            print(f"   [WARN] {rule_code}: колонка PARTNER не найдена, AUFSD пустая")
                        else:
                            try:
                                knvv_df = self.memory_manager.get_table("KNVV")
                                # Если KNVV не загружена в память (например, проверяли только ADR2) — подгружаем из БД
                                if (knvv_df is None or knvv_df.empty) and getattr(self, 'db_path', None):
                                    import sqlite3
                                    conn = connect_sqlite(self.db_path)
                                    tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                    knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                                    if knvv_name:
                                        knvv_df = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                                    conn.close()
                                if knvv_df is not None and not knvv_df.empty:
                                    kunnr_col = next((c for c in knvv_df.columns if str(c).upper() in ('KUNNR', 'KUNNR_KNVV')), None)
                                    aufsd_knvv = next((c for c in knvv_df.columns if 'AUFSD' in str(c).upper()), None)
                                    if kunnr_col and aufsd_knvv:
                                        knvv_aufsd = knvv_df[[kunnr_col, aufsd_knvv]].drop_duplicates(subset=[kunnr_col], keep='first')
                                        knvv_aufsd = knvv_aufsd.rename(columns={kunnr_col: '_partner', aufsd_knvv: 'AUFSD'})
                                        knvv_aufsd['_partner'] = knvv_aufsd['_partner'].astype(str).str.strip()
                                        error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                        error_df['_partner'] = error_df[partner_col].astype(str).str.strip()
                                        error_df = error_df.merge(knvv_aufsd[['_partner', 'AUFSD']], on='_partner', how='left')
                                        error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                        print(f"   [INFO] {rule_code}: добавлена колонка AUFSD из KNVV для ADR2")
                                    else:
                                        error_df['AUFSD'] = None
                                        print(f"   [WARN] {rule_code}: в KNVV не найдены колонки KUNNR или AUFSD. Колонки: {list(knvv_df.columns)[:20]}")
                                else:
                                    error_df['AUFSD'] = None
                                    # Разделяем: таблицы нет в БД или таблица есть, но 0 строк
                                    if getattr(self, 'db_path', None):
                                        import sqlite3
                                        conn = connect_sqlite(self.db_path)
                                        tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                        knvv_in_db = any(str(r[0]).strip().upper() == 'KNVV' for r in tables.values)
                                        conn.close()
                                        if not knvv_in_db:
                                            print(f"   [WARN] {rule_code}: таблица KNVV не найдена в БД (проверьте имя таблицы), колонка AUFSD пустая")
                                        else:
                                            print(f"   [WARN] {rule_code}: таблица KNVV в БД есть, но пуста (0 строк) или не загрузилась, колонка AUFSD пустая")
                                    else:
                                        print(f"   [WARN] {rule_code}: таблица KNVV не найдена, колонка AUFSD пустая")
                            except Exception as e:
                                print(f"   [WARN] {rule_code}: ошибка при добавлении AUFSD: {e}")
                                error_df['AUFSD'] = None
                    else:
                        # AUFSD уже есть — убеждаемся, что колонка в нужном месте (после ADDRNUMBER/PARTNER)
                        pass
                
                # Для таблиц AUSP_143, AUSP_604, AUSP_148, AUSP_151 в файлах ошибок должна быть колонка AUFSD
                # Цепочка: AUSP.PARTNER_GUID -> BUT000.PARTNER_GUID -> BUT000.PARTNER (=KUNNR) -> KNA1.AUFSD
                is_ausp_table = (str(table_name or "").strip().upper() in ("AUSP_143", "AUSP_604", "AUSP_148", "AUSP_151"))
                if is_ausp_table:
                    need_aufsd_ausp = 'AUFSD' not in error_df.columns or error_df['AUFSD'].isna().all()
                    if need_aufsd_ausp:
                        partner_guid_col = next((c for c in error_df.columns if 'PARTNER_GUID' in str(c).upper() or (str(c).upper() == 'PARTNERGUID')), None)
                        if partner_guid_col is None:
                            error_df['AUFSD'] = None
                            print(f"   [WARN] {rule_code} ({table_name}): колонка PARTNER_GUID не найдена в ошибках, AUFSD пустая")
                        else:
                            try:
                                import sqlite3
                                conn = connect_sqlite(self.db_path)
                                # BUT000: PARTNER_GUID -> PARTNER
                                tables_list = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                but000_name = None
                                for _n in tables_list['name']:
                                    if str(_n).strip().upper() == 'BUT000':
                                        but000_name = _n
                                        break
                                if but000_name is None:
                                    conn.close()
                                    error_df['AUFSD'] = None
                                    print(f"   [WARN] {rule_code}: таблица BUT000 не найдена, AUFSD пустая")
                                else:
                                    but000_cols = pd.read_sql_query(f"PRAGMA table_info(\"{but000_name}\")", conn)
                                    col_list = [r[1] for r in but000_cols.values]
                                    pg_but = next((c for c in col_list if 'PARTNER_GUID' in str(c).upper() or str(c).upper() == 'PARTNERGUID'), None)
                                    p_but = next((c for c in col_list if str(c).upper() == 'PARTNER'), None)
                                    if not pg_but or not p_but:
                                        conn.close()
                                        error_df['AUFSD'] = None
                                        print(f"   [WARN] {rule_code}: в BUT000 не найдены PARTNER_GUID или PARTNER")
                                    else:
                                        but000_df = pd.read_sql_query(f"SELECT \"{pg_but}\", \"{p_but}\" FROM \"{but000_name}\"", conn)
                                        but000_df = but000_df.rename(columns={pg_but: '_pg', p_but: '_partner'})
                                        but000_df = but000_df.drop_duplicates(subset=['_pg'], keep='first')
                                        conn.close()
                                        # Присоединяем PARTNER к ошибкам по PARTNER_GUID
                                        but000_df['_pg'] = but000_df['_pg'].astype(str).str.strip()
                                        error_df = error_df.merge(but000_df[['_pg', '_partner']], left_on=partner_guid_col, right_on='_pg', how='left')
                                        error_df = error_df.drop(columns=['_pg'], errors='ignore')
                                        if '_partner' not in error_df.columns:
                                            error_df['AUFSD'] = None
                                            print(f"   [WARN] {rule_code}: не удалось подтянуть PARTNER из BUT000")
                                        else:
                                            # KNA1: KUNNR (= PARTNER) -> AUFSD
                                            kna1_df = self._get_table_for_rules("KNA1")
                                            if kna1_df is None or kna1_df.empty:
                                                conn = connect_sqlite(self.db_path)
                                                kna1_tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                                kna1_name = next((r[0] for r in kna1_tables.values if 'KNA1' in str(r[0]).upper()), None)
                                                if kna1_name:
                                                    kna1_df = pd.read_sql_query(f'SELECT * FROM "{kna1_name}"', conn)
                                                conn.close()
                                            if kna1_df is not None and not kna1_df.empty:
                                                kunnr_col = next((c for c in kna1_df.columns if str(c).upper() == 'KUNNR'), None)
                                                aufsd_col = next((c for c in kna1_df.columns if 'AUFSD' in str(c).upper()), None)
                                                if kunnr_col and aufsd_col:
                                                    kna1_aufsd = kna1_df[[kunnr_col, aufsd_col]].drop_duplicates(subset=[kunnr_col], keep='first')
                                                    kna1_aufsd = kna1_aufsd.rename(columns={kunnr_col: '_kunnr', aufsd_col: 'AUFSD'})
                                                    kna1_aufsd['_kunnr'] = kna1_aufsd['_kunnr'].astype(str).str.strip()
                                                    error_df['_partner'] = error_df['_partner'].astype(str).str.strip()
                                                    error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                                    error_df = error_df.merge(kna1_aufsd[['_kunnr', 'AUFSD']], left_on='_partner', right_on='_kunnr', how='left')
                                                    error_df = error_df.drop(columns=['_partner', '_kunnr'], errors='ignore')
                                                    print(f"   [INFO] {rule_code} ({table_name}): добавлена колонка AUFSD из KNA1 (AUSP.PARTNER_GUID -> BUT000 -> KNA1)")
                                                else:
                                                    error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                                    error_df['AUFSD'] = None
                                                    print(f"   [WARN] {rule_code}: в KNA1 не найдены KUNNR или AUFSD")
                                            else:
                                                error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                                error_df['AUFSD'] = None
                                                print(f"   [WARN] {rule_code}: таблица KNA1 не найдена, AUFSD пустая")
                            except Exception as e:
                                error_df['AUFSD'] = None
                                if '_partner' in error_df.columns:
                                    error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                print(f"   [WARN] {rule_code}: ошибка при добавлении AUFSD для AUSP: {e}")
                    error_data['error_df'] = error_df
                
                # Для остальных таблиц (BUT000, KNB1, KNVP, KNVH, ADRC, ZBUT0000P3VVI9 и др.): подтягиваем AUFSD по KUNNR/PARTNER из KNVV или KNA1
                if self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
                    error_df = self._enrich_error_df_kna1_ktokd(error_df, table_name, rule_code)
                    error_data["error_df"] = error_df

                if not is_adr2 and not is_ausp_table:
                    need_aufsd = 'AUFSD' not in error_df.columns or error_df['AUFSD'].isna().all()
                    if need_aufsd:
                        partner_col = next((c for c in error_df.columns if str(c).upper() == 'KUNNR'), None)
                        if partner_col is None:
                            partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                        acol_gen = next((c for c in error_df.columns if 'ADDRNUMBER' in str(c).upper()), None)
                        if partner_col is None and acol_gen is not None:
                            try:
                                import sqlite3
                                conn = connect_sqlite(self.db_path)
                                tc = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name='BUT020'", conn)
                                if not tc.empty:
                                    but020_df = pd.read_sql_query("SELECT ADDRNUMBER, PARTNER FROM BUT020", conn)
                                    conn.close()
                                    _ac = next((c for c in but020_df.columns if str(c).upper() == 'ADDRNUMBER'), None)
                                    _pc = next((c for c in but020_df.columns if str(c).upper() == 'PARTNER'), None)
                                    if _ac and _pc:
                                        j = lambda x: (str(x).strip().lstrip('0') or '0')
                                        error_df = error_df.assign(_ak=error_df[acol_gen].apply(j))
                                        but020_df['_ak'] = but020_df[_ac].apply(j)
                                        error_df = error_df.merge(but020_df[['_ak', _pc]].drop_duplicates('_ak'), on='_ak', how='left')
                                        error_df = error_df.drop(columns=['_ak'], errors='ignore')
                                        if _pc in error_df.columns:
                                            partner_col = _pc
                                        else:
                                            partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                                            if partner_col is None:
                                                partner_col = _pc
                                        print(f"   [INFO] {rule_code}: добавлена колонка PARTNER из BUT020 для подтягивания AUFSD")
                                else:
                                    conn.close()
                            except Exception as e:
                                print(f"   [WARN] {rule_code}: не удалось подтянуть PARTNER для AUFSD: {e}")
                        if partner_col is not None:
                            aufsd_added = False
                            try:
                                knvv_df = self.memory_manager.get_table("KNVV")
                                if (knvv_df is None or knvv_df.empty) and getattr(self, 'db_path', None):
                                    import sqlite3
                                    conn = connect_sqlite(self.db_path)
                                    tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                    knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                                    if knvv_name:
                                        knvv_df = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                                    conn.close()
                                if knvv_df is not None and not knvv_df.empty:
                                    kunnr_knvv = next((c for c in knvv_df.columns if str(c).upper() in ('KUNNR', 'KUNNR_KNVV')), None)
                                    aufsd_knvv = next((c for c in knvv_df.columns if 'AUFSD' in str(c).upper()), None)
                                    if kunnr_knvv and aufsd_knvv:
                                        knvv_aufsd = knvv_df[[kunnr_knvv, aufsd_knvv]].drop_duplicates(subset=[kunnr_knvv], keep='first')
                                        knvv_aufsd = knvv_aufsd.rename(columns={kunnr_knvv: '_partner', aufsd_knvv: 'AUFSD'})
                                        knvv_aufsd['_partner'] = knvv_aufsd['_partner'].astype(str).str.strip()
                                        error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                        error_df['_partner'] = error_df[partner_col].astype(str).str.strip()
                                        error_df = error_df.merge(knvv_aufsd[['_partner', 'AUFSD']], on='_partner', how='left')
                                        error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                        print(f"   [INFO] {rule_code} ({table_name}): добавлена колонка AUFSD из KNVV")
                                        aufsd_added = True
                                if not aufsd_added:
                                    kna1_df = self._get_table_for_rules("KNA1")
                                    if (kna1_df is None or kna1_df.empty) and getattr(self, 'db_path', None):
                                        import sqlite3
                                        conn = connect_sqlite(self.db_path)
                                        kna1_tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                        kna1_name = next((r[0] for r in kna1_tables.values if 'KNA1' in str(r[0]).upper()), None)
                                        if kna1_name:
                                            kna1_df = pd.read_sql_query(f'SELECT * FROM "{kna1_name}"', conn)
                                        conn.close()
                                    if kna1_df is not None and not kna1_df.empty:
                                        kunnr_col = next((c for c in kna1_df.columns if str(c).upper() == 'KUNNR'), None)
                                        aufsd_col = next((c for c in kna1_df.columns if 'AUFSD' in str(c).upper()), None)
                                        if kunnr_col and aufsd_col:
                                            kna1_aufsd = kna1_df[[kunnr_col, aufsd_col]].drop_duplicates(subset=[kunnr_col], keep='first')
                                            kna1_aufsd = kna1_aufsd.rename(columns={kunnr_col: '_partner', aufsd_col: 'AUFSD'})
                                            kna1_aufsd['_partner'] = kna1_aufsd['_partner'].astype(str).str.strip()
                                            error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                            error_df['_partner'] = error_df[partner_col].astype(str).str.strip()
                                            error_df = error_df.merge(kna1_aufsd[['_partner', 'AUFSD']], on='_partner', how='left')
                                            error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                            print(f"   [INFO] {rule_code} ({table_name}): добавлена колонка AUFSD из KNA1")
                                            aufsd_added = True
                                    if not aufsd_added:
                                        error_df['AUFSD'] = None
                            except Exception as e:
                                error_df['AUFSD'] = None
                                print(f"   [WARN] {rule_code} ({table_name}): ошибка при добавлении AUFSD: {e}")
                        else:
                            error_df['AUFSD'] = None
                
                total_errors = len(error_df)
                is_truncated = error_data.get('is_truncated', False)
                original_error_count = error_data.get('error_count', total_errors)

                # Для отдельных правил/таблиц ограничение 100k снято — пишем до лимита Excel
                tbl = str(table_name or "").strip().upper()
                rule_u = str(rule_code or "").strip().upper()
                no_100k_limit = (
                    tbl in ("ADR2", "BUT000")
                    or rule_u == "RCCONF_18.2"
                    or rule_u == "RCCONF_63.1"
                    or tbl.startswith("DFKKBPTAXNUM")
                )
                limit_save = self.EXCEL_MAX_ROWS if no_100k_limit else self.MAX_ERRORS_TO_SAVE
                
                if total_errors > limit_save:
                    error_df = error_df.head(limit_save)
                    is_truncated = True
                    print(f"   [WARN] {rule_code} ({table_name}): ошибок {original_error_count:,}, в файл сохранено только {limit_save:,} (первые {limit_save:,})")

                # ADR2: дополнительно сохраняем строки ошибок в sqlite БД,
                # чтобы потом можно было анализировать именно "строки как в таблице ADR2".
                if is_adr2:
                    self._save_adr2_rule_errors_to_db(error_df, rule_code=rule_code, run_ts=timestamp)
                
                # Добавляем пометку о превышении лимита в первую строку данных (если есть)
                if is_truncated and len(error_df) > 0:
                    # Добавляем предупреждение как первую строку
                    warning_row = pd.DataFrame([{col: f"[!] ВНИМАНИЕ: Всего ошибок {original_error_count:,}, показано только первые {limit_save:,}" 
                                                if col == error_df.columns[0] else "" for col in error_df.columns}])
                    error_df = pd.concat([warning_row, error_df], ignore_index=True)
                
                # Сохраняем файл: CSV когда ошибок столько, что не поместятся в Excel (или уже > limit_save)
                # ADR2/BUT000: всегда CSV, т.к. пишем все ошибки (могут не поместиться в Excel)
                use_csv = (
                    no_100k_limit
                    or original_error_count > self.EXCEL_MAX_ROWS
                    or total_errors > limit_save
                    or len(error_df) > limit_save
                )
                safe_table_name = self._safe_filename_token(table_name)
                # Финальная страховка: KTOKD из KNA1 в файле ошибок (RCCOMP_113.1 / RCCONF_113.1)
                if (
                    str(table_name or "").strip().upper() == "KNB1"
                    and self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD
                ):
                    error_df = self._enrich_error_df_kna1_ktokd(error_df, table_name, rule_code)
                    if "KTOKD" not in error_df.columns:
                        error_df.insert(0, "KTOKD", "")
                        error_df.insert(1, "KTOKD_SOURCE", "KNA1")
                if use_csv:
                    filename = f"{rule_code}_{safe_table_name}_errors_{timestamp}.csv"
                    filepath = os.path.join(errors_dir, filename)
                    error_df.to_csv(filepath, index=False, encoding='utf-8-sig', sep=';')
                    status_msg = f" ({original_error_count:,} всего, сохранено {len(error_df):,})" if is_truncated else f" ({len(error_df):,} строк)"
                    print(f"   [INFO] Сохранены ошибки в CSV: {filename}{status_msg}")
                else:
                    filename = f"{rule_code}_{safe_table_name}_errors_{timestamp}.xlsx"
                    filepath = os.path.join(errors_dir, filename)
                    error_df.to_excel(filepath, index=False, engine='openpyxl')
                    status_msg = f" ({original_error_count:,} всего, сохранено {len(error_df):,})" if is_truncated else f" ({len(error_df):,} строк)"
                    print(f"   [INFO] Сохранены ошибки в Excel: {filename}{status_msg}")
                
                saved_count += 1
                    
            except Exception as e:
                print(f"   [ERROR] Ошибка сохранения {key}: {e}")
                traceback.print_exc()
        
        print(f"   Сохранено файлов: {saved_count}")

    def _save_totals_by_table(self, timestamp=None):
        """
        Ранее сохранял CSV «Всего записей» в quality_reports/total.
        По текущей настройке ведение этой папки отключено.
        Оставлен как no-op, чтобы не ломать вызовы.
        """
        return
    
    def _create_correct_report(self, report_name: str = "quality_check_report", timestamp=None):
        """Создает цветной отчет в Excel с датой и временем в имени"""
        try:
            if not self.results:
                print(f"\n[INFO] Нет данных для отчета")
                return
            
            # Используем переданный timestamp или создаем новый
            if timestamp is None:
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            
            # Добавляем дату и время к имени файла отчета
            excel_path = os.path.join(self.output_dir, f"{self._safe_filename_token(report_name)}_{timestamp}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Сводка проверок"
            
            # Заголовок
            ws['A1'] = "СВОДКА ПРОВЕРОК КАЧЕСТВА ДАННЫХ"
            ws['A1'].font = Font(bold=True, size=14)
            
            # Подзаголовок
            ws['A2'] = f"Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило"
            ws['A2'].font = Font(size=10)
            
            # Пояснение для ADR2: что входит в «Всего записей»
            ws['A5'] = "Для таблицы ADR2: «Всего записей» считается по внутренним фильтрам (ADDRNUMBER+PARTNER) для правила, отдельный список записей сейчас не формируется автоматически."
            ws['A5'].font = Font(size=9, italic=True)
            
            # Даты
            check_date = self.results[0].get('check_date', '') if self.results else ''
            ws['A3'] = f"Дата проверки: {check_date}"
            ws['A3'].font = Font(size=9, italic=True)
            
            # Пустая строка
            ws['A4'] = ""
            
            # Заголовки таблицы (Таблица + Тип TAXNUM для фильтра по DFKKBPTAXNUM)
            # Всего записей = passed + failed; для ADR2 = число записей по внутреннему счётчику (filtered_adr2_count)
            headers = [
                "Код правила", "Описание", "Категория", "Таблица", "Тип TAXNUM",
                "Колонка", "Всего записей", "Успешно", "Ошибок",
                "% успеха", "Статус", "Время (сек)", "Комментарии",
                "Список записей (файл)"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col_num, value=header)
                cell.fill = self.colors['header']
                cell.font = self.colors['header_font']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Данные
            row_num = 7
            for result in self.results:
                table_name = result.get('table_name', '')
                if table_name in self.DFKKBPTAXNUM_ALIASES:
                    # Для DFKKBPTAXNUM1/2/3/5 показываем DFKKBPTAXNUM + тип TAXNUM; для DFKKBPTAXNUM_ALL — тип ALL
                    table_display = "DFKKBPTAXNUM"
                    suffix = table_name.replace("DFKKBPTAXNUM", "")
                    taxnum_type = "ALL" if suffix == "_ALL" else suffix  # "1", "2", "3", "5", "ALL"
                else:
                    table_display = table_name
                    taxnum_type = ""
                # Для ADR2: «Всего записей» = число записей в файле со списком (то, что реально подсчитывается)
                is_adr2 = (str(table_name or "").strip().upper() == "ADR2")
                if is_adr2 and result.get("filtered_adr2_count") is not None:
                    total_rec_display = result.get("filtered_adr2_count")
                    list_file_display = result.get("filtered_adr2_file", "") or ""
                else:
                    total_rec_display = result.get('total_records', 0)
                    list_file_display = ""
                values = [
                    result.get('rule_code', ''),
                    result.get('rule_description', ''),
                    result.get('quality_category', ''),
                    table_display,
                    taxnum_type,
                    result.get('column_checked', ''),
                    total_rec_display,
                    result.get('passed', 0),
                    result.get('failed', 0),
                    result.get('success_rate_%', 0),
                    result.get('status', ''),
                    result.get('execution_time_sec', 0),
                    result.get('comments', ''),
                    list_file_display,
                ]
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = self.colors['normal_font']
                    # Ссылка на файл со списком записей (ADR2): открыть — увидеть, что именно входит в «Всего записей»
                    if col_num == 14 and value and isinstance(value, str) and os.path.isfile(value):
                        try:
                            path_uri = "file:///" + value.replace("\\", "/").lstrip("/")
                            cell.hyperlink = path_uri
                            cell.font = Font(color="0563C1", underline="single")
                        except Exception:
                            pass
                    
                    # Цветовая индикация статуса (колонка "Статус")
                    if col_num == 11:
                        status_color = result.get('status_color', '')
                        if status_color == 'green':
                            cell.fill = self.colors['green']
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif status_color == 'red':
                            cell.fill = self.colors['red']
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif status_color == 'orange':
                            cell.fill = self.colors['orange']
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif status_color == 'dark_red':
                            cell.fill = self.colors['dark_red']
                            cell.font = Font(color='FFFFFF', bold=True)
                    
                    # Цвет для количества ошибок (колонка "Ошибок")
                    elif col_num == 9:
                        failed_count = result.get('failed', 0)
                        if failed_count == 0:
                            cell.fill = self.colors['green']
                            cell.font = Font(color='000000', bold=True)  # Черный шрифт на светло-зеленом фоне (C6EFCE)
                        elif failed_count > self.MAX_ERRORS_TO_SAVE:
                            cell.fill = self.colors['orange']
                            cell.font = Font(color='000000', bold=True)  # Черный шрифт на оранжевом фоне
                        else:
                            cell.fill = self.colors['red']
                            cell.font = Font(color='000000', bold=True)  # Черный шрифт на светло-красном фоне (FFC7CE)
                
                row_num += 1
            
            # Автонастройка ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            print(f"\n[INFO] Цветной отчет сохранен: {excel_path}")
            
        except Exception as e:
            print(f"\n[ERROR] Ошибка при создании отчета: {e}")
            traceback.print_exc()
    
    def _print_final_statistics(self):
        total_rules = self.processed_rules + self.skipped_rules
        successful_rules = len([r for r in self.results if r.get('status') == 'УСПЕШНО'])
        failed_rules = len([r for r in self.results if r.get('status') == 'ОШИБКИ'])
        suspicious_rules = len([r for r in self.results if r.get('status') == 'ПОДОЗРИТЕЛЬНО'])
        mass_rules = len([r for r in self.results if r.get('status') == 'МАССОВЫЕ ОШИБКИ'])
        
        print(f"\n" + "="*100)
        print(f"\033[1mИТОГОВАЯ СТАТИСТИКА:\033[0m")
        print(f"="*100)
        print(f"Всего правил: \033[1m{total_rules}\033[0m")
        print(f"  [OK] Успешно:      \033[92m{successful_rules:4d}\033[0m ({successful_rules/total_rules*100:.1f}%)")
        print(f"  [!] Ошибки:       \033[91m{failed_rules:4d}\033[0m ({failed_rules/total_rules*100:.1f}%)")
        print(f"  [!] Подозрительные: \033[93m{suspicious_rules:4d}\033[0m ({suspicious_rules/total_rules*100:.1f}%)")
        print(f"  ⚡ Массовые:      \033[91m{mass_rules:4d}\033[0m ({mass_rules/total_rules*100:.1f}%)")
        print(f"  ↻ Пропущено:    \033[90m{self.skipped_rules:4d}\033[0m ({self.skipped_rules/total_rules*100:.1f}%)")
        print(f"="*100)
        
        # Выводим подозрительные правила
        if suspicious_rules > 0:
            print(f"\n\033[93mПОДОЗРИТЕЛЬНЫЕ ПРАВИЛА (требуют проверки логики):\033[0m")
            suspicious_list = [r for r in self.results if r.get('status') in ['ПОДОЗРИТЕЛЬНО', 'МАССОВЫЕ ОШИБКИ']]
            for rule in suspicious_list[:10]:
                print(f"  • {rule['rule_code']:20} - {rule['table_name']:15} - {rule['failed']:,} ошибок ({rule['success_rate_%']:.1f}% успеха)")
        
        # Выводим время выполнения
        overall_time = time.time() - self.start_time
        print(f"\nВремя выполнения: {overall_time:.2f} сек")
        if overall_time > 0:
            print(f"Скорость: {self.processed_rules/overall_time:.1f} правил/сек")

    def _safe_filename_token(self, value):
        """Возвращает безопасное имя для файла/части пути."""
        s = str(value or "").strip()
        if not s:
            return "unknown"
        s = re.sub(r'[\\/:*?"<>|]+', "_", s)
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"_+", "_", s).strip("._")
        return s or "unknown"


class InequalityValidator:
    def __init__(self, rule_info):
        self.rule_info = rule_info
    
    def validate(self, df, column_name, second_column=None, **kwargs):
        """Ошибка когда оба поля заполнены и равны (не должны быть одинаковы)."""
        total_rows = len(df)
        
        if not second_column:
            second_column = self._find_second_column(df.columns, column_name)
        
        if not second_column or second_column not in df.columns:
            print(f"      [WARN] Вторая колонка не найдена для сравнения")
            return total_rows, 0, pd.DataFrame()
        
        v1 = df[column_name].astype(str).str.strip().str.lower().fillna('')
        v2 = df[second_column].astype(str).str.strip().str.lower().fillna('')
        both_filled = (v1 != '') & (v2 != '')
        equal = (v1 == v2)
        error_mask = both_filled & equal
        error_indices = df.index[error_mask].tolist()
        error_count = len(error_indices)
        
        if error_count > 0:
            error_df = df.loc[error_indices].copy()
            error_df['error_type'] = 'DUPLICATE_VALUES'
            error_df['error_message'] = f'{column_name} не должно быть равно {second_column}'
        else:
            error_df = pd.DataFrame()
        return total_rows, error_count, error_df
    
    def _find_second_column(self, available_columns, first_column):
        first_lower = first_column.lower()
        
        if '2' in first_lower or 'org2' in first_lower or 'name2' in first_lower:
            for col in available_columns:
                col_lower = col.lower()
                if '1' in col_lower or 'org1' in col_lower or 'name1' in col_lower:
                    return col
        
        return None


__all__ = ['FastDataQualityChecker', 'InequalityValidator']