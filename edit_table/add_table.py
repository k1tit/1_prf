import os
import re
import sys
import pandas as pd
import sqlite3
from glob import glob
import gc
import time


def _bootstrap_dq_project() -> str:
    """
    Скрипт лежит в edit_table/ — utils/ и main.py в родительском каталоге (data_quality_checker).
    """
    here = os.path.dirname(os.path.abspath(__file__))
    parent = os.path.dirname(here)
    for candidate in (parent, here):
        if os.path.isfile(os.path.join(candidate, "utils", "sqlite_safe.py")):
            if candidate not in sys.path:
                sys.path.insert(0, candidate)
            return candidate
    if parent not in sys.path:
        sys.path.insert(0, parent)
    return parent


_PROJECT_ROOT = _bootstrap_dq_project()

from utils.sqlite_safe import connect_sqlite, probe_db_writable, is_lock_error  # noqa: E402

try:
    from utils.sqlite_safe import resolve_database_path  # noqa: E402
except ImportError:
    from utils.sqlite_safe import DEFAULT_DB_FILENAME  # noqa: E402

    def resolve_database_path(project_root, cli_path=None, must_exist=False):
        """Запасной вариант для старого sqlite_safe без resolve_database_path."""
        if cli_path and str(cli_path).strip():
            p = cli_path if os.path.isabs(cli_path) else os.path.join(project_root, cli_path)
        else:
            cfg = os.path.join(project_root, "config", "database.json")
            name = DEFAULT_DB_FILENAME
            if os.path.isfile(cfg):
                try:
                    import json
                    with open(cfg, encoding="utf-8") as f:
                        data = json.load(f)
                    if isinstance(data, dict) and data.get("database"):
                        name = str(data["database"]).strip()
                except Exception:
                    pass
            p = os.path.join(project_root, name)
        if must_exist and not os.path.isfile(p):
            raise FileNotFoundError(f"Файл базы данных не найден: {p}")
        return p, "config/database.json или DEFAULT_DB_FILENAME"


def _resolve_db_path(db_path=None):
    """Та же БД, что у main.py: config/database.json → DQ_DATABASE → --db."""
    if db_path is not None and str(db_path).strip():
        path, _ = resolve_database_path(_PROJECT_ROOT, db_path)
        return path
    path, _ = resolve_database_path(_PROJECT_ROOT)
    return path


def _resolve_data_path(data_folder=None):
    """Папка db/ относительно корня проекта, не текущего cwd."""
    folder = data_folder if data_folder is not None else DEFAULT_DATA_FOLDER_REL
    if os.path.isabs(folder):
        return folder
    return os.path.join(_PROJECT_ROOT, folder)


# ========== НАСТРОЙКИ ==========
# Почему загрузка может быть долгой:
#   - Чтение Excel (openpyxl): большие .xlsx парсятся целиком; для ускорения используется read_only режим.
#   - Финальная дедупликация по всей таблице (CREATE TABLE AS SELECT DISTINCT) — очень дорогая на больших объёмах.
#   - Много мелких коммитов — замедляет; сейчас один коммит в конце.
# Ускорение: вызовите merge_and_load_xlsx_files_fast(..., skip_final_dedup=True) или используйте merge_and_load_xlsx_files_ultra_fast.
DEFAULT_TABLE_NAME = "T052U"      # Имя загружаемой таблицы в БД (при одиночной загрузке)
DEFAULT_DATA_FOLDER_REL = "db"    # Подпапки = имена таблиц, внутри — .xlsx / .xls / .csv
DEFAULT_DB_PATH = _resolve_db_path()
DEFAULT_DATA_FOLDER = _resolve_data_path()
# =====================================================


sys.stdout.flush()
sys.stderr.flush()


# Логирование при импорте модуля
print("=" * 80, file=sys.stderr)
print("МОДУЛЬ add_table_to_DB.py ЗАГРУЖЕН", file=sys.stderr)
print("=" * 80, file=sys.stderr)
sys.stderr.flush()


def _dedup_table_in_db(conn, table_name):
    """
    Удаляет дубликаты в таблице по полному совпадению строк (все колонки).
    Не по одному полю — только строки, полностью идентичные по всем полям.
    Возвращает (было_строк, стало_строк, удалено).
    """
    escaped = f'"{table_name}"'
    cursor = conn.cursor()
    cursor.execute(f"SELECT COUNT(*) FROM {escaped}")
    before = cursor.fetchone()[0]
    if before == 0:
        return before, 0, 0
    # SELECT DISTINCT * — уникальность по всей строке (все колонки)
    tmp = f'"{table_name}_dedup_tmp"'
    cursor.execute(f"CREATE TABLE {tmp} AS SELECT DISTINCT * FROM {escaped}")
    cursor.execute(f"SELECT COUNT(*) FROM {tmp}")
    after = cursor.fetchone()[0]
    removed = before - after
    if removed == 0:
        cursor.execute(f"DROP TABLE {tmp}")
        return before, after, 0
    cursor.execute(f"DELETE FROM {escaped}")
    cursor.execute(f"INSERT INTO {escaped} SELECT * FROM {tmp}")
    cursor.execute(f"DROP TABLE {tmp}")
    return before, after, removed


def _sanitize_header_cell(value) -> str:
    """Имя колонки для SQLite: без пробелов/точек, без «невидимых» символов."""
    s = str(value).strip() if value is not None else ""
    if not s or s.lower() == "nan":
        return ""
    try:
        import unicodedata
        s = unicodedata.normalize("NFKC", s)
    except Exception:
        pass
    # пробелы, неразрывный пробел, табы → _
    s = re.sub(r"[\s\u00a0\u200b\u200c\u200d\ufeff]+", "_", s)
    s = re.sub(r"[.\-/\\:]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _unique_sqlite_column_names(raw_names) -> tuple[list[str], list[tuple[str, str]]]:
    """
    Уникальные имена колонок для CREATE TABLE / INSERT (SQLite не допускает дубликаты).
    Grpng_Key + Grpng.Key + GRPNG_KEY → Grpng_Key, Grpng_Key_1, Grpng_Key_2 …
    """
    result: list[str] = []
    renamed: list[tuple[str, str]] = []
    seen_count: dict[str, int] = {}
    for i, raw in enumerate(raw_names):
        base = _sanitize_header_cell(raw)
        if not base:
            base = f"col_{i}"
        # SQLite без учёта регистра в именах колонок
        key = base.upper()
        n = seen_count.get(key, 0)
        seen_count[key] = n + 1
        if n == 0:
            result.append(base)
        else:
            new_name = f"{base}_{n}"
            result.append(new_name)
            renamed.append((str(raw).strip() if raw is not None else base, new_name))
    return result, renamed


def _normalize_df_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Переименовать дубли в заголовках перед записью в SQLite (всегда, не только при WARN)."""
    if df is None or df.empty:
        return df
    raw_cols = list(df.columns)
    new_cols, renamed = _unique_sqlite_column_names(raw_cols)
    if renamed:
        print(f"  [WARN] В выгрузке {len(renamed)} дубль(ей) заголовка — переименованы для SQLite:")
        for old, new in renamed[:8]:
            print(f"         {old!r} -> {new!r}")
        if len(renamed) > 8:
            print(f"         ... и ещё {len(renamed) - 8}")
    need_assign = (
        raw_cols != new_cols
        or len(new_cols) != len(set(new_cols))
        or (hasattr(df.columns, "duplicated") and bool(df.columns.duplicated().any()))
    )
    if need_assign:
        out = df.copy()
        out.columns = new_cols
        return out
    return df


class ProgressBar:
    """Простой прогресс-бар для отображения прогресса"""
   
    @staticmethod
    def show(current, total, prefix="", suffix="", length=50):
        """
        Отображает прогресс-бар
       
        Args:
            current: текущее значение
            total: общее значение
            prefix: текст перед прогресс-баром
            suffix: текст после прогресс-баром
            length: длина прогресс-бара в символах
        """
        percent = 100 * (current / float(total))
        filled_length = int(length * current // total)
        bar = '█' * filled_length + '░' * (length - filled_length)
       
        # Форматируем проценты
        percent_display = f"{percent:.1f}%"
       
        # Форматируем счетчик
        counter_display = f"{current:,}/{total:,}"
       
        sys.stdout.write(f'\r{prefix} |{bar}| {percent_display} ({counter_display}) {suffix}')
        sys.stdout.flush()
       
        if current == total:
            print()  # Новая строка после завершения


def print_step(step_num, total_steps, message):
    """Отображает шаг процесса"""
    print(f"\n[ШАГ {step_num}/{total_steps}] {message}")
    print("-" * 60)


def _list_data_files(data_folder: str) -> list[str]:
    """Все выгрузки в папке таблицы: Excel и CSV."""
    files: list[str] = []
    for pattern in ("*.xlsx", "*.xls", "*.csv"):
        files.extend(glob(os.path.join(data_folder, pattern)))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    return sorted(files)


def _read_csv_all_strings(file_path, header=0, max_rows=None):
    """
    Читает CSV в строки (как Excel). Пробует кодировки utf-8-sig, utf-8, cp1251.
    Разделитель: авто (запятая или ;).
    """
    encodings = ("utf-8-sig", "utf-8", "cp1251", "latin-1")
    last_err: Exception | None = None
    read_kw = dict(
        header=0 if header == 0 else None,
        dtype=str,
        keep_default_na=False,
        nrows=max_rows,
        sep=None,
        engine="python",
    )
    for enc in encodings:
        try:
            try:
                df = pd.read_csv(file_path, encoding=enc, on_bad_lines="warn", **read_kw)
            except TypeError:
                df = pd.read_csv(file_path, encoding=enc, error_bad_lines=False, **read_kw)
            df = df.fillna("")
            if header == 0:
                cols, renamed = _unique_sqlite_column_names(df.columns)
                for old, new in renamed:
                    print(f"  [WARN] Дубль заголовка CSV переименован: {old!r} -> {new!r}")
                df.columns = cols
            return df.astype(str)
        except Exception as e:
            last_err = e
    raise last_err if last_err else OSError(f"Не удалось прочитать CSV: {file_path}")


def _read_data_file(file_path, header=0, max_rows=None):
    """Excel (.xlsx/.xls) или CSV — единый интерфейс для загрузчиков."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return _read_csv_all_strings(file_path, header=header, max_rows=max_rows)
    return _read_excel_all_strings(file_path, header=header, max_rows=max_rows)


def _read_excel_all_strings(file_path, header=0, max_rows=None):
    """
    Читает Excel через openpyxl, все ячейки в строки.
    Сначала пробует быстрый режим read_only=True (стриминг, меньше памяти).
    При ошибке или «битых» значениях — fallback на полную загрузку с патчем _cast_number/from_excel.
    max_rows: если задано, читать не более столько строк (для оценки размера).
    """
    from openpyxl import load_workbook


    def _to_rows_readonly(wb, max_rows_limit=None):
        ws = wb.active
        max_excel_row = 1_048_576
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_excel_row, values_only=True):
            try:
                rows.append([str(v) if v is not None else "" for v in row])
            except Exception:
                # на «битых» значениях в read_only — fallback не поможет по строке, бросаем
                raise
            if max_rows_limit is not None and len(rows) >= max_rows_limit:
                break
        return rows


    # Быстрый путь: read_only — не грузит весь файл в память, быстрее на больших файлах
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            rows = _to_rows_readonly(wb, max_rows)
        finally:
            wb.close()
    except Exception:
        # Fallback: полная загрузка с защитой от битых чисел/дат
        rows = _read_excel_all_strings_full_load(file_path, max_rows)
        if rows is None:
            raise


    if not rows:
        return pd.DataFrame()
    max_len = max(len(r) for r in rows)
    rows = [list(r) + [""] * (max_len - len(r)) for r in rows]
    if header == 0:
        columns, renamed = _unique_sqlite_column_names(rows[0])
        for old, new in renamed:
            print(f"  [WARN] Дубль заголовка Excel переименован: {old!r} -> {new!r}")
        df = pd.DataFrame(rows[1:], columns=columns)
    else:
        df = pd.DataFrame(rows)
    return df.astype(str)




def _read_excel_all_strings_full_load(file_path, max_rows=None):
    """Полная загрузка книги с патчем _cast_number/from_excel. Возвращает list of rows или None."""
    from openpyxl import load_workbook
    import openpyxl.worksheet._reader as _ox_reader
    import openpyxl.utils.datetime as _ox_dt
    _orig_cast = getattr(_ox_reader, "_cast_number", None)
    _orig_from_excel = getattr(_ox_dt, "from_excel", None)
    _reader_has_from_excel = hasattr(_ox_reader, "from_excel")


    def _safe_cast_number(value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return value


    def _safe_from_excel(value, *args, **kwargs):
        if not isinstance(value, (int, float)):
            return value
        if _orig_from_excel is None:
            return value
        try:
            return _orig_from_excel(value, *args, **kwargs)
        except (ValueError, TypeError, ZeroDivisionError):
            return value


    try:
        _ox_reader._cast_number = _safe_cast_number
        _ox_dt.from_excel = _safe_from_excel
        if _reader_has_from_excel:
            _ox_reader.from_excel = _safe_from_excel
        wb = load_workbook(filename=file_path, read_only=False, data_only=True)
        ws = wb.active
        max_excel_row = 1_048_576
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_excel_row, values_only=True):
            rows.append([str(v) if v is not None else "" for v in row])
            if max_rows is not None and len(rows) >= max_rows:
                break
        wb.close()
    finally:
        if _orig_cast is not None:
            _ox_reader._cast_number = _orig_cast
        if _orig_from_excel is not None:
            _ox_dt.from_excel = _orig_from_excel
        if _reader_has_from_excel and _orig_from_excel is not None:
            _ox_reader.from_excel = _orig_from_excel
    return rows




def merge_and_load_xlsx_files_fast(
    db_path=None,
    data_folder=None,
    target_table=None,
    skip_header_after_first=True,
    chunksize=100000,
    skip_final_dedup=False,
):
    """
    Быстрое объединение Excel файлов и загрузка в SQLite базу данных.


    Args:
        db_path: Путь к файлу базы данных SQLite
        data_folder: Папка с Excel файлами для объединения
        target_table: Название таблицы в БД
        skip_header_after_first: Если True, заголовки только из первого файла
        chunksize: Размер чанка для пакетной записи
        skip_final_dedup: Если True, не выполнять финальную дедупликацию по всей таблице (значительно быстрее на больших таблицах)
    """
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    if data_folder is None:
        data_folder = _resolve_data_path()
    else:
        data_folder = _resolve_data_path(data_folder)
    if target_table is None:
        target_table = DEFAULT_TABLE_NAME


    print("\n" + "=" * 80)
    print("ЗАГРУЗКА ТАБЛИЦЫ В БАЗУ ДАННЫХ")
    print("=" * 80)
    print(f"Папка с данными: {data_folder}")
    print(f"Таблица: {target_table}")
    print(f"База данных: {db_path}")
    print("=" * 80 + "\n")
   
    total_steps = 4
    current_step = 1
   
    # ШАГ 1: Проверка файлов
    print_step(current_step, total_steps, "Проверка файлов...")
    current_step += 1
   
    if not os.path.exists(data_folder):
        print(f"ОШИБКА: Папка '{data_folder}' не найдена!")
        return None
   
    data_files = _list_data_files(data_folder)

    if not data_files:
        print(f"ОШИБКА: Нет файлов .xlsx / .xls / .csv в папке '{data_folder}'")
        return None

    total_files = len(data_files)

    print(f"Найдено файлов: {total_files}")
    for i, f in enumerate(data_files[:10], 1):
        print(f"  {i:2}. {os.path.basename(f)}")
    if total_files > 10:
        print(f"  ... и еще {total_files - 10} файлов")


    # ШАГ 2: Подготовка базы данных
    print_step(current_step, total_steps, "Подготовка базы данных...")
    current_step += 1
   
    try:
        # Подключаемся к базе данных с оптимизациями
        print("Подключение к базе данных...")
        conn = connect_sqlite(db_path)
        cursor = conn.cursor()
       
        # Включаем оптимизации для скорости
        print("Включение оптимизаций...")
        conn.execute("PRAGMA journal_mode = OFF")
        conn.execute("PRAGMA synchronous = OFF")
        conn.execute("PRAGMA cache_size = -20000")
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("PRAGMA temp_store = MEMORY")
        print("Оптимизации включены")
       
        # Удаляем существующую таблицу
        print(f"Удаление старой таблицы '{target_table}'...")
        cursor.execute(f"DROP TABLE IF EXISTS '{target_table}'")
        conn.commit()
        print("Старая таблица удалена")
       
    except Exception as e:
        print(f"ОШИБКА при подготовке БД: {e}")
        return None
   
    # ШАГ 3: Чтение и загрузка файлов
    print_step(current_step, total_steps, f"Загрузка {total_files} файлов...")
    current_step += 1
   
    first_file_columns = None
    table_created = False
    total_rows_loaded = 0
    start_time = time.time()
   
    print("\nНачало загрузки файлов:")
    print("=" * 70)
   
    for file_idx, file_path in enumerate(data_files, 1):
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # в MB
       
        print(f"\nФайл {file_idx}/{total_files}: {file_name} ({file_size:.1f} MB)")
       
        try:
            # Определяем параметры чтения
            read_start = time.time()
           
            if file_idx == 1:  # Первый файл
                print(f"  Чтение с заголовками...")
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
                first_file_columns = list(df.columns)
                print(f"  Столбцов: {len(first_file_columns)}")
               
            else:  # Последующие файлы
                if skip_header_after_first and first_file_columns:
                    print(f"  Чтение без заголовков...")
                    df = _read_data_file(file_path, header=None)
                    if len(df) > 0:
                        rows_before = len(df)
                        df = df.iloc[1:].reset_index(drop=True)
                        rows_removed = rows_before - len(df)
                        if rows_removed > 0:
                            print(f"  Удалено заголовков: {rows_removed}")
                    if len(df.columns) == len(first_file_columns):
                        df.columns = first_file_columns
                    elif len(df.columns) > len(first_file_columns):
                        df.columns = first_file_columns + [
                            f'extra_col_{j}' for j in range(len(first_file_columns), len(df.columns))
                        ]
                    else:
                        df.columns = first_file_columns[:len(df.columns)]
                else:
                    print(f"  Чтение с заголовками...")
                    df = _normalize_df_columns(_read_data_file(file_path, header=0))
           
            df = _normalize_df_columns(df)
            read_time = time.time() - read_start
            print(f"  Прочитано: {len(df):,} строк за {read_time:.1f} сек")
           
            # Очистка от полных дублей: только строки, совпадающие по всем колонкам (не по одному полю)
            if len(df) > 0:
                before_dedup = len(df)
                df = df.drop_duplicates()  # без subset = по всем колонкам
                if len(df) < before_dedup:
                    print(f"  Удалено дубликатов: {before_dedup - len(df):,} (осталось {len(df):,})")
           
            # Создаем таблицу при первом файле
            if not table_created and len(df) > 0:
                print(f"  Создание таблицы в БД...")
                create_start = time.time()
               
                # Создаем таблицу
                sample_df = df.head(min(10000, len(df)))
                sample_df.to_sql(
                    target_table,
                    conn,
                    if_exists='fail',
                    index=False,
                    chunksize=chunksize
                )
                table_created = True
               
                create_time = time.time() - create_start
                print(f"  Таблица создана за {create_time:.1f} сек")
               
                # Если sample_df не весь df, записываем остаток
                if len(df) > len(sample_df):
                    remaining_df = df.iloc[len(sample_df):]
                    print(f"  Запись оставшихся {len(remaining_df):,} строк...")
                    remaining_df.to_sql(
                        target_table,
                        conn,
                        if_exists='append',
                        index=False,
                        chunksize=chunksize
                    )
            elif table_created and len(df) > 0:
                # Добавляем данные к существующей таблице
                print(f"  Добавление данных в таблицу...")
                write_start = time.time()
               
                df.to_sql(
                    target_table,
                    conn,
                    if_exists='append',
                    index=False,
                    chunksize=chunksize
                )
               
                write_time = time.time() - write_start
                print(f"  Записано за {write_time:.1f} сек")
           
            total_rows_loaded += len(df)
           
            # Показываем прогресс
            progress_percent = (file_idx / total_files) * 100
            ProgressBar.show(file_idx, total_files,
                           prefix=f"Прогресс:",
                           suffix=f"Файлов: {file_idx}/{total_files} | Строк: {total_rows_loaded:,}")
           
            del df
            gc.collect()
           
        except Exception as e:
            print(f"\n  ОШИБКА при обработке файла {file_name}: {str(e)}")
            print(f"  Пропускаем файл и продолжаем...")
            continue
   
    # Один коммит в конце
    conn.commit()
   
    if not skip_final_dedup:
        try:
            cursor.execute(f"SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?", (target_table,))
            if cursor.fetchone()[0]:
                before_dedup, after_dedup, removed = _dedup_table_in_db(conn, target_table)
                if removed > 0:
                    print(f"\n  Удалено дубликатов в таблице: {removed:,} (было {before_dedup:,}, стало {after_dedup:,})")
                conn.commit()
        except Exception as e:
            print(f"\n  [WARN] Очистка дубликатов не выполнена: {e}")
   
    # Восстанавливаем безопасные настройки
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
   
    # Проверяем результат (таблица может не существовать, если все файлы упали с ошибкой)
    try:
        cursor.execute(f"SELECT COUNT(*) FROM '{target_table}'")
        count = cursor.fetchone()[0]
    except sqlite3.OperationalError:
        count = 0
        print(f"\nВНИМАНИЕ: Таблица '{target_table}' не создана (все файлы завершились с ошибкой).")
   
    total_time = time.time() - start_time
   
    conn.close()
   
    # ШАГ 4: Завершение
    print_step(current_step, total_steps, "Завершение...")
   
    print("\n" + "=" * 80)
    print("ЗАГРУЗКА ЗАВЕРШЕНА")
    print("=" * 80)
   
    # Статистика
    print(f"\nСТАТИСТИКА:")
    print(f"- Таблица: {target_table}")
    print(f"- Файлов обработано: {total_files}")
    print(f"- Строк загружено: {count:,}")
    print(f"- Общее время: {total_time:.1f} секунд")
   
    if count > 0 and total_time > 0:
        rows_per_second = count / total_time
        print(f"- Скорость: {rows_per_second:,.0f} строк/сек")
   
    print(f"- Результат: УСПЕХ")
    print("\n" + "=" * 80)
   
    return {
        'table_name': target_table,
        'source_files': total_files,
        'db_rows': count,
        'total_time': total_time,
        'rows_per_second': count / total_time if total_time > 0 else 0
    }


def merge_and_load_xlsx_files_ultra_fast(
    db_path=None,
    data_folder=None,
    target_table=None,
    skip_header_after_first=True,
    batch_size=100000,
    skip_final_dedup=False,
):
    """
    УЛЬТРА-БЫСТРОЕ объединение Excel файлов и загрузка в SQLite.
    skip_final_dedup: если True, не выполнять финальную дедупликацию по всей таблице (быстрее).
    """
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    if data_folder is None:
        data_folder = _resolve_data_path()
    else:
        data_folder = _resolve_data_path(data_folder)
    if target_table is None:
        target_table = DEFAULT_TABLE_NAME


    print("\n" + "=" * 80)
    print("УЛЬТРА-БЫСТРАЯ ЗАГРУЗКА ТАБЛИЦЫ")
    print("=" * 80)
   
    try:
        if not os.path.exists(data_folder):
            raise FileNotFoundError(f"Папка '{data_folder}' не найдена!")
       
        data_files = _list_data_files(data_folder)

        if not data_files:
            print(f"ОШИБКА: Нет файлов .xlsx / .xls / .csv в папке '{data_folder}'")
            return None

        total_files = len(data_files)

        print(f"Найдено файлов: {total_files}")
       
        # Проверка целостности БД (повреждение ≠ database is locked)
        if os.path.exists(db_path):
            import datetime


            def _run_quick_check():
                c = connect_sqlite(db_path)
                try:
                    row = c.execute("PRAGMA quick_check(1)").fetchone()
                    val = (row[0] if row else "") or ""
                    if str(val).strip().lower() != "ok":
                        raise sqlite3.DatabaseError(f"quick_check: {val}")
                finally:
                    try:
                        c.close()
                    except Exception:
                        pass


            try:
                _run_quick_check()
            except sqlite3.OperationalError as e:
                if is_lock_error(e):
                    print("\nБД занята (database is locked) — ждём освобождения файла...")
                    ok_probe, err_probe = probe_db_writable(db_path)
                    if not ok_probe:
                        print("Не удалось получить доступ к БД для записи.")
                        print("Закройте программы с открытым файлом: DB Browser for SQLite, другой Python/загрузчик, Excel-плагины и т.п.")
                        if err_probe:
                            print(f"Детали: {err_probe}")
                        return None
                    try:
                        _run_quick_check()
                    except sqlite3.OperationalError as e2:
                        print(f"ОШИБКА: БД снова недоступна: {e2}")
                        return None
                else:
                    raise
            except sqlite3.DatabaseError as e:
                print(f"\nВНИМАНИЕ: База данных повреждена или не прошла проверку ({e}).")
                print("Если переименовать её в бэкап и создать новую — в новой БД будет только загружаемая таблица.")
                print("Остальные таблицы останутся только в файле-бэкапе.")
                try:
                    answer = input("Всё равно переименовать в бэкап и создать новую БД? (y/n): ").strip().lower()
                except Exception:
                    answer = "n"
                if answer != "y" and answer != "yes":
                    print("Загрузка отменена. Восстановите БД из бэкапа (db_mrt.db.corrupted_*) или исправьте файл вручную.")
                    return None
                backup_name = f"{db_path}.corrupted_{datetime.datetime.now():%Y%m%d_%H%M%S}"
                print(f"Файл переименован в: {backup_name}")
                print("Создаётся новая пустая БД.\n")
                os.rename(db_path, backup_name)
       
        # Подключаемся к БД (долгое ожидание при конкурирующих читателях)
        conn = connect_sqlite(db_path)
       
        # Максимальные оптимизации
        conn.execute("PRAGMA journal_mode = OFF")
        conn.execute("PRAGMA synchronous = OFF")
        conn.execute("PRAGMA cache_size = -50000")
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("PRAGMA temp_store = MEMORY")
        # EXCLUSIVE часто даёт database is locked, если другой процесс открыл БД (даже на чтение)
        conn.execute("PRAGMA locking_mode = NORMAL")
       
        cursor = conn.cursor()
       
        # Удаляем существующую таблицу
        cursor.execute(f"DROP TABLE IF EXISTS '{target_table}'")
       
        first_file_columns = None
        total_rows = 0
        start_time = time.time()
       
        print("\nНачало ультра-быстрой загрузки...")
        print("=" * 70)
       
        for file_idx, file_path in enumerate(data_files, 1):
            file_name = os.path.basename(file_path)
           
            print(f"\nФайл {file_idx}/{total_files}: {file_name}")
           
            # Чтение файла
            read_start = time.time()
           
            if file_idx == 1:
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
                first_file_columns = list(df.columns)
            else:
                if skip_header_after_first and first_file_columns:
                    df = _read_data_file(file_path, header=None)
                    if len(df) > 0:
                        df = df.iloc[1:].reset_index(drop=True)
                    if len(df.columns) == len(first_file_columns):
                        df.columns = first_file_columns
                    elif len(df.columns) > len(first_file_columns):
                        df.columns = first_file_columns + [
                            f'extra_col_{j}' for j in range(len(first_file_columns), len(df.columns))
                        ]
                    else:
                        df.columns = first_file_columns[:len(df.columns)]
                else:
                    df = _normalize_df_columns(_read_data_file(file_path, header=0))

            df = _normalize_df_columns(df)
            read_time = time.time() - read_start
            print(f"  Прочитано: {len(df):,} строк за {read_time:.1f} сек")
           
            # Очистка от полных дублей: только строки, совпадающие по всем колонкам (не по одному полю)
            if len(df) > 0:
                before_dedup = len(df)
                df = df.drop_duplicates()  # без subset = по всем колонкам
                if len(df) < before_dedup:
                    print(f"  Удалено дубликатов: {before_dedup - len(df):,} (осталось {len(df):,})")
           
            # Создаем таблицу при первом файле
            if file_idx == 1 and len(df) > 0:
                print(f"  Создание таблицы...")
                create_start = time.time()
               
                col_defs = []
                for col in df.columns:
                    col_defs.append(f'"{col}" TEXT')
               
                create_sql = f'CREATE TABLE "{target_table}" (\n'
                create_sql += ',\n'.join(col_defs)
                create_sql += '\n)'
               
                cursor.execute(create_sql)
                create_time = time.time() - create_start
                print(f"  Таблица создана за {create_time:.1f} сек")
           
            # Вставка данных батчами (многострочный INSERT — быстрее за счёт меньшего числа вызовов)
            if len(df) > 0:
                print(f"  Запись данных...")
                write_start = time.time()
                num_cols = len(df.columns)
                data_tuples = [tuple(x) for x in df.values]
                # Лимит SQLite на число параметров в запросе (обычно 999)
                SQLITE_MAX_VARS = 999
                rows_per_stmt = max(1, SQLITE_MAX_VARS // num_cols)
                one_row_ph = ','.join(['?' for _ in range(num_cols)])
                multi_ph = ','.join([f'({one_row_ph})' for _ in range(rows_per_stmt)])
                insert_sql_multi = f'INSERT INTO "{target_table}" VALUES {multi_ph}'
                insert_sql_single = f'INSERT INTO "{target_table}" VALUES ({one_row_ph})'
                for batch_start in range(0, len(data_tuples), batch_size):
                    batch = data_tuples[batch_start:batch_start + batch_size]
                    for i in range(0, len(batch), rows_per_stmt):
                        chunk = batch[i:i + rows_per_stmt]
                        if len(chunk) == rows_per_stmt:
                            flat = [v for row in chunk for v in row]
                            cursor.execute(insert_sql_multi, flat)
                        else:
                            for row in chunk:
                                cursor.execute(insert_sql_single, row)
                    current_progress = min(batch_start + batch_size, len(data_tuples))
                    percent = (current_progress / len(data_tuples)) * 100
                    print(f"    Записано: {current_progress:,}/{len(data_tuples):,} строк ({percent:.1f}%)", end='\r')
                print(f"    Записано: {len(data_tuples):,} строк")
                write_time = time.time() - write_start
                print(f"  Записано за {write_time:.1f} сек")
               
                total_rows += len(df)
               
                # Показываем общий прогресс
                ProgressBar.show(file_idx, total_files,
                               prefix=f"Общий прогресс:",
                               suffix=f"Файлов: {file_idx}/{total_files} | Строк: {total_rows:,}")
           
            # Освобождаем память
            del df
            gc.collect()
       
        # Финальная дедупликация по всей таблице (опционально)
        if not skip_final_dedup:
            try:
                before_dedup, after_dedup, removed = _dedup_table_in_db(conn, target_table)
                if removed > 0:
                    print(f"\n  Удалено дубликатов в таблице: {removed:,} (было {before_dedup:,}, стало {after_dedup:,})")
            except Exception as e:
                print(f"\n  [WARN] Очистка дубликатов не выполнена: {e}")
       
        # Завершаем транзакцию
        conn.commit()
       
        # Восстанавливаем безопасные настройки (после commit)
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA synchronous = NORMAL")
        conn.execute("PRAGMA locking_mode = NORMAL")
       
        # Проверяем результат
        cursor.execute(f"SELECT COUNT(*) FROM '{target_table}'")
        count = cursor.fetchone()[0]
       
        total_time = time.time() - start_time
        conn.close()
       
        print("\n" + "=" * 80)
        print("УЛЬТРА-БЫСТРАЯ ЗАГРУЗКА ЗАВЕРШЕНА")
        print("=" * 80)
       
        print(f"\nСТАТИСТИКА:")
        print(f"- Таблица: {target_table}")
        print(f"- Файлов: {total_files}")
        print(f"- Строк: {count:,}")
        print(f"- Время: {total_time:.1f} сек")
       
        if total_time > 0:
            speed = count / total_time
            print(f"- Скорость: {speed:,.0f} строк/сек")
            print(f"- Метод: Ультра-быстрый")
       
        print("\n" + "=" * 80)
       
        return {
            'table_name': target_table,
            'source_files': total_files,
            'db_rows': count,
            'total_time': total_time
        }
       
    except Exception as e:
        print(f"\nОШИБКА: {e}")
        import traceback
        traceback.print_exc()
        return None


# Функции для вызова из других модулей
def load_data_fast():
    """Быстрая загрузка данных"""
    return merge_and_load_xlsx_files_fast(
        db_path=DEFAULT_DB_PATH,
        data_folder=DEFAULT_DATA_FOLDER,
        target_table=DEFAULT_TABLE_NAME,
        skip_header_after_first=True,
        chunksize=100000
    )


def load_data_ultra_fast():
    """Ультра-быстрая загрузка данных"""
    return merge_and_load_xlsx_files_ultra_fast(
        db_path=DEFAULT_DB_PATH,
        data_folder=DEFAULT_DATA_FOLDER,
        target_table=DEFAULT_TABLE_NAME,
        skip_header_after_first=True,
        batch_size=50000
    )


# Для обратной совместимости
def merge_and_load_xlsx_files(*args, **kwargs):
    return merge_and_load_xlsx_files_fast(*args, **kwargs)


def load_data():
    return load_data_fast()


def load_data_to_custom_table(table_name):
    """Загружает данные в указанную таблицу"""
    print(f"\nЗагрузка данных в таблицу: {table_name}")
    result = merge_and_load_xlsx_files_fast(
        db_path=DEFAULT_DB_PATH,
        data_folder=DEFAULT_DATA_FOLDER,
        target_table=table_name,
        skip_header_after_first=True,
        chunksize=100000
    )
   
    if result is None:
        print(f"ОШИБКА: Не удалось загрузить данные в таблицу '{table_name}'")
    else:
        print(f"УСПЕХ: Данные загружены в таблицу '{result['table_name']}'")
   
    return result




def get_table_folders(base_folder=None):
    """
    Возвращает список подпапок в base_folder — каждая подпапка = имя таблицы.
    Игнорируются скрытые папки и служебные имена.
    """
    base_abs = _resolve_data_path(base_folder)
    if not os.path.isdir(base_abs):
        return []
    names = []
    for name in os.listdir(base_abs):
        path = os.path.join(base_abs, name)
        if os.path.isdir(path) and not name.startswith(".") and name != "__pycache__":
            names.append(name)
    return sorted(names)




def load_all_tables_from_db_folders(
    db_path=None,
    base_folder=None,
    method="fast",
    skip_final_dedup=False,
    only_tables=None,
):
    """
    Заливает в БД все таблицы по подпапкам в base_folder (по умолчанию db):
    каждая подпапка = имя таблицы, внутри — .xlsx / .xls / .csv (склеиваются, одна шапка).
   
    :param db_path: путь к db_mrt.db
    :param base_folder: базовая папка (по умолчанию db)
    :param method: "fast" или "ultra_fast"
    :param skip_final_dedup: не делать финальную дедупликацию по всей таблице (быстрее)
    :param only_tables: если задан список имён — загружать только эти таблицы; иначе все подпапки
    """
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    base_abs = _resolve_data_path(base_folder)

    if not os.path.isdir(base_abs):
        print(f"ОШИБКА: Папка не найдена: {base_abs}")
        return []
   
    tables = get_table_folders(base_folder) if only_tables is None else only_tables
    if only_tables is not None:
        existing = set(get_table_folders(base_folder))
        tables = [t for t in only_tables if t in existing]
   
    if not tables:
        print(f"В папке {base_abs} нет подпапок с именами таблиц.")
        return []
   
    print("\n" + "=" * 70)
    print("ЗАГРУЗКА ВСЕХ ТАБЛИЦ ИЗ ПОДПАПОК")
    print("=" * 70)
    print(f"Базовая папка: {base_abs}")
    print(f"БД: {db_path}")
    print(f"Таблиц к загрузке: {len(tables)}")
    print(f"Список: {', '.join(tables)}")
    print("=" * 70 + "\n")
   
    results = []
    for i, table_name in enumerate(tables, 1):
        data_folder = os.path.join(base_abs, table_name)
        if not os.path.isdir(data_folder):
            print(f"[{i}/{len(tables)}] Пропуск {table_name}: не папка")
            continue
        print(f"\n[{i}/{len(tables)}] Таблица: {table_name} (папка: {data_folder})")
        if method == "ultra_fast":
            r = merge_and_load_xlsx_files_ultra_fast(
                db_path=db_path,
                data_folder=data_folder,
                target_table=table_name,
                skip_header_after_first=True,
                batch_size=50000,
                skip_final_dedup=skip_final_dedup,
            )
        else:
            r = merge_and_load_xlsx_files_fast(
                db_path=db_path,
                data_folder=data_folder,
                target_table=table_name,
                skip_header_after_first=True,
                chunksize=100000,
                skip_final_dedup=skip_final_dedup,
            )
        if r:
            results.append(r)
   
    print("\n" + "=" * 70)
    print("ИТОГО ЗАГРУЖЕНО ТАБЛИЦ:", len(results), "из", len(tables))
    print("=" * 70)
    return results


if __name__ == "__main__":
    print("\n" + "=" * 80)
    print("СКРИПТ ЗАГРУЗКИ ТАБЛИЦ В БАЗУ ДАННЫХ")
    print("=" * 80)
    _db, _db_src = resolve_database_path(_PROJECT_ROOT)
    print(f"Корень проекта: {_PROJECT_ROOT}")
    print(f"БД ({_db_src}): {_db}")
    print(f"Папка выгрузок: {_resolve_data_path()}")

    tables_in_db = get_table_folders()
    print(f"\nВ папке '{_resolve_data_path()}' найдено подпапок (таблиц): {len(tables_in_db)}")
    if tables_in_db:
        print("   ", ", ".join(tables_in_db[:15]), "..." if len(tables_in_db) > 15 else "")
   
    print(f"\nВыберите режим:")
    print("1. Залить одну таблицу (быстрый метод) — папка из настроек DEFAULT_TABLE_NAME")
    print("2. Залить одну таблицу (ультра-быстрый метод)")
    print("3. Залить ВСЕ таблицы из подпапок db (каждая подпапка = одна таблица)")
   
    try:
        choice = input("\nВведите номер (1, 2 или 3): ").strip()
       
        if choice == "3":
            print("\nРежим: загрузка всех таблиц из подпапок db.")
            m = input("Метод: 1=быстрый, 2=ультра-быстрый [1]: ").strip() or "1"
            skip_dedup = input("Пропустить финальную дедупликацию по таблице (ускоряет загрузку)? (y/n) [n]: ").strip().lower() == "y"
            result = load_all_tables_from_db_folders(
                method="ultra_fast" if m == "2" else "fast",
                skip_final_dedup=skip_dedup,
            )
            if result:
                print(f"\nОперация завершена. Загружено таблиц: {len(result)}")
            else:
                print("\nНет таблиц для загрузки или ошибки.")
        elif choice == "2":
            print(f"\nЗапуск ультра-быстрой загрузки таблицы '{DEFAULT_TABLE_NAME}'...")
            result = load_data_ultra_fast()
            if result:
                print(f"\nОперация завершена успешно!")
            else:
                print(f"\nОперация завершена с ошибками!")
        elif choice == "1":
            print(f"\nЗапуск быстрой загрузки таблицы '{DEFAULT_TABLE_NAME}'...")
            result = load_data_fast()
            if result:
                print(f"\nОперация завершена успешно!")
            else:
                print(f"\nОперация завершена с ошибками!")
        else:
            print(f"\nНеверный выбор. Используем режим 1 (одна таблица, быстрый метод).")
            result = load_data_fast()
            if result:
                print(f"\nОперация завершена успешно!")
           
    except KeyboardInterrupt:
        print(f"\n\nОперация прервана пользователем")
    except Exception as e:
        print(f"\nКРИТИЧЕСКАЯ ОШИБКА: {e}")


